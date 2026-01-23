from pathlib import Path

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook

from exstruct.core.backends.com_backend import ComBackend, _parse_print_area_range
from exstruct.core.backends.openpyxl_backend import OpenpyxlBackend
from exstruct.core.ranges import parse_range_zero_based


def test_openpyxl_backend_extract_cells_switches_link_mode(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    calls: list[str] = []

    def fake_cells(file_path: Path) -> dict[str, list[object]]:
        calls.append("cells")
        return {}

    def fake_cells_links(file_path: Path) -> dict[str, list[object]]:
        calls.append("links")
        return {}

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.extract_sheet_cells",
        fake_cells,
    )
    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.extract_sheet_cells_with_links",
        fake_cells_links,
    )

    backend = OpenpyxlBackend(tmp_path / "book.xlsx")
    backend.extract_cells(include_links=False)
    backend.extract_cells(include_links=True)

    assert calls == ["cells", "links"]


def test_openpyxl_backend_detect_tables_handles_failure(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    def fake_detect(file_path: Path, sheet_name: str) -> list[str]:
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.detect_tables_openpyxl",
        fake_detect,
    )

    backend = OpenpyxlBackend(tmp_path / "book.xlsx")
    assert backend.detect_tables("Sheet1") == []


def test_openpyxl_backend_extract_colors_map_returns_none_on_failure(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    def fake_colors_map(
        file_path: Path,
        *,
        include_default_background: bool,
        ignore_colors: set[str] | None,
    ) -> object:
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.extract_sheet_colors_map",
        fake_colors_map,
    )

    backend = OpenpyxlBackend(tmp_path / "book.xlsx")
    assert (
        backend.extract_colors_map(include_default_background=False, ignore_colors=None)
        is None
    )


def test_openpyxl_backend_extract_formulas_map_returns_none_on_failure(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    def fake_formulas_map(file_path: Path) -> object:
        """
        Test helper that always raises a RuntimeError to simulate a failure when extracting a formulas map.
        
        Raises:
            RuntimeError: with message "boom".
        """
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.extract_sheet_formulas_map",
        fake_formulas_map,
    )

    backend = OpenpyxlBackend(tmp_path / "book.xlsx")
    assert backend.extract_formulas_map() is None


def test_com_backend_extract_colors_map_returns_none_on_failure(
    monkeypatch: MonkeyPatch,
) -> None:
    def fake_colors_map(
        workbook: object,
        *,
        include_default_background: bool,
        ignore_colors: set[str] | None,
    ) -> object:
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.com_backend.extract_sheet_colors_map_com",
        fake_colors_map,
    )

    class DummyWorkbook:
        pass

    backend = ComBackend(DummyWorkbook())
    assert (
        backend.extract_colors_map(include_default_background=False, ignore_colors=None)
        is None
    )


def test_com_backend_extract_formulas_map_returns_none_on_failure(
    monkeypatch: MonkeyPatch,
) -> None:
    def fake_formulas_map(workbook: object) -> object:
        """
        Test stub that simulates a failure by always raising a RuntimeError.
        
        Parameters:
            workbook (object): Workbook-like object (ignored); present to match the real function's signature.
        
        Raises:
            RuntimeError: Always raised with message "boom".
        """
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.com_backend.extract_sheet_formulas_map_com",
        fake_formulas_map,
    )

    class DummyWorkbook:
        pass

    backend = ComBackend(DummyWorkbook())
    assert backend.extract_formulas_map() is None


def test_com_backend_extract_print_areas_handles_sheet_error(
    monkeypatch: MonkeyPatch,
) -> None:
    class _FailingPageSetup:
        @property
        def PrintArea(self) -> str:
            raise RuntimeError("boom")

    class _FailingSheetApi:
        PageSetup = _FailingPageSetup()

    class _FailingSheet:
        name = "Sheet1"
        api = _FailingSheetApi()

    class _DummyWorkbook:
        sheets = [_FailingSheet()]

    backend = ComBackend(_DummyWorkbook())
    assert backend.extract_print_areas() == {}


def test_openpyxl_backend_extract_print_areas(tmp_path: Path) -> None:
    """
    Verifies that OpenpyxlBackend.extract_print_areas reads an openpyxl workbook's print area and returns the corresponding zero-based ranges keyed by sheet name.
    
    Creates an in-memory workbook with a single sheet named "Sheet1", sets its print area to "A1:B2", saves and loads it via OpenpyxlBackend, then asserts the sheet is present, has at least one area, and that the first area's r1 and c1 are 1 and 0 respectively.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.print_area = "A1:B2"
    file_path = tmp_path / "print_area.xlsx"
    wb.save(file_path)
    wb.close()

    backend = OpenpyxlBackend(file_path)
    areas = backend.extract_print_areas()
    assert "Sheet1" in areas
    assert areas["Sheet1"]
    assert areas["Sheet1"][0].r1 == 1
    assert areas["Sheet1"][0].c1 == 0


def test_openpyxl_backend_extract_print_areas_returns_empty_on_error(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    """
    Ensure OpenpyxlBackend.extract_print_areas returns an empty dict when the workbook loader raises an error.
    
    Verifies that the backend handles errors from the underlying workbook opening function by returning an empty mapping of print areas.
    """
    def _raise(*_args: object, **_kwargs: object) -> None:
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.openpyxl_workbook", _raise
    )
    backend = OpenpyxlBackend(tmp_path / "book.xlsx")
    assert backend.extract_print_areas() == {}


def test_parse_range_zero_based_parses_sheet_prefix() -> None:
    bounds = parse_range_zero_based("Sheet1!A1:B2")
    assert bounds is not None
    assert bounds.r1 == 0
    assert bounds.c1 == 0
    assert bounds.r2 == 1
    assert bounds.c2 == 1


def test_com_backend_extract_print_areas_success() -> None:
    class _PageSetup:
        PrintArea = "A1:B2,INVALID"

    class _SheetApi:
        PageSetup = _PageSetup()

    class _Sheet:
        name = "Sheet1"
        api = _SheetApi()

    class _DummyWorkbook:
        sheets = [_Sheet()]

    backend = ComBackend(_DummyWorkbook())
    areas = backend.extract_print_areas()
    assert "Sheet1" in areas
    assert areas["Sheet1"][0].r1 == 1
    assert areas["Sheet1"][0].c1 == 0
    assert areas["Sheet1"][0].r2 == 2
    assert areas["Sheet1"][0].c2 == 1


def test_com_backend_parse_print_area_range_invalid() -> None:
    assert _parse_print_area_range("INVALID") is None


class _Location:
    def __init__(self, row: int | None = None, col: int | None = None) -> None:
        """
        Initialize the location with row and column values.
        
        Parameters:
        	row (int | None): Row index or None.
        	col (int | None): Column index or None.
        """
        self.Row = row
        self.Column = col


class _BreakItem:
    def __init__(self, row: int | None = None, col: int | None = None) -> None:
        """
        Initialize the break item with an optional sheet location.
        
        Parameters:
            row (int | None): Row index (1-based) for the location, or None if unspecified.
            col (int | None): Column index (1-based) for the location, or None if unspecified.
        """
        self.Location = _Location(row=row, col=col)


class _Breaks:
    def __init__(self, items: list[_BreakItem]) -> None:
        """
        Initialize the Breaks collection from a list of break items.
        
        Parameters:
            items (list[_BreakItem]): Sequence of `_BreakItem` instances representing page break entries; ordering corresponds to 1-based access via `Item`.
        """
        self._items = items
        self.Count = len(items)

    def Item(self, index: int) -> _BreakItem:
        """
        Return the break item at the given 1-based position.
        
        Parameters:
            index (int): 1-based position of the break to retrieve.
        
        Returns:
            _BreakItem: The break item at the specified position.
        """
        return self._items[index - 1]


class _RangeRows:
    def __init__(self, count: int) -> None:
        """
        Initialize the breaks container with a specified item count.
        
        Parameters:
            count (int): Number of break items the container should report via its `Count` attribute.
        """
        self.Count = count


class _RangeCols:
    def __init__(self, count: int) -> None:
        """
        Initialize the breaks container with a specified item count.
        
        Parameters:
            count (int): Number of break items the container should report via its `Count` attribute.
        """
        self.Count = count


class _Range:
    Row = 1
    Column = 1
    Rows = _RangeRows(2)
    Columns = _RangeCols(2)


class _UsedRange:
    Address = "A1:B2"


class _PageSetup:
    PrintArea = "A1:B2"


class _SheetApi:
    def __init__(self) -> None:
        """
        Initialize a fake sheet API used by COM backend tests with default page and range state.
        
        Creates default attributes:
        - DisplayPageBreaks set to False.
        - PageSetup populated with a default PrintArea.
        - UsedRange populated with a default Address.
        - HPageBreaks containing one horizontal break at row 2.
        - VPageBreaks containing one vertical break at column 2.
        """
        self.DisplayPageBreaks = False
        self.PageSetup = _PageSetup()
        self.UsedRange = _UsedRange()
        self.HPageBreaks = _Breaks([_BreakItem(row=2)])
        self.VPageBreaks = _Breaks([_BreakItem(col=2)])

    def Range(self, _addr: str) -> _Range:
        """
        Create and return a Range wrapper for the given Excel-style address.
        
        Parameters:
            _addr (str): Excel-style address or range string (e.g., "A1", "A1:B2", or "Sheet1!A1:B2").
        
        Returns:
            _Range: An object representing the requested worksheet range.
        """
        return _Range()


class _Sheet:
    name = "Sheet1"

    def __init__(self) -> None:
        """
        Initialize a mock sheet and attach its API.
        
        Sets the `api` attribute to a new `_SheetApi` instance used by tests to simulate a sheet's COM-like API.
        """
        self.api = _SheetApi()


class _DummyWorkbook:
    def __init__(self) -> None:
        """
        Initialize a dummy workbook containing a single default sheet.
        
        The instance provides a `sheets` attribute set to a list with one `_Sheet` object.
        """
        self.sheets = [_Sheet()]


def test_com_backend_extract_auto_page_breaks_success() -> None:
    backend = ComBackend(_DummyWorkbook())
    areas = backend.extract_auto_page_breaks()
    assert "Sheet1" in areas
    assert areas["Sheet1"]


class _RestoreErrorSheetApi:
    def __init__(self) -> None:
        """
        Initialize a mock sheet API with default page, range, and break attributes.
        
        Creates:
        - `_display`: boolean flag for DisplayPageBreaks (defaults to False).
        - `PageSetup`: a default page setup object.
        - `UsedRange`: a default used-range object.
        - `HPageBreaks` and `VPageBreaks`: horizontal and vertical break collections, initialized empty.
        """
        self._display = False
        self.PageSetup = _PageSetup()
        self.UsedRange = _UsedRange()
        self.HPageBreaks = _Breaks([])
        self.VPageBreaks = _Breaks([])

    @property
    def DisplayPageBreaks(self) -> bool:
        """
        Get whether displaying page breaks is enabled on the sheet.
        
        Returns:
            `True` if page break display is enabled, `False` otherwise.
        """
        return self._display

    @DisplayPageBreaks.setter
    def DisplayPageBreaks(self, value: bool) -> None:
        """
        Set the sheet's DisplayPageBreaks flag.
        
        Parameters:
            value (bool): True to enable display of automatic page breaks. Passing False will trigger a restore failure.
        
        Raises:
            RuntimeError: If `value` is False (restore failed).
        """
        if value is False:
            raise RuntimeError("restore failed")
        self._display = value

    def Range(self, _addr: str) -> _Range:
        """
        Create and return a Range wrapper for the given Excel-style address.
        
        Parameters:
            _addr (str): Excel-style address or range string (e.g., "A1", "A1:B2", or "Sheet1!A1:B2").
        
        Returns:
            _Range: An object representing the requested worksheet range.
        """
        return _Range()


class _RestoreErrorSheet:
    name = "Sheet1"

    def __init__(self) -> None:
        """
        Create a sheet object whose underlying API simulates an error when restoring DisplayPageBreaks.
        
        This constructor assigns an instance of _RestoreErrorSheetApi to the `api` attribute so tests can exercise code paths that handle failures when restoring page-break state.
        """
        self.api = _RestoreErrorSheetApi()


class _RestoreErrorWorkbook:
    def __init__(self) -> None:
        """
        Create a mock workbook containing a single sheet that raises an error when restoring DisplayPageBreaks.
        
        The instance exposes a `sheets` attribute set to a list with one _RestoreErrorSheet(), which is used to simulate failures during page-break restoration in tests.
        """
        self.sheets = [_RestoreErrorSheet()]


def test_com_backend_extract_auto_page_breaks_restore_error() -> None:
    backend = ComBackend(_RestoreErrorWorkbook())
    areas = backend.extract_auto_page_breaks()
    assert "Sheet1" in areas