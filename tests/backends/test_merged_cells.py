from pathlib import Path
from typing import cast

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook
import pytest
import xlwings as xw

from exstruct.core.backends.com_backend import ComBackend
from exstruct.core.backends.openpyxl_backend import OpenpyxlBackend
from exstruct.models import MergedCell


def _make_merged_book(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    ws.merge_cells("A1:C1")
    wb.save(path)


def test_openpyxl_backend_extract_merged_cells(tmp_path: Path) -> None:
    path = tmp_path / "merged.xlsx"
    _make_merged_book(path)

    backend = OpenpyxlBackend(path)
    merged = backend.extract_merged_cells()
    assert merged["Sheet1"] == [MergedCell(r1=1, c1=0, r2=1, c2=2, v="Header")]


def test_openpyxl_backend_extract_merged_cells_handles_failure(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    def _boom(_path: Path) -> dict[str, list[MergedCell]]:
        raise RuntimeError("boom")

    monkeypatch.setattr(
        "exstruct.core.backends.openpyxl_backend.extract_sheet_merged_cells",
        _boom,
    )
    backend = OpenpyxlBackend(tmp_path / "missing.xlsx")
    assert backend.extract_merged_cells() == {}


def test_com_backend_extract_merged_cells_not_implemented() -> None:
    backend = ComBackend(cast(xw.Book, object()))
    with pytest.raises(NotImplementedError):
        backend.extract_merged_cells()
