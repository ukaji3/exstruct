from pathlib import Path

from openpyxl import Workbook

from exstruct import extract
from exstruct.core.backends.base import PrintAreaData
from exstruct.core.backends.openpyxl_backend import (
    OpenpyxlBackend,
    _append_print_areas,
    _extract_print_areas_from_defined_names,
    _extract_print_areas_from_sheet_props,
    _parse_print_area_range,
)


def _make_book_with_print_area(path: Path) -> None:
    """
    Create a simple Excel workbook with a single sheet named "Sheet1", set its print area to "A1:B2", write "x" to cell A1, save it to the given path, and close the file.

    Parameters:
        path (Path): Filesystem path where the workbook will be saved.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws.print_area = "A1:B2"
    wb.save(path)
    wb.close()


def test_light_mode_includes_print_areas_without_com(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_book_with_print_area(path)

    wb_data = extract(path, mode="light")
    areas = wb_data.sheets["Sheet1"].print_areas
    assert len(areas) == 1
    area = areas[0]
    assert (area.r1, area.c1, area.r2, area.c2) == (1, 0, 2, 1)


def test_openpyxl_backend_multiple_print_areas(tmp_path: Path) -> None:
    path = tmp_path / "multi_print_area.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws["D3"] = "y"
    ws.print_area = "A1:B2,D3:E4"
    wb.save(path)
    wb.close()

    backend = OpenpyxlBackend(path)
    areas = backend.extract_print_areas()

    assert "Sheet1" in areas
    ranges = [(a.r1, a.c1, a.r2, a.c2) for a in areas["Sheet1"]]
    assert ranges == [(1, 0, 2, 1), (3, 3, 4, 4)]


def test_extract_print_areas_from_defined_names_filters_unknown_sheets() -> None:
    """Ignore defined-name destinations for sheets that do not exist."""

    class _DefinedArea:
        destinations = [("Sheet1", "A1:B2"), ("Unknown", "C1:D2")]

    class _DefinedNames:
        def get(self, _name: str) -> _DefinedArea:
            """
            Create a default defined area object.

            Returns:
                _DefinedArea: A new, empty/default defined-area instance.
            """
            return _DefinedArea()

    class _DummyWorkbook:
        defined_names = _DefinedNames()
        sheetnames = ["Sheet1"]

    areas = _extract_print_areas_from_defined_names(_DummyWorkbook())
    assert "Sheet1" in areas
    assert "Unknown" not in areas


def test_extract_print_areas_from_defined_names_without_defined_names() -> None:
    """Return an empty mapping when defined_names is missing."""

    class _DummyWorkbook:
        defined_names = None

    assert _extract_print_areas_from_defined_names(_DummyWorkbook()) == {}


def test_extract_print_areas_from_sheet_props_skips_empty() -> None:
    """Skip sheet print areas when the property is empty."""

    class _SheetEmpty:
        title = "Sheet1"
        _print_area = None

    class _SheetWithArea:
        title = "Sheet2"
        _print_area = "A1:B2"

    class _DummyWorkbook:
        worksheets = [_SheetEmpty(), _SheetWithArea()]

    areas = _extract_print_areas_from_sheet_props(_DummyWorkbook())
    assert "Sheet2" in areas


def test_parse_print_area_range_invalid() -> None:
    """Return None for invalid range strings."""
    assert _parse_print_area_range("INVALID") is None


def test_append_print_areas_skips_invalid_ranges() -> None:
    """Append only valid print areas and skip invalid ranges."""
    areas: PrintAreaData = {}
    _append_print_areas(areas, "Sheet1", "A1:B2,INVALID")
    assert "Sheet1" in areas
    assert len(areas["Sheet1"]) == 1
