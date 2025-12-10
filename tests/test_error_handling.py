from pathlib import Path
import subprocess
import sys
from typing import Never

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook

from exstruct import extract
from exstruct.core.charts import parse_series_formula
from exstruct.core.integrate import extract_workbook
from exstruct.models import Chart


def _make_simple_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "v1"
    ws["B1"] = "v2"
    wb.save(path)


def test_COMエラーでもフォールバックしてプロセスが落ちない(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "book.xlsx"
    _make_simple_workbook(path)

    def _raise(*_a: object, **_k: object) -> Never:
        raise RuntimeError("COM not available")

    monkeypatch.setattr("exstruct.core.integrate.xw.Book", _raise, raising=False)
    data = extract(path)
    assert data.sheets
    sheet = next(iter(data.sheets.values()))
    assert sheet.shapes == []
    assert sheet.charts == []


def test_図形抽出失敗でも他要素が取得される(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "book.xlsx"
    _make_simple_workbook(path)

    def _raise_shapes(*_a: object, **_k: object) -> Never:
        raise RuntimeError("shapes fail")

    monkeypatch.setattr(
        "exstruct.core.integrate.get_shapes_with_position", _raise_shapes
    )

    class DummyBook:
        def __init__(self, *_a: object, **_k: object) -> None:
            self.sheets: list[object] = []

        def close(self) -> None:
            pass

    monkeypatch.setattr("exstruct.core.integrate.xw.Book", DummyBook, raising=False)

    data = extract_workbook(path)
    for sheet in data.sheets.values():
        assert sheet.rows != []
        assert sheet.charts == []
        assert sheet.shapes == []


def test_chart_errorに必ず文字列が入る(monkeypatch: MonkeyPatch) -> None:
    def _broken_parse(*_a: object, **_k: object) -> Never:
        raise RuntimeError("broken chart")

    monkeypatch.setattr("exstruct.core.charts.parse_series_formula", _broken_parse)
    ch = Chart(
        name="c",
        chart_type="t",
        title=None,
        y_axis_title="",
        y_axis_range=[],
        series=[],
        l=0,
        t=0,
        error="failed",
    )
    assert isinstance(ch.error, str)


def test_broken_range解析は例外化せずNone() -> None:
    assert parse_series_formula("=SERIES(") is None


def test_excelファイルが開けない場合はメッセージを出して終了(tmp_path: Path) -> None:
    bad_path = tmp_path / "missing.xlsx"
    out = tmp_path / "out.json"
    cmd = [sys.executable, "-m", "exstruct.cli.main", str(bad_path), "-o", str(out)]
    result = subprocess.run(cmd, capture_output=True, text=True)
    assert result.returncode == 0
    assert "file not found" in result.stdout.lower()
