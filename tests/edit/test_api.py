from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from exstruct.edit import (
    MakeRequest,
    PatchOp,
    PatchRequest,
    get_patch_op_schema,
    make_workbook,
    patch_workbook,
)
from exstruct.mcp.patch.models import PatchRequest as McpPatchModelRequest
from exstruct.mcp.patch_runner import PatchRequest as McpPatchRequest


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    workbook.save(path)
    workbook.close()


def test_patch_workbook_edits_without_path_policy(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    _create_workbook(source)

    result = patch_workbook(
        PatchRequest(
            xlsx_path=source,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            backend="openpyxl",
        )
    )

    assert result.error is None
    patched = Path(result.out_path)
    assert patched.exists()
    workbook = load_workbook(patched)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_make_workbook_creates_file_without_path_policy(tmp_path: Path) -> None:
    target = tmp_path / "new_book.xlsx"

    result = make_workbook(
        MakeRequest(
            out_path=target,
            ops=[
                PatchOp(op="add_sheet", sheet="Data"),
                PatchOp(op="set_value", sheet="Data", cell="A1", value="ok"),
            ],
            backend="openpyxl",
        )
    )

    assert result.error is None
    workbook = load_workbook(target)
    try:
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_edit_request_import_path_matches_mcp_compatibility_path() -> None:
    assert PatchRequest is McpPatchRequest
    assert PatchRequest is McpPatchModelRequest


def test_edit_op_schema_is_public() -> None:
    schema = get_patch_op_schema("create_chart")
    assert schema is not None
    assert schema.op == "create_chart"
