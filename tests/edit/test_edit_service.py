from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.edit import runtime as edit_runtime
from exstruct.edit.models import (
    FormulaIssue,
    MakeRequest,
    OpenpyxlEngineResult,
    PatchOp,
    PatchRequest,
)
import exstruct.edit.service as edit_service


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    workbook.save(path)
    workbook.close()


def test_edit_service_patch_prefers_com_when_available(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    calls: dict[str, bool] = {}

    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )

    def _fake_apply_xlwings_engine(
        input_path: Path,
        output_path: Path,
        ops: list[PatchOp],
        auto_formula: bool,
    ) -> list[object]:
        calls["com"] = True
        return []

    monkeypatch.setattr(
        edit_service, "apply_xlwings_engine", _fake_apply_xlwings_engine
    )
    result = edit_service.patch_workbook(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
            backend="auto",
        )
    )
    assert result.error is None
    assert result.engine == "com"
    assert calls["com"] is True


def test_edit_service_patch_auto_falls_back_to_openpyxl_on_com_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)

    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )

    def _raise_com_error(
        input_path: Path,
        output_path: Path,
        ops: list[PatchOp],
        auto_formula: bool,
    ) -> list[object]:
        raise RuntimeError("boom")

    def _fake_apply_openpyxl_engine(
        request: PatchRequest,
        input_path: Path,
        output_path: Path,
    ) -> OpenpyxlEngineResult:
        return OpenpyxlEngineResult()

    monkeypatch.setattr(edit_service, "apply_xlwings_engine", _raise_com_error)
    monkeypatch.setattr(
        edit_service, "apply_openpyxl_engine", _fake_apply_openpyxl_engine
    )
    result = edit_service.patch_workbook(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
            backend="auto",
        )
    )
    assert result.error is None
    assert result.engine == "openpyxl"
    assert any("falling back to openpyxl" in warning for warning in result.warnings)


def test_edit_service_make_applies_ops_without_mcp_policy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )
    out_path = tmp_path / "book.xlsx"
    result = edit_service.make_workbook(
        MakeRequest(
            out_path=out_path,
            ops=[
                PatchOp(op="add_sheet", sheet="Data"),
                PatchOp(op="set_value", sheet="Data", cell="A1", value="ok"),
            ],
            backend="openpyxl",
        )
    )

    assert result.error is None
    workbook = load_workbook(out_path)
    try:
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_edit_service_formula_health_check_uses_first_error_issue(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)

    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )

    def _fake_apply_openpyxl_engine(
        request: PatchRequest,
        input_path: Path,
        output_path: Path,
    ) -> OpenpyxlEngineResult:
        return OpenpyxlEngineResult(
            formula_issues=[
                FormulaIssue(
                    sheet="Sheet1",
                    cell="B1",
                    level="warning",
                    code="name_error",
                    message="warning-first",
                ),
                FormulaIssue(
                    sheet="Sheet1",
                    cell="A1",
                    level="error",
                    code="ref_error",
                    message="real-error",
                ),
            ]
        )

    monkeypatch.setattr(
        edit_service, "apply_openpyxl_engine", _fake_apply_openpyxl_engine
    )

    result = edit_service.patch_workbook(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(op="set_formula", sheet="Sheet1", cell="B1", formula="=1+1"),
                PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!"),
            ],
            preflight_formula_check=True,
            backend="openpyxl",
        )
    )

    assert result.error is not None
    assert result.error.op_index == 1
    assert result.error.cell == "A1"


def test_edit_service_dry_run_rename_cleans_reserved_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("existing", encoding="utf-8")

    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )

    def _fake_apply_openpyxl_engine(
        request: PatchRequest,
        input_path: Path,
        output_path: Path,
    ) -> OpenpyxlEngineResult:
        assert output_path.name == "book_patched_1.xlsx"
        return OpenpyxlEngineResult(
            patch_diff=[],
            inverse_ops=[],
            formula_issues=[],
            op_warnings=[],
        )

    monkeypatch.setattr(
        edit_service, "apply_openpyxl_engine", _fake_apply_openpyxl_engine
    )

    result = edit_service.patch_workbook(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
            dry_run=True,
            backend="openpyxl",
        )
    )

    assert result.error is None
    assert result.out_path.endswith("book_patched_1.xlsx")
    assert not Path(result.out_path).exists()


def test_edit_service_preflight_rename_cleans_reserved_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("existing", encoding="utf-8")

    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )

    def _fake_apply_openpyxl_engine(
        request: PatchRequest,
        input_path: Path,
        output_path: Path,
    ) -> OpenpyxlEngineResult:
        assert output_path.name == "book_patched_1.xlsx"
        return OpenpyxlEngineResult(
            patch_diff=[],
            inverse_ops=[],
            formula_issues=[
                FormulaIssue(
                    sheet="Sheet1",
                    cell="A1",
                    level="error",
                    code="ref_error",
                    message="real-error",
                )
            ],
            op_warnings=[],
        )

    monkeypatch.setattr(
        edit_service, "apply_openpyxl_engine", _fake_apply_openpyxl_engine
    )

    result = edit_service.patch_workbook(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!")
            ],
            on_conflict="rename",
            preflight_formula_check=True,
            backend="openpyxl",
        )
    )

    assert result.error is not None
    assert result.out_path.endswith("book_patched_1.xlsx")
    assert not Path(result.out_path).exists()


def test_edit_service_openpyxl_reraise_cleans_reserved_output(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        edit_runtime,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )

    for expected_exc in (ValueError, FileNotFoundError, OSError):
        input_path = tmp_path / f"{expected_exc.__name__}.xlsx"
        _create_workbook(input_path)
        default_out = tmp_path / f"{expected_exc.__name__}_patched.xlsx"
        default_out.write_text("existing", encoding="utf-8")

        def _raise_openpyxl_error(
            request: PatchRequest,
            input_path: Path,
            output_path: Path,
            exc_type: type[Exception] = expected_exc,
        ) -> OpenpyxlEngineResult:
            assert output_path.name == f"{exc_type.__name__}_patched_1.xlsx"
            raise exc_type("boom")

        monkeypatch.setattr(
            edit_service, "apply_openpyxl_engine", _raise_openpyxl_error
        )

        with pytest.raises(expected_exc):
            edit_service.patch_workbook(
                PatchRequest(
                    xlsx_path=input_path,
                    ops=[
                        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")
                    ],
                    on_conflict="rename",
                    backend="openpyxl",
                )
            )
        assert not (tmp_path / f"{expected_exc.__name__}_patched_1.xlsx").exists()
