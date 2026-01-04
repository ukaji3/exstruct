from pathlib import Path

from openpyxl import Workbook

from exstruct.core.cells import extract_sheet_merged_cells


def _make_merged_book(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Title"
    ws.merge_cells("A1:B2")
    ws.merge_cells("D4:E4")
    wb.save(path)


def test_extract_sheet_merged_cells_basic(tmp_path: Path) -> None:
    path = tmp_path / "merged.xlsx"
    _make_merged_book(path)

    merged = extract_sheet_merged_cells(path)
    ranges = merged["Sheet1"]
    assert len(ranges) == 2
    tuples = {(r.r1, r.c1, r.r2, r.c2, r.v) for r in ranges}
    assert (1, 0, 2, 1, "Title") in tuples
    assert (4, 3, 4, 4, "") in tuples


def test_extract_sheet_merged_cells_empty(tmp_path: Path) -> None:
    path = tmp_path / "plain.xlsx"
    wb = Workbook()
    wb.save(path)

    merged = extract_sheet_merged_cells(path)
    assert merged["Sheet"] == []
