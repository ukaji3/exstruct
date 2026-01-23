from pathlib import Path

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from exstruct.core import cells
from exstruct.core.cells import (
    _coerce_numeric_preserve_format,
    _normalize_formula_from_com,
    _normalize_formula_value,
    detect_tables_openpyxl,
    extract_sheet_formulas_map,
    extract_sheet_formulas_map_com,
)


def test_coerce_numeric_preserve_format() -> None:
    assert _coerce_numeric_preserve_format("42") == 42
    assert _coerce_numeric_preserve_format("-3.14") == -3.14
    # non-numeric stays string
    assert _coerce_numeric_preserve_format("3.14e2") == "3.14e2"
    assert _coerce_numeric_preserve_format("abc") == "abc"


def test_detect_tables_openpyxl_detects_tables(tmp_path: Path) -> None:
    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    tab = Table(displayName="Table1", ref="A1:B2")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    wb.save(path)
    wb.close()

    tables = detect_tables_openpyxl(path, "Sheet1")
    assert "A1:B2" in tables


def test_detect_tables_openpyxl_respects_table_params(
    tmp_path: Path, monkeypatch: MonkeyPatch
) -> None:
    # Ensure detection still runs after modifying global thresholds
    path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    tab = Table(displayName="Table1", ref="A1:B2")
    ws.add_table(tab)
    wb.save(path)
    wb.close()

    # Force density/coverage to high thresholds to ensure Table objects are still returned
    monkeypatch.setattr(
        cells,
        "_DETECTION_CONFIG",
        {
            "table_score_threshold": 0.99,
            "density_min": 0.99,
            "coverage_min": 0.99,
            "min_nonempty_cells": 1,
        },
    )
    tables = detect_tables_openpyxl(path, "Sheet1")
    assert "A1:B2" in tables


def test_normalize_formula_value_prefers_array_text() -> None:
    """
    Verify that _normalize_formula_value prefers an array-like object's text and treats an empty string as no formula.

    Asserts that an object with a `text` attribute is converted to a formula string prefixed with '=' (e.g., "=SUM(A1:A3)"), and that an empty string is normalized to None.
    """

    class _ArrayFormulaLike:
        text = "SUM(A1:A3)"

    assert _normalize_formula_value(_ArrayFormulaLike()) == "=SUM(A1:A3)"
    assert _normalize_formula_value("") is None


def test_extract_sheet_formulas_map_collects_formulas(tmp_path: Path) -> None:
    path = tmp_path / "formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = "=SUM(A1:A2)"
    wb.save(path)
    wb.close()

    result = extract_sheet_formulas_map(path)
    sheet = result.get_sheet("Sheet1")
    assert sheet is not None
    assert sheet.formulas_map == {"=SUM(A1:A2)": [(1, 1)]}


def test_normalize_formula_from_com() -> None:
    assert _normalize_formula_from_com("=A1") == "=A1"
    assert _normalize_formula_from_com("A1") is None
    assert _normalize_formula_from_com("") is None
    assert _normalize_formula_from_com(None) is None


def test_extract_sheet_formulas_map_com_empty_range() -> None:
    class _DummyLastCell:
        row = 0
        column = 0

    class _DummyUsedRange:
        row = 1
        column = 1
        last_cell = _DummyLastCell()

    class _DummySheet:
        name = "Sheet1"
        used_range = _DummyUsedRange()

    class _DummyWorkbook:
        sheets = [_DummySheet()]

    result = extract_sheet_formulas_map_com(_DummyWorkbook())
    sheet = result.get_sheet("Sheet1")
    assert sheet is not None
    assert sheet.formulas_map == {}


def test_extract_sheet_formulas_map_com_collects_formulas() -> None:
    class _DummyLastCell:
        row = 2
        column = 2

    class _DummyUsedRange:
        row = 1
        column = 1
        last_cell = _DummyLastCell()

    class _DummyRange:
        formula = [["=A1", "B1"], ["=SUM(A1)", ""]]

    class _DummySheet:
        name = "Sheet1"
        used_range = _DummyUsedRange()

        def range(self, _start: object, _end: object) -> _DummyRange:
            """
            Return a new _DummyRange representing a requested cell range.

            Parameters:
                _start (object): Start coordinate or cell reference for the range request (ignored by this dummy implementation).
                _end (object): End coordinate or cell reference for the range request (ignored by this dummy implementation).

            Returns:
                _DummyRange: A fresh _DummyRange instance corresponding to the requested range.
            """
            return _DummyRange()

    class _DummyWorkbook:
        sheets = [_DummySheet()]

    result = extract_sheet_formulas_map_com(_DummyWorkbook())
    sheet = result.get_sheet("Sheet1")
    assert sheet is not None
    assert sheet.formulas_map == {
        "=A1": [(1, 0)],
        "=SUM(A1)": [(2, 0)],
    }
