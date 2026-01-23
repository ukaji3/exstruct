from pathlib import Path
from typing import Never

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import pytest

from exstruct.core.cells import (
    TableScanLimits,
    detect_tables_openpyxl,
    extract_sheet_cells,
    load_border_maps_xlsx,
)
from exstruct.core.integrate import extract_workbook


def _make_workbook_with_table(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append(["v1", "v2"])
    ws.append(["v3", "v4"])
    tbl = Table(displayName="Tbl", ref="A1:B3")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    wb.save(path)
    wb.close()


def test_数値セルは型保持で抽出される(tmp_path: Path) -> None:
    path = tmp_path / "numbers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["123", "1.50", "text"])
    wb.save(path)
    wb.close()

    data = extract_sheet_cells(path)
    row = data["Sheet1"][0]
    assert row.c["0"] == 123
    assert isinstance(row.c["0"], int)
    assert row.c["1"] == pytest.approx(1.5)
    assert isinstance(row.c["1"], float)
    assert row.c["2"] == "text"


def test_openpyxlで正式テーブルを検出できる(tmp_path: Path) -> None:
    path = tmp_path / "table.xlsx"
    _make_workbook_with_table(path)
    tables = detect_tables_openpyxl(path, "Sheet1")
    assert "A1:B3" in tables


def test_openpyxl_border_scan_defers_column_shrink(tmp_path: Path) -> None:
    path = tmp_path / "border.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws.cell(row=1, column=1, value="x").border = border
    ws.cell(row=50, column=120, value="y").border = border
    wb.save(path)
    wb.close()

    limits = TableScanLimits(
        max_rows=200,
        max_cols=200,
        empty_row_run=200,
        empty_col_run=1,
        min_rows_before_col_shrink=60,
    )
    has_border, *_ = load_border_maps_xlsx(path, "Sheet1", scan_limits=limits)
    assert bool(has_border[50, 120]) is True


def test_excelなし環境ではセルとテーブルのみ返す(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "fallback.xlsx"
    _make_workbook_with_table(path)

    def _raise(*_args: object, **_kwargs: object) -> Never:
        raise RuntimeError("no COM")

    monkeypatch.setattr("exstruct.core.integrate._open_workbook", _raise, raising=False)

    wb_data = extract_workbook(path)
    assert wb_data.book_name == path.name
    sheet = wb_data.sheets["Sheet1"]
    assert sheet.shapes == []
    assert sheet.charts == []
    assert "A1:B3" in sheet.table_candidates
