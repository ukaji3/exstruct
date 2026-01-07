from pathlib import Path

from openpyxl import Workbook

from exstruct import extract
from exstruct.core.cells import extract_sheet_cells_with_links
from exstruct.models import CellRow


def _make_link_book(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    c = ws["A1"]
    c.value = "link"
    c.hyperlink = "http://example.com"
    ws["B1"] = "no-link"
    wb.save(path)
    wb.close()


def test_extract_sheet_cells_with_links_returns_links(tmp_path: Path) -> None:
    path = tmp_path / "links.xlsx"
    _make_link_book(path)
    data = extract_sheet_cells_with_links(path)
    row = data["Sheet1"][0]
    assert isinstance(row, CellRow)
    assert row.links == {"0": "http://example.com"}


def test_extract_verbose_includes_links(tmp_path: Path) -> None:
    path = tmp_path / "links.xlsx"
    _make_link_book(path)
    wb = extract(path, mode="verbose")
    row = wb.sheets["Sheet1"].rows[0]
    assert row.links == {"0": "http://example.com"}


def test_extract_standard_excludes_links_by_default(tmp_path: Path) -> None:
    path = tmp_path / "links.xlsx"
    _make_link_book(path)
    wb = extract(path, mode="standard")
    row = wb.sheets["Sheet1"].rows[0]
    assert row.links is None
