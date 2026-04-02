from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest
import xlwings as xw

from exstruct.cli.availability import ComAvailability
import exstruct.edit.internal as edit_internal
import exstruct.edit.runtime as edit_runtime
import exstruct.edit.service as edit_service
from exstruct.mcp import patch_runner
from exstruct.mcp.io import PathPolicy
from exstruct.mcp.patch_runner import MakeRequest, PatchOp, run_make


def _disable_com(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )


def _write_workbook(path: Path, value: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = value
    workbook.save(path)
    workbook.close()


def test_run_make_creates_xlsx_with_sheet1(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(out_path=out_path, ops=[]),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert result.engine == "openpyxl"
    workbook = load_workbook(result.out_path)
    try:
        assert "Sheet1" in workbook.sheetnames
    finally:
        workbook.close()


def test_run_make_applies_ops(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[
                PatchOp(op="add_sheet", sheet="Data"),
                PatchOp(op="set_value", sheet="Data", cell="A1", value="ok"),
            ],
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_make_preserves_patch_runner_get_com_availability_override(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    out_path = tmp_path / "book.xlsx"
    seen: dict[str, object] = {}

    def _fake_get_com_availability() -> ComAvailability:
        return ComAvailability(available=False, reason="patched-by-patch-runner")

    def _fake_make_workbook(request: MakeRequest) -> patch_runner.PatchResult:
        seen["availability"] = edit_runtime.get_com_availability()
        return patch_runner.PatchResult(
            out_path=str(out_path),
            patch_diff=[],
            warnings=[],
            engine="openpyxl",
        )

    monkeypatch.setattr(
        patch_runner, "get_com_availability", _fake_get_com_availability
    )
    monkeypatch.setattr(edit_service, "make_workbook", _fake_make_workbook)

    result = run_make(
        MakeRequest(out_path=out_path, ops=[]),
        policy=PathPolicy(root=tmp_path),
    )

    availability = seen["availability"]
    assert isinstance(availability, ComAvailability)
    assert availability.available is False
    assert availability.reason == "patched-by-patch-runner"
    assert result.engine == "openpyxl"


def test_run_make_syncs_patch_runner_override_into_edit_internal_xls_path(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    out_path = tmp_path / "book.xls"

    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="patched-by-patch-runner"),
    )
    monkeypatch.setattr(
        edit_internal,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )

    class _SentinelApp:
        def __init__(self, *args: object, **kwargs: object) -> None:
            raise AssertionError("COM seed creation should not be attempted")

    monkeypatch.setattr(xw, "App", _SentinelApp)

    with pytest.raises(ValueError, match=r"\.xls editing requires Windows Excel COM"):
        run_make(
            MakeRequest(out_path=out_path, ops=[]),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_make_uses_top_level_sheet_as_initial_sheet_when_no_matching_add_sheet(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(
            out_path=out_path,
            sheet="Data",
            ops=[PatchOp(op="set_value", sheet="Data", cell="A1", value="ok")],
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        assert "Data" in workbook.sheetnames
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_make_keeps_sheet1_when_matching_add_sheet_exists(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(
            out_path=out_path,
            sheet="Data",
            ops=[
                PatchOp(op="add_sheet", sheet="Data"),
                PatchOp(op="set_value", sheet="Data", cell="A1", value="ok"),
            ],
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        assert "Sheet1" in workbook.sheetnames
        assert "Data" in workbook.sheetnames
        assert workbook["Data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_make_keeps_sheet1_when_add_sheet_differs_only_by_case(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    result = run_make(
        MakeRequest(
            out_path=out_path,
            sheet="Data",
            ops=[
                PatchOp(op="add_sheet", sheet="data"),
                PatchOp(op="set_value", sheet="data", cell="A1", value="ok"),
            ],
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        assert "Sheet1" in workbook.sheetnames
        assert "data" in workbook.sheetnames
        assert workbook["data"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_make_conflict_overwrite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="overwrite",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_run_make_conflict_skip(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="skip",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.patch_diff == []
    workbook = load_workbook(out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "old"
    finally:
        workbook.close()


def test_run_make_conflict_rename(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    out_path = tmp_path / "book.xlsx"
    _write_workbook(out_path, "old")
    result = run_make(
        MakeRequest(
            out_path=out_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert Path(result.out_path) != out_path
    assert Path(result.out_path).exists()


def test_run_make_rejects_path_outside_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    root = tmp_path / "root"
    root.mkdir()
    with pytest.raises(ValueError):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xlsx"),
            policy=PathPolicy(root=root),
        )


def test_run_make_resolves_relative_out_path_from_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    root = tmp_path / "root"
    root.mkdir()
    elsewhere = tmp_path / "elsewhere"
    elsewhere.mkdir()
    monkeypatch.chdir(elsewhere)
    result = run_make(
        MakeRequest(out_path=Path("outputs/book.xlsx")),
        policy=PathPolicy(root=root),
    )
    assert result.error is None
    assert Path(result.out_path) == (root / "outputs" / "book.xlsx").resolve()


def test_run_make_rejects_xls_when_com_unavailable(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    with pytest.raises(ValueError, match=r"requires Windows Excel COM"):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls"),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_make_rejects_xls_with_openpyxl_backend(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match=r"backend='openpyxl' cannot edit \.xls files"):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls", backend="openpyxl"),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_make_rejects_xls_with_dry_run_flags(tmp_path: Path) -> None:
    with pytest.raises(
        ValueError,
        match=r"\.xls creation does not support dry_run, return_inverse_ops, or preflight_formula_check",
    ):
        run_make(
            MakeRequest(out_path=tmp_path / "book.xls", dry_run=True),
            policy=PathPolicy(root=tmp_path),
        )
