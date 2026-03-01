from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from pydantic import ValidationError
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.mcp import patch_runner
from exstruct.mcp.io import PathPolicy
from exstruct.mcp.patch_runner import PatchOp, PatchRequest, run_patch


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    sheet["B1"] = 1
    workbook.save(path)
    workbook.close()


def _seed_table_source(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Sheet1"]
        sheet["A1"] = "Name"
        sheet["B1"] = "Amount"
        sheet["A2"] = "A"
        sheet["B2"] = 100
        sheet["A3"] = "B"
        sheet["B3"] = 200
        workbook.save(path)
    finally:
        workbook.close()


def _disable_com(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )


def test_run_patch_set_value_and_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new"),
        PatchOp(op="set_formula", sheet="Sheet1", cell="B1", formula="=SUM(1,1)"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert Path(result.out_path).exists()
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].value == "new"
        formula_value = sheet["B1"].value
        if isinstance(formula_value, str) and not formula_value.startswith("="):
            formula_value = f"={formula_value}"
        assert formula_value == "=SUM(1,1)"
    finally:
        workbook.close()
    assert len(result.patch_diff) == 2
    assert result.patch_diff[0].after is not None
    assert result.engine == "openpyxl"


def test_run_patch_backend_auto_uses_openpyxl_when_com_unavailable(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
            on_conflict="rename",
            backend="auto",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert result.engine == "openpyxl"


def test_run_patch_backend_com_requires_com_available(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    with pytest.raises(ValueError, match=r"backend='com' requires"):
        run_patch(
            PatchRequest(
                xlsx_path=input_path,
                ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
                on_conflict="rename",
                backend="com",
            ),
            policy=PathPolicy(root=tmp_path),
        )


def test_run_patch_backend_openpyxl_rejects_xls(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )
    input_path = tmp_path / "book.xls"
    input_path.write_text("dummy", encoding="utf-8")
    with pytest.raises(ValueError, match=r"backend='openpyxl' cannot edit \.xls"):
        run_patch(
            PatchRequest(
                xlsx_path=input_path,
                ops=[PatchOp(op="add_sheet", sheet="Sheet2")],
                on_conflict="rename",
                backend="openpyxl",
            ),
            policy=PathPolicy(root=tmp_path),
        )


def test_patch_request_backend_com_rejects_dry_run() -> None:
    with pytest.raises(ValidationError, match=r"backend='com' does not support"):
        PatchRequest(
            xlsx_path=Path("book.xlsx"),
            ops=[PatchOp(op="add_sheet", sheet="S2")],
            dry_run=True,
            backend="com",
        )


def test_patch_request_backend_com_rejects_restore_design_snapshot() -> None:
    with pytest.raises(
        ValidationError,
        match=r"backend='com' does not support restore_design_snapshot operation",
    ):
        PatchRequest(
            xlsx_path=Path("book.xlsx"),
            ops=[
                PatchOp(
                    op="restore_design_snapshot",
                    sheet="Sheet1",
                    design_snapshot=patch_runner.DesignSnapshot(),
                )
            ],
            backend="com",
        )


def test_patch_request_backend_openpyxl_rejects_create_chart() -> None:
    with pytest.raises(
        ValidationError,
        match=r"create_chart is supported only on COM backend",
    ):
        PatchRequest(
            xlsx_path=Path("book.xlsx"),
            ops=[
                PatchOp(
                    op="create_chart",
                    sheet="Sheet1",
                    chart_type="line",
                    data_range="A1:B3",
                    anchor_cell="D2",
                )
            ],
            backend="openpyxl",
        )


def test_patch_request_backend_openpyxl_rejects_mixed_chart_and_table() -> None:
    with pytest.raises(
        ValidationError,
        match=r"create_chart is supported only on COM backend",
    ):
        PatchRequest(
            xlsx_path=Path("book.xlsx"),
            ops=[
                PatchOp(
                    op="create_chart",
                    sheet="Sheet1",
                    chart_type="line",
                    data_range="A1:B3",
                    anchor_cell="D2",
                ),
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="A1:B3",
                    style="TableStyleMedium2",
                ),
            ],
            backend="openpyxl",
        )


def test_run_patch_add_sheet_and_set_value(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="add_sheet", sheet="NewSheet"),
        PatchOp(op="set_value", sheet="NewSheet", cell="A1", value="ok"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    workbook = load_workbook(result.out_path)
    try:
        assert "NewSheet" in workbook.sheetnames
        assert workbook["NewSheet"]["A1"].value == "ok"
    finally:
        workbook.close()


def test_run_patch_add_sheet_rejects_duplicate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="add_sheet", sheet="Sheet1")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.error.op_index == 0


def test_patch_op_allows_value_starting_with_equal() -> None:
    op = PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")
    assert op.value == "=SUM(1,1)"


def test_run_patch_rejects_formula_without_equal() -> None:
    with pytest.raises(ValidationError):
        PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="SUM(1,1)")


def test_run_patch_set_value_with_equal_requires_auto_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert "rejects values starting with" in result.error.message


def test_run_patch_set_value_with_equal_auto_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="=SUM(1,1)")]
    request = PatchRequest(
        xlsx_path=input_path, ops=ops, on_conflict="rename", auto_formula=True
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    workbook = load_workbook(result.out_path)
    try:
        formula_value = workbook["Sheet1"]["A1"].value
        if isinstance(formula_value, str) and not formula_value.startswith("="):
            formula_value = f"={formula_value}"
        assert formula_value == "=SUM(1,1)"
    finally:
        workbook.close()


def test_run_patch_rejects_path_outside_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    root = tmp_path / "root"
    root.mkdir()
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    with pytest.raises(ValueError):
        run_patch(request, policy=PathPolicy(root=root))


def test_run_patch_conflict_rename(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path != str(default_out)
    assert Path(result.out_path).exists()


def test_run_patch_conflict_skip(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="skip")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path == str(default_out)
    assert result.patch_diff == []
    assert any("skipping" in warning for warning in result.warnings)


def test_run_patch_conflict_skip_dry_run_still_simulates(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    default_out.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="skip",
        dry_run=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.patch_diff) == 1
    assert any("ignores on_conflict=skip" in warning for warning in result.warnings)
    assert not any("may drop shapes/charts" in warning for warning in result.warnings)
    assert default_out.read_text(encoding="utf-8") == "dummy"


def test_run_patch_conflict_overwrite(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    default_out = tmp_path / "book_patched.xlsx"
    _create_workbook(default_out)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="overwrite")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.out_path == str(default_out)
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_run_patch_default_output_name_does_not_chain_patched_suffix(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book_patched.xlsx"
    _create_workbook(input_path)
    request = PatchRequest(
        xlsx_path=input_path,
        ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert result.out_path == str(input_path)
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Sheet1"]["A1"].value == "new"
    finally:
        workbook.close()


def test_run_patch_atomicity(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x"),
        PatchOp(op="set_value", sheet="Missing", cell="A1", value="y"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    output_path = tmp_path / "book_patched.xlsx"
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert not output_path.exists()


def test_run_patch_creates_out_dir(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    out_dir = tmp_path / "nested" / "output"
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        out_dir=out_dir,
        on_conflict="rename",
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    out_path = Path(result.out_path)
    assert out_path.exists()
    assert out_path.parent == out_dir


def test_run_patch_xls_requires_com(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xls"
    input_path.write_text("dummy", encoding="utf-8")
    ops = [PatchOp(op="add_sheet", sheet="Sheet2")]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    with pytest.raises(ValueError, match=r"requires Windows Excel COM"):
        run_patch(request, policy=PathPolicy(root=tmp_path))


def test_run_patch_xlsm_openpyxl_uses_keep_vba(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsm"
    input_path.write_bytes(b"dummy")
    calls: dict[str, object] = {}

    class FakeCell:
        def __init__(self, value: str | int | float | None) -> None:
            self.value = value
            self.data_type: str | None = None

    class FakeSheet:
        def __init__(self) -> None:
            self._cells: dict[str, FakeCell] = {"A1": FakeCell("old")}

        def __getitem__(self, key: str) -> FakeCell:
            if key not in self._cells:
                self._cells[key] = FakeCell(None)
            return self._cells[key]

    class FakeWorkbook:
        def __init__(self) -> None:
            self._sheets: dict[str, FakeSheet] = {"Sheet1": FakeSheet()}
            self.sheetnames = ["Sheet1"]

        def __getitem__(self, key: str) -> FakeSheet:
            return self._sheets[key]

        def create_sheet(self, title: str) -> FakeSheet:
            sheet = FakeSheet()
            self._sheets[title] = sheet
            self.sheetnames.append(title)
            return sheet

        def save(self, filename: str | Path) -> None:
            calls["saved"] = str(filename)

        def close(self) -> None:
            calls["closed"] = True

    fake_workbook = FakeWorkbook()

    def _fake_load_workbook(path: Path, **kwargs: object) -> FakeWorkbook:
        calls["path"] = str(path)
        calls["keep_vba"] = kwargs.get("keep_vba", False)
        return fake_workbook

    monkeypatch.setattr("openpyxl.load_workbook", _fake_load_workbook)

    request = PatchRequest(
        xlsx_path=input_path,
        ops=[PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")],
        on_conflict="rename",
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert calls["keep_vba"] is True
    assert calls["closed"] is True


def test_run_patch_dry_run_does_not_write(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    output_path = tmp_path / "book_patched.xlsx"
    assert not output_path.exists()
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        dry_run=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert not output_path.exists()
    assert len(result.patch_diff) == 1


def test_run_patch_return_inverse_ops(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="new")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        return_inverse_ops=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.inverse_ops) == 1
    inverse = result.inverse_ops[0]
    assert inverse.op == "set_value"
    assert inverse.cell == "A1"
    assert inverse.value == "old"


def test_run_patch_set_range_values(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_range_values",
            sheet="Sheet1",
            range="A2:B3",
            values=[["r1c1", "r1c2"], ["r2c1", "r2c2"]],
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A2"].value == "r1c1"
        assert sheet["B3"].value == "r2c2"
    finally:
        workbook.close()


def test_run_patch_set_range_values_size_mismatch(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_range_values",
            sheet="Sheet1",
            range="A2:B3",
            values=[["only_one_column"], ["still_one_column"]],
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert "width does not match range" in result.error.message


def test_run_patch_set_value_if_skipped(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_value_if",
            sheet="Sheet1",
            cell="A1",
            expected="not_old",
            value="new",
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.patch_diff) == 1
    assert result.patch_diff[0].status == "skipped"


def test_run_patch_fill_formula(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        sheet = workbook["Sheet1"]
        sheet["A2"] = 1
        sheet["A3"] = 2
        sheet["A4"] = 3
        sheet["B2"] = 10
        sheet["B3"] = 20
        sheet["B4"] = 30
        workbook.save(input_path)
    finally:
        workbook.close()
    ops = [
        PatchOp(
            op="fill_formula",
            sheet="Sheet1",
            range="C2:C4",
            base_cell="C2",
            formula="=A2+B2",
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["C2"].value == "=A2+B2"
        assert sheet["C3"].value == "=A3+B3"
        assert sheet["C4"].value == "=A4+B4"
    finally:
        workbook.close()


def test_run_patch_formula_health_check(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!+1")]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        preflight_formula_check=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.formula_issues
    assert result.formula_issues[0].code == "ref_error"


def test_run_patch_formula_health_check_reports_matching_op(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_formula", sheet="Sheet1", cell="B1", formula="=SUM(1,1)"),
        PatchOp(op="set_formula", sheet="Sheet1", cell="A1", formula="=#REF!+1"),
    ]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        preflight_formula_check=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is not None
    assert result.error.op_index == 1
    assert result.error.op == "set_formula"


def test_patch_op_add_sheet_rejects_unrelated_fields() -> None:
    with pytest.raises(ValidationError, match="add_sheet does not accept range"):
        PatchOp(op="add_sheet", sheet="NewSheet", range="A1:A1")


def test_patch_op_set_value_rejects_expected() -> None:
    with pytest.raises(ValidationError, match="set_value does not accept expected"):
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value="x", expected="old")


def test_run_patch_draw_grid_border_and_inverse_restore(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="draw_grid_border",
            sheet="Sheet1",
            base_cell="A1",
            row_count=2,
            col_count=2,
        )
    ]
    request = PatchRequest(
        xlsx_path=input_path,
        ops=ops,
        on_conflict="rename",
        return_inverse_ops=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.inverse_ops) == 1
    workbook = load_workbook(result.out_path)
    try:
        assert workbook["Sheet1"]["A1"].border.top.style == "thin"
    finally:
        workbook.close()

    restored = run_patch(
        PatchRequest(
            xlsx_path=Path(result.out_path),
            ops=result.inverse_ops,
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    restored_book = load_workbook(restored.out_path)
    try:
        assert restored_book["Sheet1"]["A1"].border.top.style is None
    finally:
        restored_book.close()


def test_run_patch_set_bold_and_fill_color(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_bold", sheet="Sheet1", range="A1:B1"),
        PatchOp(op="set_fill_color", sheet="Sheet1", cell="A1", fill_color="#112233"),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        cell = workbook["Sheet1"]["A1"]
        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"
        assert cell.fill.start_color.rgb == "FF112233"
    finally:
        workbook.close()


def test_run_patch_set_font_size_cell_and_range(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_font_size", sheet="Sheet1", cell="A1", font_size=14.5),
        PatchOp(op="set_font_size", sheet="Sheet1", range="A1:B1", font_size=16.0),
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].font.size == 16.0
        assert sheet["B1"].font.size == 16.0
    finally:
        workbook.close()


def test_run_patch_set_font_size_preserves_other_font_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"]["A1"].font = Font(
            name="Calibri", bold=True, italic=True, size=11.0
        )
        workbook.save(input_path)
    finally:
        workbook.close()

    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(op="set_font_size", sheet="Sheet1", cell="A1", font_size=18.0)
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    out_book = load_workbook(result.out_path)
    try:
        font = out_book["Sheet1"]["A1"].font
        assert font.name == "Calibri"
        assert font.bold is True
        assert font.italic is True
        assert font.size == 18.0
    finally:
        out_book.close()


def test_run_patch_set_dimensions(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(
            op="set_dimensions",
            sheet="Sheet1",
            rows=[1, 2],
            row_height=24.5,
            columns=["A", 2],
            column_width=18.0,
        )
    ]
    request = PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename")
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet.row_dimensions[1].height == 24.5
        assert sheet.column_dimensions["A"].width == 18.0
        assert sheet.column_dimensions["B"].width == 18.0
    finally:
        workbook.close()
    after_value = result.patch_diff[0].after
    assert after_value is not None
    assert isinstance(after_value.value, str)
    assert "columns=A, B (2)" in after_value.value


def test_run_patch_auto_fit_columns_with_bounds(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        sheet = workbook["Sheet1"]
        sheet["A1"] = "short"
        sheet["B1"] = "this is a much longer sample text"
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="auto_fit_columns",
                    sheet="Sheet1",
                    min_width=8,
                    max_width=20,
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        width_a = sheet.column_dimensions["A"].width
        width_b = sheet.column_dimensions["B"].width
        assert width_a is not None
        assert width_b is not None
        assert 8 <= width_a <= 20
        assert 8 <= width_b <= 20
        assert width_b >= width_a
    finally:
        workbook.close()


def test_run_patch_auto_fit_columns_accepts_mixed_column_identifiers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="auto_fit_columns",
                    sheet="Sheet1",
                    columns=["A", 2],
                    min_width=9,
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    workbook = load_workbook(result.out_path)
    try:
        sheet = workbook["Sheet1"]
        assert sheet.column_dimensions["A"].width is not None
        assert sheet.column_dimensions["B"].width is not None
        assert sheet.column_dimensions["A"].width >= 9
        assert sheet.column_dimensions["B"].width >= 9
    finally:
        workbook.close()


def test_patch_op_auto_fit_columns_rejects_invalid_bounds() -> None:
    with pytest.raises(
        ValidationError, match="auto_fit_columns requires min_width <= max_width"
    ):
        PatchOp(
            op="auto_fit_columns",
            sheet="Sheet1",
            min_width=20,
            max_width=10,
        )


def test_run_patch_warns_when_ops_exceed_soft_threshold(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    ops = [
        PatchOp(op="set_value", sheet="Sheet1", cell="A1", value=f"v{i}")
        for i in range(201)
    ]
    result = run_patch(
        PatchRequest(xlsx_path=input_path, ops=ops, on_conflict="rename"),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert any("Recommended maximum is 200" in warning for warning in result.warnings)


def test_patch_op_set_bold_rejects_cell_and_range() -> None:
    with pytest.raises(
        ValidationError, match="set_bold requires exactly one of cell or range"
    ):
        PatchOp(op="set_bold", sheet="Sheet1", cell="A1", range="A1:A1")


def test_patch_op_set_fill_color_rejects_invalid_color() -> None:
    with pytest.raises(
        ValidationError,
        match="Invalid fill_color format. Use 'RRGGBB', 'AARRGGBB', '#RRGGBB', or '#AARRGGBB'",
    ):
        PatchOp(op="set_fill_color", sheet="Sheet1", cell="A1", fill_color="red")


def test_patch_op_normalizes_hex_inputs() -> None:
    fill_op = PatchOp(
        op="set_fill_color",
        sheet="Sheet1",
        cell="A1",
        fill_color="1f4e79",
    )
    font_op = PatchOp(
        op="set_font_color",
        sheet="Sheet1",
        cell="A1",
        color="cc336699",
    )
    assert fill_op.fill_color == "#1F4E79"
    assert font_op.color == "#CC336699"


def test_patch_op_set_font_color_rejects_fill_color() -> None:
    with pytest.raises(
        ValidationError, match="set_font_color does not accept fill_color"
    ):
        PatchOp(
            op="set_font_color",
            sheet="Sheet1",
            cell="A1",
            color="#112233",
            fill_color="#445566",
        )


def test_patch_op_set_fill_color_rejects_color() -> None:
    with pytest.raises(ValidationError, match="set_fill_color does not accept color"):
        PatchOp(
            op="set_fill_color",
            sheet="Sheet1",
            cell="A1",
            fill_color="#112233",
            color="#445566",
        )


def test_patch_op_set_style_requires_at_least_one_style_field() -> None:
    with pytest.raises(ValidationError, match="set_style requires at least one style"):
        PatchOp(op="set_style", sheet="Sheet1", cell="A1")


def test_patch_op_set_style_rejects_cell_and_range() -> None:
    with pytest.raises(
        ValidationError, match="set_style requires exactly one of cell or range"
    ):
        PatchOp(
            op="set_style",
            sheet="Sheet1",
            cell="A1",
            range="A1:B1",
            bold=True,
        )


def test_patch_op_apply_table_style_requires_style() -> None:
    with pytest.raises(ValidationError, match="apply_table_style requires style"):
        PatchOp(op="apply_table_style", sheet="Sheet1", range="A1:B3")


def test_patch_op_apply_table_style_rejects_cell() -> None:
    with pytest.raises(
        ValidationError, match="apply_table_style does not accept cell or base_cell"
    ):
        PatchOp(
            op="apply_table_style",
            sheet="Sheet1",
            cell="A1",
            range="A1:B3",
            style="TableStyleMedium2",
        )


def test_run_patch_set_font_color(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(op="set_font_color", sheet="Sheet1", cell="A1", color="112233")
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    out_book = load_workbook(result.out_path)
    try:
        color = out_book["Sheet1"]["A1"].font.color
        assert color is not None
        assert str(getattr(color, "rgb", "")).upper() == "FF112233"
    finally:
        out_book.close()


def test_patch_op_set_font_size_rejects_non_positive() -> None:
    with pytest.raises(ValidationError, match="set_font_size font_size must be > 0"):
        PatchOp(op="set_font_size", sheet="Sheet1", cell="A1", font_size=0)


def test_patch_op_set_font_size_rejects_cell_and_range() -> None:
    with pytest.raises(
        ValidationError, match="set_font_size requires exactly one of cell or range"
    ):
        PatchOp(
            op="set_font_size",
            sheet="Sheet1",
            cell="A1",
            range="A1:A1",
            font_size=12,
        )


def test_patch_op_set_font_size_requires_target() -> None:
    with pytest.raises(
        ValidationError, match="set_font_size requires exactly one of cell or range"
    ):
        PatchOp(op="set_font_size", sheet="Sheet1", font_size=12)


def test_patch_op_set_dimensions_requires_dimension_pair() -> None:
    with pytest.raises(
        ValidationError,
        match="set_dimensions requires column_width when columns is provided",
    ):
        PatchOp(op="set_dimensions", sheet="Sheet1", columns=["A"])


def test_patch_op_style_target_limit() -> None:
    with pytest.raises(ValidationError, match="target exceeds max cells"):
        PatchOp(op="set_bold", sheet="Sheet1", range="A1:Z500")


def test_run_patch_rejects_design_op_for_xls(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )
    input_path = tmp_path / "book.xls"
    input_path.write_text("dummy", encoding="utf-8")
    request = PatchRequest(
        xlsx_path=input_path,
        ops=[PatchOp(op="set_bold", sheet="Sheet1", cell="A1")],
        on_conflict="rename",
    )
    with pytest.raises(
        ValueError, match=r"Design operations are not supported for \.xls files"
    ):
        run_patch(request, policy=PathPolicy(root=tmp_path))


def test_patch_op_merge_cells_requires_multi_cell_range() -> None:
    with pytest.raises(
        ValidationError, match="merge_cells requires a multi-cell range"
    ):
        PatchOp(op="merge_cells", sheet="Sheet1", range="A1:A1")


def test_run_patch_merge_cells_and_inverse_restore(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    request = PatchRequest(
        xlsx_path=input_path,
        ops=[PatchOp(op="merge_cells", sheet="Sheet1", range="A1:B1")],
        on_conflict="rename",
        return_inverse_ops=True,
    )
    result = run_patch(request, policy=PathPolicy(root=tmp_path))
    assert result.error is None
    assert len(result.inverse_ops) == 1

    workbook = load_workbook(result.out_path)
    try:
        ranges = [str(item) for item in workbook["Sheet1"].merged_cells.ranges]
        assert ranges == ["A1:B1"]
    finally:
        workbook.close()

    restored = run_patch(
        PatchRequest(
            xlsx_path=Path(result.out_path),
            ops=result.inverse_ops,
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    restored_book = load_workbook(restored.out_path)
    try:
        assert list(restored_book["Sheet1"].merged_cells.ranges) == []
    finally:
        restored_book.close()


def test_run_patch_merge_cells_rejects_overlap(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"].merge_cells("A1:B1")
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="merge_cells", sheet="Sheet1", range="B1:C1")],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is not None
    assert "overlaps existing merged ranges" in result.error.message


def test_run_patch_merge_cells_warns_on_value_loss(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"]["B1"] = "drop-me"
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="merge_cells", sheet="Sheet1", range="A1:B1")],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert any(
        "may clear non-top-left values" in warning for warning in result.warnings
    )


def test_run_patch_unmerge_cells_for_intersections(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        sheet = workbook["Sheet1"]
        sheet["C1"] = "v"
        sheet["D1"] = "w"
        sheet.merge_cells("A1:B1")
        sheet.merge_cells("C1:D1")
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[PatchOp(op="unmerge_cells", sheet="Sheet1", range="B1:C1")],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    out_book = load_workbook(result.out_path)
    try:
        assert list(out_book["Sheet1"].merged_cells.ranges) == []
    finally:
        out_book.close()


def test_run_patch_set_alignment_preserves_unspecified_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"]["A1"].alignment = Alignment(vertical="top", wrap_text=True)
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="set_alignment",
                    sheet="Sheet1",
                    cell="A1",
                    horizontal_align="center",
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    out_book = load_workbook(result.out_path)
    try:
        alignment = out_book["Sheet1"]["A1"].alignment
        assert alignment.horizontal == "center"
        assert alignment.vertical == "top"
        assert alignment.wrap_text is True
    finally:
        out_book.close()


def test_run_patch_set_alignment_inverse_restore(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"]["A1"].alignment = Alignment(
            horizontal="left", vertical="bottom", wrap_text=True
        )
        workbook.save(input_path)
    finally:
        workbook.close()
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="set_alignment",
                    sheet="Sheet1",
                    range="A1:B1",
                    horizontal_align="center",
                    vertical_align="center",
                    wrap_text=False,
                )
            ],
            on_conflict="rename",
            return_inverse_ops=True,
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert len(result.inverse_ops) == 1
    restored = run_patch(
        PatchRequest(
            xlsx_path=Path(result.out_path),
            ops=result.inverse_ops,
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    restored_book = load_workbook(restored.out_path)
    try:
        alignment = restored_book["Sheet1"]["A1"].alignment
        assert alignment.horizontal == "left"
        assert alignment.vertical == "bottom"
        assert alignment.wrap_text is True
    finally:
        restored_book.close()


def test_run_patch_set_style_and_inverse_restore(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        workbook["Sheet1"]["A1"].alignment = Alignment(horizontal="left")
        workbook.save(input_path)
    finally:
        workbook.close()

    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="set_style",
                    sheet="Sheet1",
                    range="A1:B1",
                    bold=True,
                    color="#112233",
                    fill_color="#D9E1F2",
                    horizontal_align="center",
                    wrap_text=True,
                )
            ],
            on_conflict="rename",
            return_inverse_ops=True,
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    assert len(result.inverse_ops) == 1

    out_book = load_workbook(result.out_path)
    try:
        cell = out_book["Sheet1"]["A1"]
        assert cell.font.bold is True
        assert str(getattr(cell.font.color, "rgb", "")).upper() == "FF112233"
        assert cell.fill.fill_type == "solid"
        assert str(getattr(cell.fill.start_color, "rgb", "")).upper() == "FFD9E1F2"
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.wrap_text is True
    finally:
        out_book.close()

    restored = run_patch(
        PatchRequest(
            xlsx_path=Path(result.out_path),
            ops=result.inverse_ops,
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    restored_book = load_workbook(restored.out_path)
    try:
        restored_cell = restored_book["Sheet1"]["A1"]
        assert restored_cell.font.bold is False
        assert restored_cell.fill.fill_type is None
        assert restored_cell.alignment.horizontal == "left"
    finally:
        restored_book.close()


def test_run_patch_apply_table_style(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    _seed_table_source(input_path)
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="A1:B3",
                    style="TableStyleMedium2",
                    table_name="SalesTable",
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is None
    out_book = load_workbook(result.out_path)
    try:
        sheet = out_book["Sheet1"]
        table = sheet.tables["SalesTable"]
        assert table.ref == "A1:B3"
        style_info = table.tableStyleInfo
        assert style_info is not None
        assert style_info.name == "TableStyleMedium2"
    finally:
        out_book.close()


def test_run_patch_apply_table_style_rejects_duplicate_table_name(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    _seed_table_source(input_path)
    result = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="A1:B3",
                    style="TableStyleMedium2",
                    table_name="SalesTable",
                ),
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="D1:E3",
                    style="TableStyleMedium2",
                    table_name="SalesTable",
                ),
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is not None
    assert "Table name already exists" in result.error.message


def test_run_patch_apply_table_style_rejects_intersection(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    _seed_table_source(input_path)
    first = run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="A1:B3",
                    style="TableStyleMedium2",
                    table_name="SalesTable",
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert first.error is None
    second = run_patch(
        PatchRequest(
            xlsx_path=Path(first.out_path),
            ops=[
                PatchOp(
                    op="apply_table_style",
                    sheet="Sheet1",
                    range="B2:C4",
                    style="TableStyleMedium2",
                    table_name="SalesTable2",
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert second.error is not None
    assert "intersects existing table" in second.error.message


def test_run_patch_rejects_alignment_design_op_for_xls(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(
        patch_runner,
        "get_com_availability",
        lambda: ComAvailability(available=True, reason=None),
    )
    input_path = tmp_path / "book.xls"
    input_path.write_text("dummy", encoding="utf-8")
    request = PatchRequest(
        xlsx_path=input_path,
        ops=[
            PatchOp(
                op="set_alignment",
                sheet="Sheet1",
                cell="A1",
                horizontal_align="center",
            )
        ],
        on_conflict="rename",
    )
    with pytest.raises(
        ValueError, match=r"Design operations are not supported for \.xls files"
    ):
        run_patch(request, policy=PathPolicy(root=tmp_path))
