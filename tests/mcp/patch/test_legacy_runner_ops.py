from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.mcp.io import PathPolicy
from exstruct.mcp.patch import internal as legacy_runner
from exstruct.mcp.patch.internal import PatchOp, PatchRequest


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    sheet["B1"] = 1
    workbook.save(path)
    workbook.close()


def _disable_com(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        legacy_runner,
        "get_com_availability",
        lambda: ComAvailability(available=False, reason="test"),
    )


def test_run_patch_auto_fit_columns_openpyxl_uses_single_pass_collector(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)
    workbook = load_workbook(input_path)
    try:
        sheet = workbook["Sheet1"]
        sheet["A1"] = "a"
        sheet["B1"] = "bbbbbbbb"
        sheet["C1"] = "cccccccccccc"
        workbook.save(input_path)
    finally:
        workbook.close()

    call_count = 0
    original = legacy_runner._collect_openpyxl_target_column_max_lengths

    def _counting_collector(
        sheet: legacy_runner.OpenpyxlWorksheetProtocol, target_indexes: set[int]
    ) -> dict[int, int]:
        nonlocal call_count
        call_count += 1
        return original(sheet, target_indexes)

    monkeypatch.setattr(
        legacy_runner,
        "_collect_openpyxl_target_column_max_lengths",
        _counting_collector,
    )

    result = legacy_runner.run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="auto_fit_columns",
                    sheet="Sheet1",
                    columns=["A", "B", "C"],
                )
            ],
            on_conflict="rename",
        ),
        policy=PathPolicy(root=tmp_path),
    )

    assert result.error is None
    assert call_count == 1


def test_apply_xlwings_set_font_size() -> None:
    class _FakeFontApi:
        Size: float = 0.0

    class _FakeRangeApi:
        Font: _FakeFontApi

        def __init__(self) -> None:
            self.Font = _FakeFontApi()

    class _FakeRange:
        value: object | None = None
        formula: str | None = None
        api: object

        def __init__(self, api: _FakeRangeApi) -> None:
            self.api = api

    class _FakeSheet:
        name = "Sheet1"
        api = object()

        def __init__(self) -> None:
            self.range_api = _FakeRangeApi()
            self.last_ref = ""

        def range(self, cell: str) -> legacy_runner.XlwingsRangeProtocol:
            self.last_ref = cell
            return _FakeRange(self.range_api)

    sheet = _FakeSheet()
    op = PatchOp(op="set_font_size", sheet="Sheet1", range="A1:B2", font_size=13.0)
    diff = legacy_runner._apply_xlwings_set_font_size(sheet, op, index=0)

    assert sheet.last_ref == "A1:B2"
    assert sheet.range_api.Font.Size == 13.0
    assert diff.after is not None
    assert diff.after.value == "font_size=13.0"


def test_run_patch_error_includes_hint_for_known_set_fill_color_mistake(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _disable_com(monkeypatch)
    input_path = tmp_path / "book.xlsx"
    _create_workbook(input_path)

    def _raise_known_error(
        sheet: legacy_runner.OpenpyxlWorksheetProtocol,
        op: PatchOp,
        index: int,
    ) -> tuple[legacy_runner.PatchDiffItem, PatchOp | None]:
        raise ValueError("set_fill_color does not accept color.")

    monkeypatch.setattr(
        legacy_runner,
        "_apply_openpyxl_set_fill_color",
        _raise_known_error,
    )
    result = legacy_runner.run_patch(
        PatchRequest(
            xlsx_path=input_path,
            ops=[
                PatchOp(
                    op="set_fill_color",
                    sheet="Sheet1",
                    cell="A1",
                    fill_color="#112233",
                )
            ],
            on_conflict="rename",
            backend="openpyxl",
        ),
        policy=PathPolicy(root=tmp_path),
    )
    assert result.error is not None
    assert result.error.hint is not None
    assert "fill_color" in result.error.hint
    assert result.error.expected_fields
    assert result.error.example_op is not None
