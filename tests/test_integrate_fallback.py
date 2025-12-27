from pathlib import Path
from typing import Never

from _pytest.monkeypatch import MonkeyPatch
from openpyxl import Workbook

from exstruct.core import integrate
from exstruct.models import CellRow


def test_extract_workbook_fallback_on_com_failure(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    # create a tiny workbook
    xlsx = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    from openpyxl.worksheet.table import Table

    ws.add_table(Table(displayName="Table1", ref="A1:B2"))
    wb.save(xlsx)
    wb.close()

    def boom(_path: Path, *args: object, **kwargs: object) -> Never:
        raise RuntimeError("COM unavailable")

    monkeypatch.setattr("exstruct.core.pipeline.xlwings_workbook", boom)
    result = integrate.extract_workbook(xlsx, mode="standard")
    assert result.sheets["Sheet1"].shapes == []
    assert result.sheets["Sheet1"].charts == []
    # table_candidates populated via openpyxl detection fallback
    assert result.sheets["Sheet1"].table_candidates


def test_extract_workbook_fallback_includes_print_areas(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    xlsx = tmp_path / "print_area.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.print_area = "A1:B2"
    wb.save(xlsx)
    wb.close()

    def boom(_path: Path, *args: object, **kwargs: object) -> Never:
        raise RuntimeError("COM unavailable")

    monkeypatch.setattr("exstruct.core.pipeline.xlwings_workbook", boom)
    result = integrate.extract_workbook(xlsx, mode="standard")
    assert result.sheets["Sheet1"].print_areas


def test_extract_workbook_with_links(monkeypatch: MonkeyPatch, tmp_path: Path) -> None:
    # create workbook with hyperlink
    path = tmp_path / "links.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    cell = ws["A1"]
    cell.value = "click"
    cell.hyperlink = "http://example.com"
    wb.save(path)
    wb.close()

    def _raise(*_args: object, **_kwargs: object) -> Never:
        raise RuntimeError("no COM")

    monkeypatch.setattr("exstruct.core.pipeline.xlwings_workbook", _raise)

    result = integrate.extract_workbook(path, mode="standard", include_cell_links=True)
    row = result.sheets["Sheet1"].rows[0]
    assert isinstance(row, CellRow)
    assert row.links == {"0": "http://example.com"}
