from __future__ import annotations

from contextlib import redirect_stderr, redirect_stdout
import io
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
import pytest

from exstruct.cli.availability import ComAvailability
import exstruct.cli.edit as edit_cli_module
from exstruct.cli.edit import is_edit_subcommand
import exstruct.cli.main as cli_main_module
from exstruct.cli.main import build_parser, main as cli_main


class CliResult(BaseModel):
    """Captured result from one in-process CLI run."""

    returncode: int
    stdout: str
    stderr: str


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet["A1"] = "old"
    workbook.save(path)
    workbook.close()


def _read_cell(path: Path, sheet_name: str, cell: str) -> object:
    workbook = load_workbook(path)
    try:
        return workbook[sheet_name][cell].value
    finally:
        workbook.close()


def _run_cli(args: list[str], *, stdin_text: str | None = None) -> CliResult:
    stdout_buffer = io.StringIO()
    stderr_buffer = io.StringIO()
    original_stdin = io.StringIO(stdin_text) if stdin_text is not None else None
    with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
        if original_stdin is None:
            returncode = cli_main(argv=args)
        else:
            import sys

            previous_stdin = sys.stdin
            try:
                sys.stdin = original_stdin
                returncode = cli_main(argv=args)
            finally:
                sys.stdin = previous_stdin
    return CliResult(
        returncode=returncode,
        stdout=stdout_buffer.getvalue(),
        stderr=stderr_buffer.getvalue(),
    )


def test_patch_cli_returns_patch_result_json(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text(
        json.dumps(
            [{"op": "set_value", "sheet": "Sheet1", "cell": "A1", "value": "new"}]
        ),
        encoding="utf-8",
    )

    result = _run_cli(
        [
            "patch",
            "--input",
            str(source),
            "--ops",
            str(ops_path),
            "--backend",
            "openpyxl",
        ]
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["engine"] == "openpyxl"
    assert payload["error"] is None
    assert Path(payload["out_path"]).exists()
    assert result.stderr == ""


def test_patch_cli_reads_ops_from_stdin(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    _create_workbook(source)

    result = _run_cli(
        [
            "patch",
            "--input",
            str(source),
            "--ops",
            "-",
            "--backend",
            "openpyxl",
        ],
        stdin_text=json.dumps(
            [{"op": "set_value", "sheet": "Sheet1", "cell": "A1", "value": "stdin"}]
        ),
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["error"] is None
    assert Path(payload["out_path"]).exists()


def test_patch_cli_applies_top_level_sheet_fallback(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text(
        json.dumps([{"op": "set_value", "cell": "A1", "value": "fallback"}]),
        encoding="utf-8",
    )

    result = _run_cli(
        [
            "patch",
            "--input",
            str(source),
            "--ops",
            str(ops_path),
            "--sheet",
            "Sheet1",
            "--backend",
            "openpyxl",
        ]
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["error"] is None
    assert _read_cell(Path(payload["out_path"]), "Sheet1", "A1") == "fallback"


def test_patch_cli_returns_nonzero_for_invalid_ops_json(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text("{bad json", encoding="utf-8")

    result = _run_cli(["patch", "--input", str(source), "--ops", str(ops_path)])

    assert result.returncode == 1
    assert "Invalid JSON in --ops" in result.stderr


def test_patch_cli_converts_argparse_errors_to_exit_one() -> None:
    result = _run_cli(["patch"])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "required" in result.stderr


def test_patch_cli_help_keeps_exit_zero() -> None:
    result = _run_cli(["patch", "--help"])

    assert result.returncode == 0
    assert "--input INPUT" in result.stdout
    assert "--backend {auto,com,openpyxl}" in result.stdout
    assert "--ops OPS" in result.stdout
    assert result.stderr == ""


def test_patch_cli_returns_nonzero_when_patch_result_contains_error(
    tmp_path: Path,
) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text(
        json.dumps(
            [{"op": "set_value", "sheet": "Missing", "cell": "A1", "value": "new"}]
        ),
        encoding="utf-8",
    )

    result = _run_cli(
        [
            "patch",
            "--input",
            str(source),
            "--ops",
            str(ops_path),
            "--backend",
            "openpyxl",
        ]
    )

    assert result.returncode == 1
    payload = json.loads(result.stdout)
    assert payload["error"] is not None
    assert payload["error"]["sheet"] == "Missing"


def test_patch_cli_returns_nonzero_when_backend_raises_runtime_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text("[]", encoding="utf-8")

    def _raise_runtime_error(_request: object) -> object:
        raise RuntimeError("backend boom")

    monkeypatch.setattr(edit_cli_module, "patch_workbook", _raise_runtime_error)

    result = _run_cli(["patch", "--input", str(source), "--ops", str(ops_path)])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "Error: backend boom" in result.stderr


def test_patch_cli_returns_nonzero_when_sheet_resolution_breaks_contract(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    ops_path = tmp_path / "ops.json"
    _create_workbook(source)
    ops_path.write_text("[]", encoding="utf-8")

    monkeypatch.setattr(
        edit_cli_module, "resolve_top_level_sheet_for_payload", lambda _payload: []
    )

    result = _run_cli(["patch", "--input", str(source), "--ops", str(ops_path)])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "Top-level sheet resolution must return a dict payload." in result.stderr


def test_make_cli_creates_workbook_and_returns_json(tmp_path: Path) -> None:
    output = tmp_path / "created.xlsx"
    ops_path = tmp_path / "ops.json"
    ops_path.write_text(
        json.dumps(
            [
                {"op": "add_sheet", "sheet": "Data"},
                {"op": "set_value", "sheet": "Data", "cell": "A1", "value": "ok"},
            ]
        ),
        encoding="utf-8",
    )

    result = _run_cli(
        [
            "make",
            "--output",
            str(output),
            "--ops",
            str(ops_path),
            "--backend",
            "openpyxl",
        ]
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["error"] is None
    assert output.exists()


def test_make_cli_applies_top_level_sheet_fallback(tmp_path: Path) -> None:
    output = tmp_path / "created.xlsx"
    ops_path = tmp_path / "ops.json"
    ops_path.write_text(
        json.dumps(
            [
                {"op": "add_sheet", "sheet": "Data"},
                {"op": "set_value", "cell": "A1", "value": "fallback"},
            ]
        ),
        encoding="utf-8",
    )

    result = _run_cli(
        [
            "make",
            "--output",
            str(output),
            "--ops",
            str(ops_path),
            "--sheet",
            "Data",
            "--backend",
            "openpyxl",
        ]
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["error"] is None
    assert _read_cell(output, "Data", "A1") == "fallback"


def test_make_cli_returns_nonzero_when_backend_raises_runtime_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output = tmp_path / "created.xlsx"

    def _raise_runtime_error(_request: object) -> object:
        raise RuntimeError("make boom")

    monkeypatch.setattr(edit_cli_module, "make_workbook", _raise_runtime_error)

    result = _run_cli(["make", "--output", str(output)])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "Error: make boom" in result.stderr


def test_make_cli_returns_nonzero_when_resolved_ops_contract_breaks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output = tmp_path / "created.xlsx"
    ops_path = tmp_path / "ops.json"
    ops_path.write_text("[]", encoding="utf-8")

    monkeypatch.setattr(
        edit_cli_module,
        "resolve_top_level_sheet_for_payload",
        lambda _payload: {"ops": "bad"},
    )

    result = _run_cli(["make", "--output", str(output), "--ops", str(ops_path)])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "Resolved patch ops payload must contain an ops list." in result.stderr


def test_make_cli_defaults_to_empty_ops(tmp_path: Path) -> None:
    output = tmp_path / "empty.xlsx"

    result = _run_cli(["make", "--output", str(output), "--backend", "openpyxl"])

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["error"] is None
    assert output.exists()


def test_ops_list_cli_returns_compact_schema_summary() -> None:
    result = _run_cli(["ops", "list"])

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert "ops" in payload
    assert any(item["op"] == "set_value" for item in payload["ops"])


def test_ops_describe_cli_returns_schema_detail() -> None:
    result = _run_cli(["ops", "describe", "create_chart"])

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["op"] == "create_chart"
    assert "required" in payload
    assert "example" in payload


def test_ops_describe_cli_rejects_unknown_op() -> None:
    result = _run_cli(["ops", "describe", "missing_op"])

    assert result.returncode == 1
    assert "Unknown patch operation" in result.stderr


def test_validate_cli_returns_json_for_readable_file(tmp_path: Path) -> None:
    path = tmp_path / "input.xlsx"
    path.write_bytes(b"x")

    result = _run_cli(["validate", "--input", str(path)])

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["is_readable"] is True


def test_validate_cli_returns_nonzero_for_missing_file(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"

    result = _run_cli(["validate", "--input", str(path)])

    assert result.returncode == 1
    payload = json.loads(result.stdout)
    assert payload["is_readable"] is False
    assert payload["errors"]


def test_validate_cli_returns_nonzero_when_validation_raises(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "input.xlsx"
    path.write_bytes(b"x")

    def _raise_os_error(_request: object) -> object:
        raise OSError("boom")

    monkeypatch.setattr(edit_cli_module, "validate_input", _raise_os_error)

    result = _run_cli(["validate", "--input", str(path)])

    assert result.returncode == 1
    assert result.stdout == ""
    assert "Error: boom" in result.stderr


def test_extraction_help_mentions_editing_commands() -> None:
    help_text = build_parser(
        availability=ComAvailability(available=False, reason="test")
    ).format_help()

    assert "Editing commands:" in help_text
    assert "exstruct patch --input book.xlsx --ops ops.json" in help_text


def test_main_keeps_legacy_input_on_extraction_path(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "patch.xlsx"
    source.write_bytes(b"x")
    called: dict[str, Path] = {}

    def _fake_process_excel(*, file_path: Path, **_kwargs: object) -> None:
        called["file_path"] = file_path

    monkeypatch.setattr(cli_main_module, "process_excel", _fake_process_excel)

    result = _run_cli([str(source)])

    assert result.returncode == 0
    assert called["file_path"] == source
    assert is_edit_subcommand([str(source)]) is False


@pytest.mark.parametrize("name", ["patch", "make", "ops", "validate"])  # type: ignore[misc]
def test_main_prefers_existing_legacy_input_for_ambiguous_command_names(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, name: str
) -> None:
    source = tmp_path / name
    source.write_bytes(b"x")
    called: dict[str, Path] = {}

    def _fake_process_excel(*, file_path: Path, **_kwargs: object) -> None:
        called["file_path"] = file_path

    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(cli_main_module, "process_excel", _fake_process_excel)

    result = _run_cli([name])

    assert result.returncode == 0
    assert called["file_path"] == Path(name)
    assert is_edit_subcommand([name]) is False


@pytest.mark.parametrize(  # type: ignore[misc]
    "argv",
    [
        ["patch", "--help"],
        ["patch", "--input", "book.xlsx", "--ops", "ops.json"],
        ["patch", "--input=book.xlsx", "--ops=ops.json"],
        ["make", "--help"],
        ["make", "--ops", "ops.json"],
        ["make", "--ops=ops.json"],
        ["ops", "list"],
        ["validate", "--input", "book.xlsx"],
        ["validate", "--input=book.xlsx"],
    ],
)
def test_is_edit_subcommand_keeps_explicit_edit_syntax_even_when_name_collides(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, argv: list[str]
) -> None:
    (tmp_path / argv[0]).write_bytes(b"x")
    monkeypatch.chdir(tmp_path)

    assert is_edit_subcommand(argv) is True
