from pathlib import Path
import subprocess
import sys

from openpyxl import Workbook
import pytest
import xlwings as xw

from exstruct import extract, process_excel


def _make_basic_book(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "v1"
    ws["B1"] = "v2"
    wb.save(path)


def _ensure_excel() -> None:
    try:
        app = xw.App(add_book=False, visible=False)
        app.quit()
    except Exception:
        pytest.skip("Excel COM is unavailable; skipping Excel-dependent test.")


def _make_shapes_book(path: Path) -> None:
    app = xw.App(add_book=False, visible=False)
    try:
        wb = app.books.add()
        sht = wb.sheets[0]
        sht.name = "Sheet1"
        rect = sht.api.Shapes.AddShape(1, 50, 50, 120, 60)
        rect.TextFrame2.TextRange.Text = "rect"
        oval = sht.api.Shapes.AddShape(5, 200, 50, 80, 40)  # no text
        line = sht.api.Shapes.AddLine(10, 10, 110, 10)
        line.Line.EndArrowheadStyle = 3
        wb.save(str(path))
        wb.close()
    finally:
        try:
            app.quit()
        except Exception:
            pass


def test_lightモードではCOMに触れずセルとテーブルのみ(
    monkeypatch, tmp_path: Path
) -> None:
    path = tmp_path / "book.xlsx"
    _make_basic_book(path)

    def _boom(*_a, **_k):
        raise AssertionError("COM should not be accessed in light mode")

    monkeypatch.setattr("exstruct.core.integrate.xw.Book", _boom, raising=False)
    data = extract(path, mode="light")
    sheet = next(iter(data.sheets.values()))
    assert sheet.shapes == []
    assert sheet.charts == []


def test_standardモードはテキストなし図形を除外する(tmp_path: Path) -> None:
    _ensure_excel()
    path = tmp_path / "shapes.xlsx"
    _make_shapes_book(path)

    data = extract(path, mode="standard")
    shapes = data.sheets["Sheet1"].shapes
    texts = [s.text for s in shapes]
    assert "rect" in texts
    # textless shapes that remain must be矢印系のみ
    for s in shapes:
        if s.text != "":
            continue
        assert s.type is not None
        assert ("Line" in s.type) or ("Connector" in s.type) or ("Arrow" in s.type)


def test_verboseモードでは全図形と幅高さが出力される(tmp_path: Path) -> None:
    _ensure_excel()
    path = tmp_path / "shapes.xlsx"
    _make_shapes_book(path)

    data = extract(path, mode="verbose")
    shapes = data.sheets["Sheet1"].shapes
    texts = [s.text for s in shapes]
    assert "" in texts  # textless oval included
    for s in shapes:
        assert s.w is not None
        assert s.h is not None


def test_process_excelにモードが伝搬する(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    out = tmp_path / "out.json"
    _make_basic_book(path)
    process_excel(path, out, mode="light")
    assert out.exists()


def test_invalidモードはエラーになる(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_basic_book(path)
    with pytest.raises(ValueError):
        extract(path, mode="invalid")  # type: ignore[arg-type]

    out = tmp_path / "out.json"
    with pytest.raises(ValueError):
        process_excel(path, out, mode="invalid")  # type: ignore[arg-type]


def test_CLIのmode引数バリデーション(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_basic_book(path)
    cmd = [
        sys.executable,
        "-m",
        "exstruct.cli.main",
        str(path),
        "--mode",
        "invalid",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    assert result.returncode != 0


def _make_multi_sheet_book(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "v1"
    ws2 = wb.create_sheet("Data 02")
    ws2["A1"] = "v2"
    wb.save(path)


def test_process_excel_defaults_to_stdout(tmp_path: Path, capsys) -> None:
    path = tmp_path / "book.xlsx"
    _make_basic_book(path)

    process_excel(path, output_path=None, mode="light", pretty=True)
    captured = capsys.readouterr().out
    assert '"book_name": "book.xlsx"' in captured
    assert "Sheet1" in captured


def test_process_excel_sheets_dir_output(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_multi_sheet_book(path)
    sheets_dir = tmp_path / "sheets"

    process_excel(path, output_path=None, mode="light", sheets_dir=sheets_dir)

    files = list(sheets_dir.glob("*.json"))
    assert len(files) == 2
    names = {f.stem for f in files}
    assert "Sheet1" in names
    assert "Data 02" in names


def test_CLI_defaults_to_stdout(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _make_basic_book(path)
    cmd = [
        sys.executable,
        "-m",
        "exstruct.cli.main",
        str(path),
        "--mode",
        "light",
        "--pretty",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    assert result.returncode == 0
    assert '"book_name": "book.xlsx"' in result.stdout
