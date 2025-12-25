from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from exstruct.core.cells import extract_sheet_colors_map


def _write_color_workbook(path: Path) -> None:
    """Create a workbook with predefined fill colors for testing.

    Args:
        path: Target path to save the workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].fill = PatternFill(patternType="solid", fgColor="FFFFFF")
    ws["B1"].fill = PatternFill(patternType="solid", fgColor="AD3815")
    ws["C1"].fill = PatternFill(patternType="solid", fgColor="00FF00")
    wb.save(path)
    wb.close()


def test_colors_map_excludes_default_background(tmp_path: Path) -> None:
    """Exclude default background when include_default_background is False."""
    path = tmp_path / "colors.xlsx"
    _write_color_workbook(path)

    data = extract_sheet_colors_map(
        path, include_default_background=False, ignore_colors=None
    )
    sheet = data.get_sheet("Sheet1")
    assert sheet is not None
    assert "FFFFFF" not in sheet.colors_map
    assert "AD3815" in sheet.colors_map
    assert "00FF00" in sheet.colors_map


def test_colors_map_ignores_configured_colors(tmp_path: Path) -> None:
    """Ignore configured colors during colors_map extraction."""
    path = tmp_path / "colors.xlsx"
    _write_color_workbook(path)

    data = extract_sheet_colors_map(
        path,
        include_default_background=True,
        ignore_colors={"#ad3815", "00ff00"},
    )
    sheet = data.get_sheet("Sheet1")
    assert sheet is not None
    assert "AD3815" not in sheet.colors_map
    assert "00FF00" not in sheet.colors_map
    assert "FFFFFF" in sheet.colors_map
