from collections.abc import Callable, Iterator
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from importlib import util
import io
import json
import os
from pathlib import Path
from typing import TypeVar, cast

from openpyxl import Workbook
from pydantic import BaseModel
import pytest

from exstruct.cli.availability import ComAvailability
from exstruct.cli.main import build_parser, main as cli_main

F = TypeVar("F", bound=Callable[..., object])
render = cast(Callable[[F], F], pytest.mark.render)

_ALLOWED_CLI_FLAGS: set[str] = {
    "-f",
    "-o",
    "--format",
    "--image",
    "--mode",
    "--pdf",
    "--print-areas-dir",
}


class CliResult(BaseModel):
    """Result of running the CLI inside the test process."""

    returncode: int
    stdout: bytes | str
    stderr: bytes | str


def _toon_available() -> bool:
    try:
        import toon  # noqa: F401

        return True
    except Exception:
        return False


def _prepare_sample_excel(tmp_path: Path) -> Path:
    """Prepare a minimal Excel workbook for CLI tests.

    Args:
        tmp_path: Temporary directory provided by pytest for test artifacts.

    Returns:
        Path to a workbook copied from the repository sample when available or
        a generated fallback workbook created with openpyxl.
    """
    sample = Path("sample") / "sample.xlsx"
    dest = tmp_path / "sample.xlsx"
    if sample.exists():
        import shutil

        shutil.copy(sample, dest)
        return dest

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    wb.save(dest)
    wb.close()
    return dest


def _prepare_print_area_excel(tmp_path: Path) -> Path:
    """Prepare a workbook with a defined print area for CLI print-area tests.

    Args:
        tmp_path: Temporary directory provided by pytest for test artifacts.

    Returns:
        Path to the generated workbook that defines a print area on ``Sheet1``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.print_area = "A1:B2"
    dest = tmp_path / "print_area.xlsx"
    wb.save(dest)
    wb.close()
    return dest


def _prepare_unicode_excel(tmp_path: Path) -> Path:
    """Create a workbook containing varied Unicode characters for CLI tests.

    Args:
        tmp_path: Temporary directory provided by pytest for test artifacts.

    Returns:
        Path to the generated Excel workbook that includes mixed Unicode
        content such as Japanese text, check marks, and emoji.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "ユニコード"
    ws.append(["ラベル", "値"])
    ws.append(["チェック", "☑︎ テスト ✓ こんにちは 🌸"])
    dest = tmp_path / "unicode.xlsx"
    wb.save(dest)
    wb.close()
    return dest


def _run_cli(
    args: list[str], *, text: bool = True, env: dict[str, str] | None = None
) -> CliResult:
    """Execute the ExStruct CLI with a fixed command prefix.

    Args:
        args: Argument list appended after the module invocation.
        text: Whether to decode stdout/stderr as text.
        env: Optional environment variables overriding the current process.

    Returns:
        Result data captured from the in-process CLI execution.
    """

    safe_args = _sanitize_cli_args(args)
    stdout_buffer = io.StringIO()
    stderr_buffer = io.StringIO()
    with (
        _temporary_env(env),
        redirect_stdout(stdout_buffer),
        redirect_stderr(stderr_buffer),
    ):
        returncode = cli_main(argv=safe_args)

    stdout_text = stdout_buffer.getvalue()
    stderr_text = stderr_buffer.getvalue()
    if text:
        return CliResult(returncode=returncode, stdout=stdout_text, stderr=stderr_text)
    return CliResult(
        returncode=returncode,
        stdout=stdout_text.encode("utf-8"),
        stderr=stderr_text.encode("utf-8"),
    )


def _sanitize_cli_args(args: list[str]) -> list[str]:
    """Validate CLI arguments to mitigate command-injection risks in tests.

    Args:
        args: Argument list appended after the module invocation.

    Returns:
        Sanitized argument list safe for subprocess execution in tests.

    Raises:
        ValueError: If control characters or disallowed flags are present.
    """

    validated: list[str] = []
    for arg in args:
        _ensure_no_control_chars(arg)
        _ensure_allowed_cli_flag(arg)
        validated.append(arg)
    return validated


def _stdout_text(result: CliResult) -> str:
    """Return stdout as text for assertions.

    Args:
        result: Completed CLI result.

    Returns:
        Decoded stdout string or empty string when stdout is absent.
    """

    stdout = result.stdout
    if isinstance(stdout, bytes):
        return stdout.decode("utf-8", errors="replace")
    return stdout


def _stderr_text(result: CliResult) -> str:
    """Return stderr as text for assertions.

    Args:
        result: Completed CLI result.

    Returns:
        Decoded stderr string or empty string when stderr is absent.
    """

    stderr = result.stderr
    if isinstance(stderr, bytes):
        return stderr.decode("utf-8", errors="replace")
    return stderr


def _ensure_no_control_chars(arg: str) -> None:
    """Reject control characters in CLI arguments.

    Args:
        arg: CLI argument to validate.

    Raises:
        ValueError: If control characters are found.
    """

    if "\x00" in arg or "\n" in arg or "\r" in arg:
        msg = "CLI arguments must not contain control characters"
        raise ValueError(msg)


def _ensure_allowed_cli_flag(arg: str) -> None:
    """Ensure only known CLI flags are passed to the subprocess.

    Args:
        arg: CLI argument to validate.

    Raises:
        ValueError: If a flag is not in the allowlist.
    """

    if arg.startswith("-") and arg not in _ALLOWED_CLI_FLAGS:
        msg = f"CLI flag is not allowed in tests: {arg}"
        raise ValueError(msg)


@contextmanager
def _temporary_env(env: dict[str, str] | None) -> Iterator[None]:
    """Temporarily override environment variables for CLI execution.

    Args:
        env: Environment mapping to apply for the duration of the context.
    """

    if env is None:
        yield
        return
    original = os.environ.copy()
    os.environ.clear()
    os.environ.update(env)
    try:
        yield
    finally:
        os.environ.clear()
        os.environ.update(original)


def test_CLIでjson出力が成功する(tmp_path: Path) -> None:
    xlsx = _prepare_sample_excel(tmp_path)
    out_json = tmp_path / "out.json"
    result = _run_cli([str(xlsx), "-o", str(out_json)])
    assert result.returncode == 0
    assert out_json.exists()
    # stdout may be empty when writing to a file; ensure no errors surfaced
    assert "Error" not in _stdout_text(result)


def test_CLIでyamlやtoon指定は未サポート(tmp_path: Path) -> None:
    xlsx = _prepare_sample_excel(tmp_path)
    out_yaml = tmp_path / "out.yaml"
    result = _run_cli([str(xlsx), "-o", str(out_yaml), "-f", "yaml"])
    if util.find_spec("yaml") is not None:
        assert result.returncode == 0
        assert out_yaml.exists()
    else:
        assert result.returncode != 0
        assert "pyyaml" in _stdout_text(result) or "pyyaml" in _stderr_text(result)

    out_toon = tmp_path / "out.toon"
    result = _run_cli([str(xlsx), "-o", str(out_toon), "-f", "toon"])
    if _toon_available():
        assert result.returncode == 0
        assert out_toon.exists()
    else:
        assert result.returncode != 0
        assert "TOON export requires python-toon" in _stdout_text(result)


@render
def test_CLIでpdfと画像が出力される(tmp_path: Path) -> None:
    xlsx = _prepare_sample_excel(tmp_path)
    out_json = tmp_path / "out.json"
    result = _run_cli([str(xlsx), "-o", str(out_json), "--pdf", "--image"])
    assert result.returncode == 0
    pdf_path = out_json.with_suffix(".pdf")
    images_dir = out_json.parent / f"{out_json.stem}_images"
    assert pdf_path.exists()
    assert images_dir.exists()
    assert any(images_dir.glob("*.png"))


def test_CLIで無効ファイルは安全終了する(tmp_path: Path) -> None:
    bad_path = tmp_path / "nope.xlsx"
    out_json = tmp_path / "out.json"
    result = _run_cli([str(bad_path), "-o", str(out_json)])
    assert result.returncode == 0
    combined_output = _stdout_text(result) + _stderr_text(result)
    assert "not found" in combined_output.lower() or combined_output == ""


def test_CLI_print_areas_dir_outputs_files(tmp_path: Path) -> None:
    xlsx = _prepare_print_area_excel(tmp_path)
    areas_dir = tmp_path / "areas"
    result = _run_cli(
        [str(xlsx), "--print-areas-dir", str(areas_dir), "--mode", "standard"]
    )
    assert result.returncode == 0
    files = list(areas_dir.glob("*.json"))
    assert files, (
        "No print area files created. "
        f"stdout={_stdout_text(result)} stderr={_stderr_text(result)}"
    )


def test_cli_parser_includes_auto_page_breaks_option() -> None:
    """Ensure the auto page-breaks option is registered when COM is available."""
    availability = ComAvailability(available=True, reason=None)
    parser = build_parser(availability=availability)
    help_text = parser.format_help()
    assert "--auto-page-breaks-dir" in help_text


def test_cli_parser_excludes_auto_page_breaks_option() -> None:
    """Ensure the auto page-breaks option is hidden when COM is unavailable."""
    availability = ComAvailability(available=False, reason="disabled")
    parser = build_parser(availability=availability)
    help_text = parser.format_help()
    assert "--auto-page-breaks-dir" not in help_text


def test_CLI_stdout_is_utf8_with_cp932_env(tmp_path: Path) -> None:
    """Ensure stdout remains UTF-8 even when PYTHONIOENCODING forces cp932.

    Args:
        tmp_path: Temporary directory provided by pytest for test artifacts.
    """

    xlsx = _prepare_unicode_excel(tmp_path)
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "cp932"
    result = _run_cli([str(xlsx), "--format", "json"], text=False, env=env)

    assert result.returncode == 0

    stdout_text = _stdout_text(result)
    assert "☑︎ テスト ✓ こんにちは 🌸" in stdout_text
    json.loads(stdout_text)
