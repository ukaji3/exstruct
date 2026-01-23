from __future__ import annotations

from collections import deque
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import logging
from pathlib import Path
import re

import numpy as np
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import xlwings as xw

from ..models import CellRow
from .workbook import openpyxl_workbook

logger = logging.getLogger(__name__)
_warned_keys: set[str] = set()
XL_LINESTYLE_NONE = -4142
XL_INSIDE_VERTICAL = 11
XL_INSIDE_HORIZONTAL = 12
XL_EDGE_LEFT = 7
XL_EDGE_TOP = 8
XL_EDGE_BOTTOM = 9
XL_EDGE_RIGHT = 10
MatrixInput = Sequence[Sequence[object]] | Sequence[object]

# Detection tuning parameters (can be overridden via set_table_detection_params)
_DETECTION_CONFIG = {
    "table_score_threshold": 0.35,
    "density_min": 0.05,
    "coverage_min": 0.2,
    "min_nonempty_cells": 3,
}
_DEFAULT_BACKGROUND_HEX = "FFFFFF"
_XL_COLOR_NONE = -4142


# Use dataclasses for lightweight models
@dataclass(frozen=True)
class SheetColorsMap:
    """Background color map for a single worksheet."""

    sheet_name: str
    colors_map: dict[str, list[tuple[int, int]]]


@dataclass(frozen=True)
class WorkbookColorsMap:
    """Background color maps for all worksheets in a workbook."""

    sheets: dict[str, SheetColorsMap]

    def get_sheet(self, sheet_name: str) -> SheetColorsMap | None:
        """
        Retrieve the SheetColorsMap for a worksheet by name.

        Parameters:
            sheet_name (str): Name of the worksheet to retrieve.

        Returns:
            SheetColorsMap | None: The sheet's color map if present, `None` otherwise.
        """
        return self.sheets.get(sheet_name)


@dataclass(frozen=True)
class SheetFormulasMap:
    """Formula map for a single worksheet."""

    sheet_name: str
    formulas_map: dict[str, list[tuple[int, int]]]


@dataclass(frozen=True)
class WorkbookFormulasMap:
    """Formula maps for all worksheets in a workbook."""

    sheets: dict[str, SheetFormulasMap]

    def get_sheet(self, sheet_name: str) -> SheetFormulasMap | None:
        """
        Retrieve the formulas map for a worksheet.

        Parameters:
            sheet_name (str): Name of the worksheet to look up.

        Returns:
            SheetFormulasMap | None: The sheet's formulas map if present, `None` if the worksheet is not found.
        """
        return self.sheets.get(sheet_name)


@dataclass(frozen=True)
class MergedCellRange:
    """Merged cell range with normalized value."""

    r1: int
    c1: int
    r2: int
    c2: int
    v: str


def extract_sheet_colors_map(
    file_path: Path, *, include_default_background: bool, ignore_colors: set[str] | None
) -> WorkbookColorsMap:
    """Extract background colors for each worksheet.

    Args:
        file_path: Excel workbook path.
        include_default_background: Whether to include default (white) backgrounds
            within the used range.
        ignore_colors: Optional set of color keys to ignore.

    Returns:
        WorkbookColorsMap containing per-sheet color maps.
    """
    sheets: dict[str, SheetColorsMap] = {}
    with openpyxl_workbook(file_path, data_only=True, read_only=False) as wb:
        for ws in wb.worksheets:
            sheet_map = _extract_sheet_colors(
                ws, include_default_background, ignore_colors
            )
            sheets[ws.title] = sheet_map
    return WorkbookColorsMap(sheets=sheets)


def extract_sheet_formulas_map(file_path: Path) -> WorkbookFormulasMap:
    """
    Extract normalized formula strings from every worksheet in the workbook.

    Parameters:
        file_path (Path): Path to the Excel workbook to read.

    Returns:
        WorkbookFormulasMap: Mapping of sheet names to SheetFormulasMap objects. Each SheetFormulasMap contains a mapping from normalized formula strings (each beginning with "=") to a list of cell coordinates (row, column) where that formula occurs.
    """
    sheets: dict[str, SheetFormulasMap] = {}
    with openpyxl_workbook(file_path, data_only=False, read_only=False) as wb:
        for ws in wb.worksheets:
            sheet_map = _extract_sheet_formulas(ws)
            sheets[ws.title] = sheet_map
    return WorkbookFormulasMap(sheets=sheets)


def extract_sheet_formulas_map_com(workbook: xw.Book) -> WorkbookFormulasMap:
    """
    Collects and normalizes formulas from every worksheet in an xlwings workbook into per-sheet mappings.

    Parameters:
        workbook: xlwings Book instance whose sheets will be scanned for formulas.

    Returns:
        WorkbookFormulasMap: maps sheet names to SheetFormulasMap objects. Each SheetFormulasMap.formulas_map maps a normalized formula string (consistent representation, e.g., beginning with "=") to a list of (row, column) tuples representing cell locations using Excel 1-based indices.
    """
    sheets: dict[str, SheetFormulasMap] = {}
    for sheet in workbook.sheets:
        formulas_map: dict[str, list[tuple[int, int]]] = {}
        used = sheet.used_range
        start_row = int(getattr(used, "row", 1))
        start_col = int(getattr(used, "column", 1))
        max_row = used.last_cell.row
        max_col = used.last_cell.column
        if max_row <= 0 or max_col <= 0:
            sheets[sheet.name] = SheetFormulasMap(
                sheet_name=sheet.name, formulas_map=formulas_map
            )
            continue
        rng = sheet.range((start_row, start_col), (max_row, max_col))
        matrix = _normalize_matrix(rng.formula)
        for r_offset, row in enumerate(matrix):
            for c_offset, value in enumerate(row):
                normalized = _normalize_formula_from_com(value)
                if normalized is None:
                    continue
                row_index = start_row + r_offset
                col_index = start_col + c_offset - 1
                formulas_map.setdefault(normalized, []).append((row_index, col_index))
        sheets[sheet.name] = SheetFormulasMap(
            sheet_name=sheet.name, formulas_map=formulas_map
        )
    return WorkbookFormulasMap(sheets=sheets)


def extract_sheet_colors_map_com(
    workbook: xw.Book,
    *,
    include_default_background: bool,
    ignore_colors: set[str] | None,
) -> WorkbookColorsMap:
    """
    Extract per-sheet background color maps using the workbook's COM/display-format interfaces.

    Parameters:
        workbook (xw.Book): xlwings workbook whose sheets will be inspected.
        include_default_background (bool): If true, include default background colors (e.g., white) for cells inside each sheet's used range.
        ignore_colors (set[str] | None): Optional set of normalized color keys to exclude from results.

    Returns:
        WorkbookColorsMap: Mapping of sheet names to SheetColorsMap containing detected background color positions for each worksheet.
    """
    _prepare_workbook_for_display_format(workbook)
    sheets: dict[str, SheetColorsMap] = {}
    for sheet in workbook.sheets:
        _prepare_sheet_for_display_format(sheet)
        sheet_map = _extract_sheet_colors_com(
            sheet, include_default_background, ignore_colors
        )
        sheets[sheet.name] = sheet_map
    return WorkbookColorsMap(sheets=sheets)


def _extract_sheet_colors(
    ws: Worksheet, include_default_background: bool, ignore_colors: set[str] | None
) -> SheetColorsMap:
    """
    Extract the background color locations present on a single worksheet.

    Parameters:
        ws (Worksheet): Worksheet to scan.
        include_default_background (bool): If true, treat cells with the workbook default/background color as having a color key.
        ignore_colors (set[str] | None): Optional set of color keys to ignore (keys are normalized before comparison).

    Returns:
        SheetColorsMap: Mapping from normalized color key to a list of cell coordinates where that color appears. Coordinates are tuples (row, col) where `row` is 1-based and `col` is 0-based.
    """
    min_row, min_col, max_row, max_col = _get_used_range_bounds(ws)
    colors_map: dict[str, list[tuple[int, int]]] = {}
    if min_row > max_row or min_col > max_col:
        return SheetColorsMap(sheet_name=ws.title, colors_map=colors_map)

    ignore_set = _normalize_ignore_colors(ignore_colors)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            color_key = _resolve_cell_background(cell, include_default_background)
            if color_key is None:
                continue
            normalized_key = _normalize_color_key(color_key)
            if _should_ignore_color(normalized_key, ignore_set):
                continue
            colors_map.setdefault(normalized_key, []).append(
                (cell.row, cell.col_idx - 1)
            )
    return SheetColorsMap(sheet_name=ws.title, colors_map=colors_map)


def _extract_sheet_formulas(ws: Worksheet) -> SheetFormulasMap:
    """
    Collect normalized formula strings from a worksheet and group their cell coordinates.

    Parameters:
        ws (Worksheet): Worksheet to scan for formulas.

    Returns:
        SheetFormulasMap: container with the sheet's name and a mapping from each normalized formula string (prefixed with "=") to a list of cell coordinates as (row, zero-based-column).
    """
    min_row, min_col, max_row, max_col = _get_used_range_bounds(ws)
    formulas_map: dict[str, list[tuple[int, int]]] = {}
    if min_row > max_row or min_col > max_col:
        return SheetFormulasMap(sheet_name=ws.title, formulas_map=formulas_map)

    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if getattr(cell, "data_type", None) != "f":
                continue
            normalized = _normalize_formula_value(getattr(cell, "value", None))
            if normalized is None:
                continue
            formulas_map.setdefault(normalized, []).append((cell.row, cell.col_idx - 1))
    return SheetFormulasMap(sheet_name=ws.title, formulas_map=formulas_map)


def _normalize_formula_value(value: object) -> str | None:
    """Normalize a formula string for openpyxl cells.

    Args:
        value: Raw cell value.

    Returns:
        Formula string with leading "=", or None when empty.
    """
    if value is None:
        return None
    array_text = getattr(value, "text", None)
    if array_text is not None:
        text = str(array_text)
    else:
        text = str(value)
    if text == "":
        return None
    if not text.startswith("="):
        return f"={text}"
    return text


def _normalize_formula_from_com(value: object) -> str | None:
    """
    Normalize a COM-returned cell formula into a string that begins with '='.

    Parameters:
        value (object): Raw value returned from COM for a cell's formula.

    Returns:
        str | None: The input string if it is non-empty and starts with '=', `None` otherwise.
    """
    if value is None or not isinstance(value, str):
        return None
    text = value
    if text == "":
        return None
    if not text.startswith("="):
        return None
    return text


def _extract_sheet_colors_com(
    sheet: xw.Sheet, include_default_background: bool, ignore_colors: set[str] | None
) -> SheetColorsMap:
    """
    Extract per-sheet background color mapping using COM/DisplayFormat.

    Parameters:
        sheet (xw.Sheet): xlwings sheet object to inspect.
        include_default_background (bool): If True, include cells whose background is the workbook default color.
        ignore_colors (set[str] | None): Optional set of normalized color keys to exclude from the result.

    Returns:
        SheetColorsMap: Mapping from normalized color key (hex/theme/index canonical form) to a list of cell coordinates where that color appears. Each coordinate is a tuple (row, col) where `row` is the worksheet row number (1-based) and `col` is the zero-based column index.
    """
    colors_map: dict[str, list[tuple[int, int]]] = {}
    used = sheet.used_range
    start_row = int(getattr(used, "row", 1))
    start_col = int(getattr(used, "column", 1))
    max_row = used.last_cell.row
    max_col = used.last_cell.column
    if max_row <= 0 or max_col <= 0:
        return SheetColorsMap(sheet_name=sheet.name, colors_map=colors_map)

    ignore_set = _normalize_ignore_colors(ignore_colors)
    for row in range(start_row, max_row + 1):
        for col in range(start_col, max_col + 1):
            color_key = _resolve_cell_background_com(
                sheet, row, col, include_default_background
            )
            if color_key is None:
                continue
            normalized_key = _normalize_color_key(color_key)
            if _should_ignore_color(normalized_key, ignore_set):
                continue
            colors_map.setdefault(normalized_key, []).append((row, col - 1))
    return SheetColorsMap(sheet_name=sheet.name, colors_map=colors_map)


def _get_used_range_bounds(ws: Worksheet) -> tuple[int, int, int, int]:
    """Return used range bounds for a worksheet.

    Args:
        ws: Target worksheet.

    Returns:
        Tuple of (min_row, min_col, max_row, max_col).
    """
    try:
        if _is_effectively_empty_sheet(ws):
            return 1, 1, 0, 0
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        return min_row, min_col, max_row, max_col
    except Exception:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return 1, 1, 0, 0
        return 1, 1, max_row, max_col


def _is_effectively_empty_sheet(ws: Worksheet) -> bool:
    """Check whether a worksheet has no content or styling.

    Args:
        ws: Target worksheet.

    Returns:
        True if the sheet has no meaningful content or style, otherwise False.
    """
    if ws.max_row != 1 or ws.max_column != 1:
        return False
    cell = ws.cell(row=1, column=1)
    return cell.value is None and not cell.has_style


def _resolve_cell_background(
    cell: object, include_default_background: bool
) -> str | None:
    """Resolve a cell's background color key.

    Args:
        cell: Worksheet cell object.
        include_default_background: Whether to treat default fills as white.

    Returns:
        Normalized color key or None when excluded.
    """
    fill = getattr(cell, "fill", None)
    if fill is None:
        return _DEFAULT_BACKGROUND_HEX if include_default_background else None
    pattern_type = getattr(fill, "patternType", None)
    if pattern_type in (None, "none"):
        return _DEFAULT_BACKGROUND_HEX if include_default_background else None
    color_key = _resolve_fill_color_key(fill)
    if color_key == _DEFAULT_BACKGROUND_HEX and not include_default_background:
        return None
    return color_key


def _resolve_fill_color_key(fill: object) -> str | None:
    """Normalize the foreground/background color of a fill.

    Args:
        fill: openpyxl fill object.

    Returns:
        Normalized color key or None when unavailable.
    """
    fg_color = getattr(fill, "fgColor", None)
    if fg_color is not None:
        fg_key = _color_to_key(fg_color)
        if fg_key is not None:
            return fg_key
    bg_color = getattr(fill, "bgColor", None)
    return _color_to_key(bg_color) if bg_color is not None else None


def _resolve_cell_background_com(
    sheet: xw.Sheet, row: int, col: int, include_default_background: bool
) -> str | None:
    """Resolve a cell's background color key via COM display format.

    Args:
        sheet: Target worksheet.
        row: 1-based row index.
        col: 1-based column index.
        include_default_background: Whether to include default (white) backgrounds.

    Returns:
        Normalized color key or None when excluded.
    """
    color_value = _get_display_format_color(sheet, row, col)
    if color_value is None:
        return _DEFAULT_BACKGROUND_HEX if include_default_background else None
    if color_value == _XL_COLOR_NONE:
        return _DEFAULT_BACKGROUND_HEX if include_default_background else None
    color_key = _excel_color_int_to_rgb_hex(color_value)
    if color_key == _DEFAULT_BACKGROUND_HEX and not include_default_background:
        return None
    return color_key


def _prepare_workbook_for_display_format(workbook: xw.Book) -> None:
    """Prepare a workbook so DisplayFormat reflects conditional formatting.

    Args:
        workbook: xlwings workbook instance.
    """
    try:
        # Force calculation to ensure DisplayFormat.Interior reflects conditional formatting rules
        workbook.app.calculate()
    except Exception:
        return


def _prepare_sheet_for_display_format(sheet: xw.Sheet) -> None:
    """Prepare a sheet so DisplayFormat reflects conditional formatting.

    Args:
        sheet: Target worksheet.
    """
    try:
        # Activate sheet so DisplayFormat is available
        sheet.api.Activate()
    except Exception:
        return
    try:
        # Calculate to apply conditional formatting to DisplayFormat
        sheet.api.Calculate()
    except Exception:
        return


def _get_display_format_color(sheet: xw.Sheet, row: int, col: int) -> int | None:
    """Read DisplayFormat.Interior.Color from COM.

    Args:
        sheet: Target worksheet.
        row: 1-based row index.
        col: 1-based column index.

    Returns:
        BGR integer color or None if unavailable.
    """
    try:
        cell = sheet.api.Cells(row, col)
        display_format = cell.DisplayFormat
        interior = display_format.Interior
        return int(interior.Color)
    except Exception:
        return None


def _excel_color_int_to_rgb_hex(color_value: int) -> str:
    """Convert an Excel color integer into an RGB hex string.

    Args:
        color_value: Excel color integer from COM.

    Returns:
        RGB hex string (uppercase).
    """
    red = color_value & 0xFF
    green = (color_value >> 8) & 0xFF
    blue = (color_value >> 16) & 0xFF
    return f"{red:02X}{green:02X}{blue:02X}"


def _normalize_color_key(color_key: str) -> str:
    """Normalize a color key into a canonical representation.

    Args:
        color_key: Raw color key (hex or themed/indexed).

    Returns:
        Normalized color key.
    """
    trimmed = color_key.strip()
    if not trimmed:
        return ""
    lowered = trimmed.lower()
    if lowered.startswith(("theme:", "indexed:", "auto:")) or lowered == "auto":
        return lowered
    hex_key = trimmed.lstrip("#").upper()
    if len(hex_key) == 8:
        hex_key = hex_key[2:]
    return hex_key


def _normalize_ignore_colors(ignore_colors: set[str] | None) -> set[str]:
    """Normalize ignore color keys.

    Args:
        ignore_colors: Optional set of color keys to ignore.

    Returns:
        Normalized set of color keys.
    """
    if not ignore_colors:
        return set()
    normalized = {_normalize_color_key(color) for color in ignore_colors}
    return {color for color in normalized if color}


def _should_ignore_color(color_key: str, ignore_colors: set[str]) -> bool:
    """Check whether a color key should be ignored.

    Args:
        color_key: Normalized color key.
        ignore_colors: Normalized ignore color set.

    Returns:
        True when the color key is ignored.
    """
    return color_key in ignore_colors


def _color_to_key(color: Color | object) -> str | None:
    """Convert an openpyxl color object into a normalized key.

    Args:
        color: openpyxl color object.

    Returns:
        Normalized color key string or None when unavailable.
    """
    rgb = getattr(color, "rgb", None)
    if rgb:
        return _normalize_rgb(str(rgb))
    color_type = getattr(color, "type", None)
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        tint = getattr(color, "tint", None)
        theme_id = "unknown" if theme is None else str(theme)
        if tint is None:
            return f"theme:{theme_id}"
        return f"theme:{theme_id}:{tint}"
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if indexed is not None:
            return f"indexed:{indexed}"
    if color_type == "auto":
        auto = getattr(color, "auto", None)
        return "auto" if auto is None else f"auto:{auto}"
    return None


def _normalize_rgb(rgb: str) -> str:
    """Normalize an RGB/ARGB string into 6-hex format.

    Args:
        rgb: Raw RGB/ARGB string from openpyxl.

    Returns:
        Normalized RGB hex string (uppercase, 6 chars when possible).
    """
    cleaned = rgb.strip().upper()
    if cleaned.startswith("0X"):
        cleaned = cleaned[2:]
    if len(cleaned) == 8:
        cleaned = cleaned[2:]
    return cleaned


def warn_once(key: str, message: str) -> None:
    if key not in _warned_keys:
        logger.warning(message)
        _warned_keys.add(key)


def extract_sheet_cells(file_path: Path) -> dict[str, list[CellRow]]:
    """Read all sheets via pandas and convert to CellRow list while skipping empty cells."""
    dfs = pd.read_excel(file_path, header=None, sheet_name=None, dtype=str)
    result: dict[str, list[CellRow]] = {}
    for sheet_name, df in dfs.items():
        df = df.fillna("")
        rows: list[CellRow] = []
        for excel_row, row in enumerate(df.itertuples(index=False, name=None), start=1):
            filtered: dict[str, int | float | str] = {}
            for j, v in enumerate(row):
                s = "" if v is None else str(v)
                if s.strip() == "":
                    continue
                filtered[str(j)] = _coerce_numeric_preserve_format(s)
            if not filtered:
                continue
            rows.append(CellRow(r=excel_row, c=filtered))
        result[sheet_name] = rows
    return result


def extract_sheet_cells_with_links(file_path: Path) -> dict[str, list[CellRow]]:
    """
    Extract cells and hyperlinks per sheet.

    Returns:
        {sheet_name: [CellRow(r=..., c=..., links={"col_index": url, ...}), ...]}

    Notes:
        - Uses pandas extraction for values (same filtering as extract_sheet_cells).
        - Collects hyperlinks via openpyxl (requires read_only=False because border maps/hyperlinks need full objects).
        - Links are mapped by column index string (e.g., "0") to hyperlink.target.
    """
    cell_rows = extract_sheet_cells(file_path)
    links_by_sheet: dict[str, dict[int, dict[str, str]]] = {}
    with openpyxl_workbook(file_path, data_only=True, read_only=False) as wb:
        for ws in wb.worksheets:
            sheet_links: dict[int, dict[str, str]] = {}
            for row in ws.iter_rows():
                for cell in row:
                    link = getattr(cell, "hyperlink", None)
                    target = getattr(link, "target", None) if link else None
                    if not target:
                        continue
                    col_str = str(
                        cell.col_idx - 1
                    )  # zero-based to align with extract_sheet_cells
                    sheet_links.setdefault(cell.row, {})[col_str] = target
            links_by_sheet[ws.title] = sheet_links

    merged: dict[str, list[CellRow]] = {}
    for sheet_name, rows in cell_rows.items():
        sheet_links = links_by_sheet.get(sheet_name, {})
        merged_rows: list[CellRow] = []
        for row in rows:
            links = sheet_links.get(row.r, {})
            merged_rows.append(CellRow(r=row.r, c=row.c, links=links or None))
        merged[sheet_name] = merged_rows
    wb.close()
    return merged


def extract_sheet_merged_cells(file_path: Path) -> dict[str, list[MergedCellRange]]:
    """Extract merged cell ranges per sheet via openpyxl.

    Args:
        file_path: Excel workbook path.

    Returns:
        Mapping of sheet name to merged cell ranges.
    """
    merged_by_sheet: dict[str, list[MergedCellRange]] = {}
    with openpyxl_workbook(file_path, data_only=True, read_only=False) as wb:
        for ws in wb.worksheets:
            merged_ranges = getattr(ws, "merged_cells", None)
            if merged_ranges is None:
                merged_by_sheet[ws.title] = []
                continue
            results: list[MergedCellRange] = []
            for merged_range in getattr(merged_ranges, "ranges", []):
                bounds = range_boundaries(str(merged_range))
                min_col, min_row, max_col, max_row = bounds
                cell_value = ws.cell(row=min_row, column=min_col).value
                value_str = "" if cell_value is None else str(cell_value)
                if value_str == "":
                    value_str = " "
                results.append(
                    MergedCellRange(
                        r1=min_row,
                        c1=min_col - 1,
                        r2=max_row,
                        c2=max_col - 1,
                        v=value_str,
                    )
                )
            merged_by_sheet[ws.title] = results
    return merged_by_sheet


def shrink_to_content(  # noqa: C901
    sheet: xw.Sheet,
    top: int,
    left: int,
    bottom: int,
    right: int,
    require_inside_border: bool = False,
    min_nonempty_ratio: float = 0.0,
) -> tuple[int, int, int, int]:
    """Trim a rectangle based on cell contents and optional border heuristics."""
    rng = sheet.range((top, left), (bottom, right))
    vals = rng.value
    if vals is None:
        vals = []
    if not isinstance(vals, list):
        vals = [[vals]]
    elif vals and not isinstance(vals[0], list):
        vals = [vals]
    rows_n = len(vals)
    cols_n = len(vals[0]) if rows_n else 0

    def to_str(x: object) -> str:
        return "" if x is None else str(x)

    def is_empty_value(x: object) -> bool:
        return to_str(x).strip() == ""

    def row_empty(i: int) -> bool:
        return cols_n == 0 or all(is_empty_value(vals[i][j]) for j in range(cols_n))

    def col_empty(j: int) -> bool:
        return rows_n == 0 or all(is_empty_value(vals[i][j]) for i in range(rows_n))

    def row_nonempty_ratio(i: int) -> float:
        if cols_n == 0:
            return 0.0
        cnt = sum(1 for j in range(cols_n) if not is_empty_value(vals[i][j]))
        return cnt / cols_n

    def col_nonempty_ratio(j: int) -> float:
        if rows_n == 0:
            return 0.0
        cnt = sum(1 for i in range(rows_n) if not is_empty_value(vals[i][j]))
        return cnt / rows_n

    def column_has_inside_border(col_idx: int) -> bool:
        if not require_inside_border:
            return False
        try:
            for r in range(top, bottom + 1):
                ls = (
                    sheet.api.Cells(r, left + col_idx)
                    .Borders(XL_INSIDE_VERTICAL)
                    .LineStyle
                )
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    return True
        except Exception:
            pass
        return False

    def row_has_inside_border(row_idx: int) -> bool:
        if not require_inside_border:
            return False
        try:
            for c in range(left, right + 1):
                ls = (
                    sheet.api.Cells(top + row_idx, c)
                    .Borders(XL_INSIDE_HORIZONTAL)
                    .LineStyle
                )
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    return True
        except Exception:
            pass
        return False

    def should_trim_col(j: int) -> bool:
        if col_empty(j):
            return True
        if require_inside_border and not column_has_inside_border(j):
            return True
        if min_nonempty_ratio > 0.0 and col_nonempty_ratio(j) < min_nonempty_ratio:
            return True
        return False

    def should_trim_row(i: int) -> bool:
        if row_empty(i):
            return True
        if require_inside_border and not row_has_inside_border(i):
            return True
        if min_nonempty_ratio > 0.0 and row_nonempty_ratio(i) < min_nonempty_ratio:
            return True
        return False

    while left <= right and cols_n > 0:
        if should_trim_col(0):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(0)
            cols_n = len(vals[0]) if rows_n else 0
            left += 1
        else:
            break
    while top <= bottom and rows_n > 0:
        if should_trim_row(0):
            vals.pop(0)
            rows_n = len(vals)
            top += 1
        else:
            break
    while left <= right and cols_n > 0:
        if should_trim_col(cols_n - 1):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(cols_n - 1)
            cols_n = len(vals[0]) if rows_n else 0
            right -= 1
        else:
            break
    while top <= bottom and rows_n > 0:
        if should_trim_row(rows_n - 1):
            vals.pop(rows_n - 1)
            rows_n = len(vals)
            bottom -= 1
        else:
            break
    return top, left, bottom, right


def load_border_maps_xlsx(  # noqa: C901
    xlsx_path: Path, sheet_name: str
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, int, int]:
    with openpyxl_workbook(xlsx_path, data_only=True, read_only=False) as wb:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in {xlsx_path}")

        ws = wb[sheet_name]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(
                ws.calculate_dimension()
            )
        except Exception:
            min_col, min_row, max_col, max_row = (
                1,
                1,
                ws.max_column or 1,
                ws.max_row or 1,
            )

    shape = (max_row + 1, max_col + 1)
    has_border = np.zeros(shape, dtype=bool)
    top_edge = np.zeros(shape, dtype=bool)
    bottom_edge = np.zeros(shape, dtype=bool)
    left_edge = np.zeros(shape, dtype=bool)
    right_edge = np.zeros(shape, dtype=bool)

    def edge_has_style(edge: object) -> bool:
        if edge is None:
            return False
        style = getattr(edge, "style", None)
        return style is not None and style != "none"

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            b = getattr(cell, "border", None)
            if b is None:
                continue

            t = edge_has_style(b.top)
            btm = edge_has_style(b.bottom)
            left_border = edge_has_style(b.left)
            rgt = edge_has_style(b.right)

            if t or btm or left_border or rgt:
                has_border[r, c] = True
                if t:
                    top_edge[r, c] = True
                if btm:
                    bottom_edge[r, c] = True
                if left_border:
                    left_edge[r, c] = True
                if rgt:
                    right_edge[r, c] = True

    return has_border, top_edge, bottom_edge, left_edge, right_edge, max_row, max_col


def _detect_border_clusters_numpy(
    has_border: np.ndarray, min_size: int
) -> list[tuple[int, int, int, int]]:
    from scipy.ndimage import label

    structure = np.array([[0, 1, 0], [1, 1, 1], [0, 1, 0]], dtype=np.uint8)
    lbl, num = label(has_border.astype(np.uint8), structure=structure)
    rects: list[tuple[int, int, int, int]] = []
    for k in range(1, int(num) + 1):
        ys, xs = np.where(lbl == k)
        if int(len(ys)) < min_size:
            continue
        rects.append((int(ys.min()), int(xs.min()), int(ys.max()), int(xs.max())))
    return rects


def _detect_border_clusters_python(
    has_border: np.ndarray, min_size: int
) -> list[tuple[int, int, int, int]]:
    h, w = has_border.shape
    visited = np.zeros_like(has_border, dtype=bool)
    rects: list[tuple[int, int, int, int]] = []
    for r in range(h):
        for c in range(w):
            if not has_border[r, c] or visited[r, c]:
                continue
            q = deque([(r, c)])
            visited[r, c] = True
            ys = [r]
            xs = [c]
            while q:
                yy, xx = q.popleft()
                for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    ny, nx = yy + dy, xx + dx
                    if (
                        0 <= ny < h
                        and 0 <= nx < w
                        and has_border[ny, nx]
                        and not visited[ny, nx]
                    ):
                        visited[ny, nx] = True
                        q.append((ny, nx))
                        ys.append(ny)
                        xs.append(nx)
            if len(ys) >= min_size:
                rects.append((min(ys), min(xs), max(ys), max(xs)))
    return rects


def detect_border_clusters(
    has_border: np.ndarray, min_size: int = 4
) -> list[tuple[int, int, int, int]]:
    try:
        return _detect_border_clusters_numpy(has_border, min_size)
    except Exception:
        warn_once(
            "scipy-missing",
            "scipy is not available. Falling back to pure-Python BFS for connected components, which may be significantly slower.",
        )
        return _detect_border_clusters_python(has_border, min_size)


def _get_values_block(
    ws: Worksheet, top: int, left: int, bottom: int, right: int
) -> list[list[object]]:
    vals: list[list[object]] = []
    for row in ws.iter_rows(
        min_row=top, max_row=bottom, min_col=left, max_col=right, values_only=True
    ):
        vals.append(list(row))
    return vals


def _ensure_matrix(matrix: MatrixInput) -> list[list[object]]:
    rows_seq = list(matrix)
    if not rows_seq:
        return []
    first = rows_seq[0]
    if isinstance(first, Sequence) and not isinstance(first, str | bytes | bytearray):
        normalized: list[list[object]] = []
        for row in rows_seq:
            if isinstance(row, Sequence) and not isinstance(
                row, str | bytes | bytearray
            ):
                normalized.append(list(row))
            else:
                normalized.append([row])
        return normalized
    return [list(rows_seq)]


def _table_density_metrics(matrix: MatrixInput) -> tuple[float, float]:
    """
    Given a 2D matrix (list of rows), return (density, coverage).
    density: nonempty / total cells.
    coverage: area of tight bounding box of nonempty cells divided by total area.
    """
    normalized = _ensure_matrix(matrix)
    if not normalized:
        return 0.0, 0.0
    rows = len(normalized)
    cols = len(normalized[0]) if rows else 0
    if rows == 0 or cols == 0:
        return 0.0, 0.0

    nonempty_coords = []
    for i, row in enumerate(normalized):
        for j, v in enumerate(row):
            if not (v is None or str(v).strip() == ""):
                nonempty_coords.append((i, j))

    total = rows * cols
    if not nonempty_coords:
        return 0.0, 0.0

    nonempty = len(nonempty_coords)
    density = nonempty / total

    ys = [p[0] for p in nonempty_coords]
    xs = [p[1] for p in nonempty_coords]
    bbox_h = max(ys) - min(ys) + 1
    bbox_w = max(xs) - min(xs) + 1
    coverage = (bbox_h * bbox_w) / total if total > 0 else 0.0
    return density, coverage


def _is_plausible_table(matrix: MatrixInput) -> bool:
    """
    Heuristic: require at least 2 rows and 2 cols with meaningful data.
    - At least 2 rows have 2 以上の非空セル
    - At least 2 columns have 2 以上の非空セル
    """
    normalized = _ensure_matrix(matrix)
    if not normalized:
        return False

    rows = len(normalized)
    cols = (
        max((len(r) if isinstance(r, list) else 1) for r in normalized) if rows else 0
    )
    if rows < 2 or cols < 2:
        return False

    row_counts: list[int] = []
    col_counts = [0] * cols
    for r in normalized:
        cnt = 0
        for j in range(cols):
            v = r[j] if j < len(r) else None
            if not (v is None or str(v).strip() == ""):
                cnt += 1
                col_counts[j] += 1
        row_counts.append(cnt)

    rows_with_two = sum(1 for c in row_counts if c >= 2)
    cols_with_two = sum(1 for c in col_counts if c >= 2)
    return rows_with_two >= 2 and cols_with_two >= 2


def _nonempty_clusters(
    matrix: Sequence[Sequence[object]],
) -> list[tuple[int, int, int, int]]:
    """Return bounding boxes of connected components of nonempty cells (4-neighbor)."""
    if not matrix:
        return []
    rows = len(matrix)
    cols = max(len(r) for r in matrix) if rows else 0
    grid = [[False] * cols for _ in range(rows)]
    for i, row in enumerate(matrix):
        for j in range(cols):
            v = row[j] if j < len(row) else None
            if not (v is None or str(v).strip() == ""):
                grid[i][j] = True
    visited = [[False] * cols for _ in range(rows)]
    boxes: list[tuple[int, int, int, int]] = []

    def bfs(sr: int, sc: int) -> tuple[int, int, int, int]:
        q = deque([(sr, sc)])
        visited[sr][sc] = True
        ys = [sr]
        xs = [sc]
        while q:
            r, c = q.popleft()
            for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                nr, nc = r + dr, c + dc
                if (
                    0 <= nr < rows
                    and 0 <= nc < cols
                    and grid[nr][nc]
                    and not visited[nr][nc]
                ):
                    visited[nr][nc] = True
                    q.append((nr, nc))
                    ys.append(nr)
                    xs.append(nc)
        return min(ys), min(xs), max(ys), max(xs)

    for i in range(rows):
        for j in range(cols):
            if grid[i][j] and not visited[i][j]:
                boxes.append(bfs(i, j))
    return boxes


def _normalize_matrix(matrix: object) -> list[list[object]]:
    if matrix is None:
        return []
    if isinstance(matrix, list):
        return _ensure_matrix(matrix)
    if isinstance(matrix, Sequence) and not isinstance(matrix, str | bytes | bytearray):
        return _ensure_matrix(matrix)
    return [[matrix]]


def _header_like_row(row: list[object]) -> bool:
    nonempty = [v for v in row if not (v is None or str(v).strip() == "")]
    if len(nonempty) < 2:
        return False
    str_like = 0
    num_like = 0
    for v in nonempty:
        s = str(v)
        if _INT_RE.match(s) or _FLOAT_RE.match(s):
            num_like += 1
        else:
            str_like += 1
    return str_like >= num_like and str_like >= 1


def _table_signal_score(matrix: Sequence[Sequence[object]]) -> float:
    normalized = _ensure_matrix(matrix)
    density, coverage = _table_density_metrics(normalized)
    header = any(_header_like_row(r) for r in normalized[:2])  # check first 2 rows

    rows = len(normalized)
    cols = (
        max((len(r) if isinstance(r, list) else 1) for r in normalized) if rows else 0
    )
    row_counts: list[int] = []
    col_counts = [0] * cols if cols else []
    for r in normalized:
        cnt = 0
        for j in range(cols):
            v = r[j] if j < len(r) else None
            if not (v is None or str(v).strip() == ""):
                cnt += 1
                if j < len(col_counts):
                    col_counts[j] += 1
        row_counts.append(cnt)
    rows_with_two = sum(1 for c in row_counts if c >= 2)
    cols_with_two = sum(1 for c in col_counts if c >= 2)
    structure_score = 0.1 if (rows_with_two >= 2 and cols_with_two >= 2) else 0.0

    score = density
    if header:
        score += 0.2
    if coverage > 0.5:
        score += 0.1
    score += structure_score
    return score


def set_table_detection_params(
    *,
    table_score_threshold: float | None = None,
    density_min: float | None = None,
    coverage_min: float | None = None,
    min_nonempty_cells: int | None = None,
) -> None:
    """
    Configure table detection heuristics at runtime.
    Any parameter left as None keeps its current value.
    """
    if table_score_threshold is not None:
        _DETECTION_CONFIG["table_score_threshold"] = table_score_threshold
    if density_min is not None:
        _DETECTION_CONFIG["density_min"] = density_min
    if coverage_min is not None:
        _DETECTION_CONFIG["coverage_min"] = coverage_min
    if min_nonempty_cells is not None:
        _DETECTION_CONFIG["min_nonempty_cells"] = min_nonempty_cells


def shrink_to_content_openpyxl(  # noqa: C901
    ws: Worksheet,
    top: int,
    left: int,
    bottom: int,
    right: int,
    require_inside_border: bool,
    top_edge: np.ndarray,
    bottom_edge: np.ndarray,
    left_edge: np.ndarray,
    right_edge: np.ndarray,
    min_nonempty_ratio: float = 0.0,
) -> tuple[int, int, int, int]:
    vals = _get_values_block(ws, top, left, bottom, right)
    rows_n = bottom - top + 1
    cols_n = right - left + 1

    def to_str(x: object) -> str:
        return "" if x is None else str(x)

    def is_empty_value(x: object) -> bool:
        return to_str(x).strip() == ""

    def row_nonempty_ratio_local(i: int) -> float:
        if cols_n <= 0:
            return 0.0
        row = vals[i]
        cnt = sum(1 for v in row if not is_empty_value(v))
        return cnt / cols_n

    def col_nonempty_ratio_local(j: int) -> float:
        if rows_n <= 0:
            return 0.0
        cnt = 0
        for i in range(rows_n):
            if not is_empty_value(vals[i][j]):
                cnt += 1
        return cnt / rows_n

    def col_has_inside_border(j_abs: int) -> bool:
        if not require_inside_border:
            return False
        count_pairs = 0
        for r_abs in range(top, bottom + 1):
            if (
                j_abs > left
                and right_edge[r_abs, j_abs - 1]
                and left_edge[r_abs, j_abs]
            ):
                count_pairs += 1
        return count_pairs > 0

    def row_has_inside_border(i_abs: int) -> bool:
        if not require_inside_border:
            return False
        count_pairs = 0
        for c_abs in range(left, right + 1):
            if i_abs > top and bottom_edge[i_abs - 1, c_abs] and top_edge[i_abs, c_abs]:
                count_pairs += 1
        return count_pairs > 0

    while left <= right and cols_n > 0:
        empty_col = all(
            not (
                top_edge[i, left]
                or bottom_edge[i, left]
                or left_edge[i, left]
                or right_edge[i, left]
            )
            for i in range(top, bottom + 1)
        )
        if (
            empty_col
            or (require_inside_border and not col_has_inside_border(left))
            or (
                min_nonempty_ratio > 0.0
                and col_nonempty_ratio_local(0) < min_nonempty_ratio
            )
        ):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(0)
            cols_n -= 1
            left += 1
        else:
            break
    while top <= bottom and rows_n > 0:
        empty_row = all(
            not (
                top_edge[top, j]
                or bottom_edge[top, j]
                or left_edge[top, j]
                or right_edge[top, j]
            )
            for j in range(left, right + 1)
        )
        if (
            empty_row
            or (require_inside_border and not row_has_inside_border(top))
            or (
                min_nonempty_ratio > 0.0
                and row_nonempty_ratio_local(0) < min_nonempty_ratio
            )
        ):
            vals.pop(0)
            rows_n -= 1
            top += 1
        else:
            break
    while left <= right and cols_n > 0:
        empty_col = all(
            not (
                top_edge[i, right]
                or bottom_edge[i, right]
                or left_edge[i, right]
                or right_edge[i, right]
            )
            for i in range(top, bottom + 1)
        )
        if (
            empty_col
            or (require_inside_border and not col_has_inside_border(right))
            or (
                min_nonempty_ratio > 0.0
                and col_nonempty_ratio_local(cols_n - 1) < min_nonempty_ratio
            )
        ):
            for i in range(rows_n):
                if cols_n > 0:
                    vals[i].pop(cols_n - 1)
            cols_n -= 1
            right -= 1
        else:
            break
    while top <= bottom and rows_n > 0:
        empty_row = all(
            not (
                top_edge[bottom, j]
                or bottom_edge[bottom, j]
                or left_edge[bottom, j]
                or right_edge[bottom, j]
            )
            for j in range(left, right + 1)
        )
        if (
            empty_row
            or (require_inside_border and not row_has_inside_border(bottom))
            or (
                min_nonempty_ratio > 0.0
                and row_nonempty_ratio_local(rows_n - 1) < min_nonempty_ratio
            )
        ):
            vals.pop(rows_n - 1)
            rows_n -= 1
            bottom -= 1
        else:
            break
    return top, left, bottom, right


def _extract_listobject_tables(sheet: xw.Sheet) -> list[str]:
    """Extract table ranges from Excel ListObjects via COM.

    Args:
        sheet: xlwings worksheet.

    Returns:
        List of table ranges as Excel A1 strings.
    """
    tables: list[str] = []
    try:
        for lo in sheet.api.ListObjects:
            rng = lo.Range
            addr = rng.Address(RowAbsolute=False, ColumnAbsolute=False)
            tables.append(addr)
    except Exception:
        pass
    return tables


def _detect_border_rectangles_xlwings(
    sheet: xw.Sheet,
) -> list[tuple[int, int, int, int]]:
    """Detect bordered rectangles in a sheet using COM border inspection.

    Args:
        sheet: xlwings worksheet.

    Returns:
        List of rectangles as (top_row, left_col, bottom_row, right_col).
    """
    used = sheet.used_range
    max_row = used.last_cell.row
    max_col = used.last_cell.column

    def cell_has_any_border(r: int, c: int) -> bool:
        try:
            b = sheet.api.Cells(r, c).Borders
            for idx in (
                XL_EDGE_LEFT,
                XL_EDGE_TOP,
                XL_EDGE_RIGHT,
                XL_EDGE_BOTTOM,
                XL_INSIDE_VERTICAL,
                XL_INSIDE_HORIZONTAL,
            ):
                ls = b(idx).LineStyle
                if ls is not None and ls != XL_LINESTYLE_NONE:
                    try:
                        if getattr(b(idx), "Weight", 0) == 0:
                            continue
                    except Exception:
                        pass
                    return True
            return False
        except Exception:
            return False

    grid = [[False] * (max_col + 1) for _ in range(max_row + 1)]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if cell_has_any_border(r, c):
                grid[r][c] = True
    return _detect_border_rectangles(grid, min_size=4)


def _detect_border_rectangles(
    has_border: np.ndarray | Sequence[Sequence[bool]], *, min_size: int
) -> list[tuple[int, int, int, int]]:
    """Detect border rectangles from a boolean grid.

    Args:
        has_border: Boolean grid of border presence.
        min_size: Minimum cluster size to keep.

    Returns:
        List of rectangles as (top_row, left_col, bottom_row, right_col).
    """
    return detect_border_clusters(np.asarray(has_border, dtype=bool), min_size=min_size)


def _merge_rectangles(
    rects: Sequence[tuple[int, int, int, int]],
) -> list[tuple[int, int, int, int]]:
    """Merge overlapping rectangles while preserving contained regions.

    Args:
        rects: Sequence of rectangles (top, left, bottom, right).

    Returns:
        Merged rectangles sorted by coordinates.
    """
    merged_rects: list[tuple[int, int, int, int]] = []
    for rect in sorted(rects):
        merged = False
        for i, existing in enumerate(merged_rects):
            if _rectangles_overlap_for_merge(rect, existing):
                merged_rects[i] = (
                    min(rect[0], existing[0]),
                    min(rect[1], existing[1]),
                    max(rect[2], existing[2]),
                    max(rect[3], existing[3]),
                )
                merged = True
                break
        if not merged:
            merged_rects.append(rect)
    return merged_rects


def _rectangles_overlap_for_merge(
    a: tuple[int, int, int, int], b: tuple[int, int, int, int]
) -> bool:
    """Return True when rectangles should be merged.

    Args:
        a: First rectangle (top, left, bottom, right).
        b: Second rectangle (top, left, bottom, right).

    Returns:
        True if rectangles overlap and neither fully contains the other.
    """
    contains = (a[0] <= b[0] and a[1] <= b[1] and a[2] >= b[2] and a[3] >= b[3]) or (
        b[0] <= a[0] and b[1] <= a[1] and b[2] >= a[2] and b[3] >= a[3]
    )
    if contains:
        return False
    return not (a[1] > b[3] or a[3] < b[1] or a[0] > b[2] or a[2] < b[0])


def _collect_table_candidates_from_values(
    values: Sequence[Sequence[object]],
    *,
    base_top: int,
    base_left: int,
    col_name: Callable[[int], str],
) -> list[str]:
    """Collect table candidates from a normalized value matrix.

    Args:
        values: Normalized matrix of cell values.
        base_top: Top row index of the matrix in worksheet coordinates (1-based).
        base_left: Left column index of the matrix in worksheet coordinates (1-based).
        col_name: Function to convert column index to Excel letters.

    Returns:
        List of detected table candidate range strings.
    """
    normalized = [list(row) for row in values]
    nonempty = _count_nonempty_cells(normalized)
    if nonempty < _DETECTION_CONFIG["min_nonempty_cells"]:
        return []

    results: list[str] = []
    clusters = _nonempty_clusters(normalized)
    for r0, c0, r1, c1 in clusters:
        sub = [row[c0 : c1 + 1] for row in normalized[r0 : r1 + 1]]
        density, coverage = _table_density_metrics(sub)
        if (
            density < _DETECTION_CONFIG["density_min"]
            and coverage < _DETECTION_CONFIG["coverage_min"]
        ):
            continue
        if not _is_plausible_table(sub):
            continue
        score = _table_signal_score(sub)
        if score < _DETECTION_CONFIG["table_score_threshold"]:
            continue
        addr = (
            f"{col_name(base_left + c0)}{base_top + r0}:"
            f"{col_name(base_left + c1)}{base_top + r1}"
        )
        results.append(addr)
    return results


def _count_nonempty_cells(values: Sequence[Sequence[object]]) -> int:
    """Count non-empty cells in a normalized matrix.

    Args:
        values: Normalized matrix of values.

    Returns:
        Number of non-empty cells.
    """
    return sum(
        1 for row in values for v in row if not (v is None or str(v).strip() == "")
    )


def _extract_openpyxl_table_refs(ws: Worksheet) -> list[str]:
    """Extract table reference strings from an openpyxl worksheet.

    Args:
        ws: Target worksheet.

    Returns:
        List of table reference strings.
    """
    tables: list[str] = []
    try:
        openpyxl_tables: list[object] = []
        if hasattr(ws, "tables") and ws.tables:
            if isinstance(ws.tables, dict):
                openpyxl_tables = list(ws.tables.values())
            else:
                openpyxl_tables = list(ws.tables)
        elif hasattr(ws, "_tables") and ws._tables:
            openpyxl_tables = list(ws._tables)
        for t in openpyxl_tables:
            addr = getattr(t, "ref", None)
            if addr:
                tables.append(str(addr))
    except Exception:
        pass
    return tables


def detect_tables_xlwings(sheet: xw.Sheet) -> list[str]:
    """Detect table-like ranges via COM: ListObjects first, then border clusters."""
    tables: list[str] = []
    tables.extend(_extract_listobject_tables(sheet))

    rects = _detect_border_rectangles_xlwings(sheet)
    merged_rects = _merge_rectangles(rects)
    dedup: set[str] = set(tables)

    for top_row, left_col, bottom_row, right_col in merged_rects:
        top_row, left_col, bottom_row, right_col = shrink_to_content(
            sheet, top_row, left_col, bottom_row, right_col, require_inside_border=False
        )
        rng_vals: object | None = None
        try:
            rng_vals = sheet.range((top_row, left_col), (bottom_row, right_col)).value
        except Exception as exc:
            logger.warning(
                "Failed to read range for table detection (%s). (%r)",
                sheet.name,
                exc,
            )
        if rng_vals is None:
            continue
        candidates = _collect_table_candidates_from_values(
            _normalize_matrix(rng_vals),
            base_top=top_row,
            base_left=left_col,
            col_name=xw.utils.col_name,
        )
        for addr in candidates:
            if addr not in dedup:
                dedup.add(addr)
                tables.append(addr)
    return tables


def detect_tables_openpyxl(xlsx_path: Path, sheet_name: str) -> list[str]:
    """Detect table-like ranges via openpyxl tables and border clusters."""
    with openpyxl_workbook(xlsx_path, data_only=True, read_only=False) as wb:
        ws = wb[sheet_name]
        tables = _extract_openpyxl_table_refs(ws)

        has_border, top_edge, bottom_edge, left_edge, right_edge, max_row, max_col = (
            load_border_maps_xlsx(xlsx_path, sheet_name)
        )
        rects = _detect_border_rectangles(has_border, min_size=4)
        merged_rects = _merge_rectangles(rects)
        dedup: set[str] = set(tables)

        for top_row, left_col, bottom_row, right_col in merged_rects:
            top_row, left_col, bottom_row, right_col = shrink_to_content_openpyxl(
                ws,
                top_row,
                left_col,
                bottom_row,
                right_col,
                require_inside_border=False,
                top_edge=top_edge,
                bottom_edge=bottom_edge,
                left_edge=left_edge,
                right_edge=right_edge,
                min_nonempty_ratio=0.0,
            )
            vals_block = _get_values_block(ws, top_row, left_col, bottom_row, right_col)
            candidates = _collect_table_candidates_from_values(
                _normalize_matrix(vals_block),
                base_top=top_row,
                base_left=left_col,
                col_name=get_column_letter,
            )
            for addr in candidates:
                if addr not in dedup:
                    dedup.add(addr)
                    tables.append(addr)
        return tables


def detect_tables(sheet: xw.Sheet) -> list[str]:
    excel_path: Path | None = None
    try:
        excel_path = Path(sheet.book.fullname)
    except Exception:
        excel_path = None

    if excel_path and excel_path.suffix.lower() == ".xls":
        warn_once(
            f"xls-fallback::{excel_path}",
            f"File '{excel_path.name}' is .xls (BIFF); openpyxl cannot read it. Falling back to COM-based detection (slower). Consider converting to .xlsx.",
        )
        return detect_tables_xlwings(sheet)

    if excel_path and excel_path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # noqa: F401
        except Exception:
            warn_once(
                "openpyxl-missing",
                "openpyxl is not installed. Falling back to COM-based detection (slower).",
            )
            return detect_tables_xlwings(sheet)

        try:
            return detect_tables_openpyxl(excel_path, sheet.name)
        except Exception as e:
            warn_once(
                f"openpyxl-parse-fallback::{excel_path}::{sheet.name}",
                f"openpyxl failed to parse '{excel_path.name}' (sheet '{sheet.name}'): {e!r}. Falling back to COM-based detection (slower).",
            )
            return detect_tables_xlwings(sheet)

    warn_once(
        "unknown-ext-fallback",
        "Workbook path or extension is unavailable; falling back to COM-based detection (slower).",
    )
    return detect_tables_xlwings(sheet)


_INT_RE = re.compile(r"^[+-]?\d+$")
_FLOAT_RE = re.compile(r"^[+-]?\d*\.\d+$")


def _coerce_numeric_preserve_format(val: str) -> int | float | str:
    """
    Convert numeric-looking strings to int/float while keeping precision.
    Integers stay int; decimals keep scale via Decimal before casting to float.
    """
    if _INT_RE.match(val):
        try:
            return int(val)
        except Exception:
            return val
    if _FLOAT_RE.match(val):
        try:
            dec = Decimal(val)
            exponent = int(dec.as_tuple().exponent)
            scale = max(1, -exponent)
            quantized = dec.quantize(Decimal("1." + "0" * scale))
            return float(quantized)
        except (InvalidOperation, Exception):
            return val
    return val
