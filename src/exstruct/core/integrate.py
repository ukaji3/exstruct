from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import xlwings as xw

from ..models import CellRow, PrintArea, Shape, SheetData, WorkbookData
from .cells import (
    detect_tables,
    detect_tables_openpyxl,
    extract_sheet_cells,
    extract_sheet_cells_with_links,
)
from .charts import get_charts
from .shapes import get_shapes_with_position

logger = logging.getLogger(__name__)
_ALLOWED_MODES: set[str] = {"light", "standard", "verbose"}


def _find_open_workbook(file_path: Path) -> xw.Book | None:
    """Return an existing workbook if already open in Excel; otherwise None."""
    try:
        for app in xw.apps:
            for wb in app.books:
                try:
                    if Path(wb.fullname).resolve() == file_path.resolve():
                        return wb
                except Exception:
                    continue
    except Exception:
        return None
    return None


def _open_workbook(file_path: Path) -> tuple[xw.Book, bool]:
    """
    Open workbook:
    - If already open, reuse and do not close Excel on exit.
    - Otherwise create invisible Excel (visible=False) and close when done.
    Returns (workbook, should_close_app).
    """
    existing = _find_open_workbook(file_path)
    if existing:
        return existing, False
    app = xw.App(add_book=False, visible=False)
    wb = app.books.open(str(file_path))
    return wb, True


def _parse_print_area_range(
    range_str: str, *, zero_based: bool = True
) -> tuple[int, int, int, int] | None:
    """
    Parse an Excel range string into (r1, c1, r2, c2). Returns None on failure.
    """
    cleaned = range_str.strip()
    if not cleaned:
        return None
    if "!" in cleaned:
        cleaned = cleaned.split("!", 1)[1]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    except Exception:
        return None
    if zero_based:
        return (min_row - 1, min_col - 1, max_row - 1, max_col - 1)
    return (min_row, min_col, max_row, max_col)


def _extract_print_areas_openpyxl(  # noqa: C901
    file_path: Path,
) -> dict[str, list[PrintArea]]:
    """
    Extract print areas per sheet using openpyxl defined names.

    Returns {sheet_name: [PrintArea, ...]}.
    """
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return {}

    try:
        defined = wb.defined_names.get("_xlnm.Print_Area")
        areas: dict[str, list[PrintArea]] = {}
        if defined:
            for sheet_name, range_str in defined.destinations:
                if sheet_name not in wb.sheetnames:
                    continue
                # A single destination can contain multiple comma-separated ranges.
                for part in str(range_str).split(","):
                    parsed = _parse_print_area_range(part)
                    if not parsed:
                        continue
                    r1, c1, r2, c2 = parsed
                    areas.setdefault(sheet_name, []).append(
                        PrintArea(r1=r1, c1=c1, r2=r2, c2=c2)
                    )
        # Fallback: some files carry sheet-level print_area without defined name.
        if not areas:
            for ws in wb.worksheets:
                pa = getattr(ws, "_print_area", None)
                if not pa:
                    continue
                for part in str(pa).split(","):
                    parsed = _parse_print_area_range(part)
                    if not parsed:
                        continue
                    r1, c1, r2, c2 = parsed
                    areas.setdefault(ws.title, []).append(
                        PrintArea(r1=r1, c1=c1, r2=r2, c2=c2)
                    )
        return areas
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _extract_print_areas_com(workbook: xw.Book) -> dict[str, list[PrintArea]]:
    """
    Extract print areas per sheet via xlwings/COM.

    Uses Sheet.PageSetup.PrintArea which may contain comma-separated ranges.
    """
    areas: dict[str, list[PrintArea]] = {}
    for sheet in workbook.sheets:
        try:
            raw = sheet.api.PageSetup.PrintArea or ""
        except Exception:
            continue
        if not raw:
            continue
        parts = str(raw).split(",")
        for part in parts:
            parsed = _parse_print_area_range(part, zero_based=True)
            if not parsed:
                continue
            r1, c1, r2, c2 = parsed
            areas.setdefault(sheet.name, []).append(
                PrintArea(r1=r1, c1=c1, r2=r2, c2=c2)
            )
    return areas


def integrate_sheet_content(
    cell_data: dict[str, list[CellRow]],
    shape_data: dict[str, list[Shape]],
    workbook: xw.Book,
    mode: Literal["light", "standard", "verbose"] = "standard",
    print_area_data: dict[str, list[PrintArea]] | None = None,
) -> dict[str, SheetData]:
    """Integrate cells, shapes, charts, and tables into SheetData per sheet."""
    result: dict[str, SheetData] = {}
    for sheet_name, rows in cell_data.items():
        sheet_shapes = shape_data.get(sheet_name, [])
        sheet = workbook.sheets[sheet_name]

        sheet_model = SheetData(
            rows=rows,
            shapes=sheet_shapes,
            charts=[] if mode == "light" else get_charts(sheet, mode=mode),
            table_candidates=detect_tables(sheet),
            print_areas=print_area_data.get(sheet_name, []) if print_area_data else [],
        )

        result[sheet_name] = sheet_model
    return result


def extract_workbook(  # noqa: C901
    file_path: Path,
    mode: Literal["light", "standard", "verbose"] = "standard",
    *,
    include_cell_links: bool = False,
    include_print_areas: bool = True,
) -> WorkbookData:
    """Extract workbook and return WorkbookData; fallback to cells+tables if Excel COM is unavailable."""
    if mode not in _ALLOWED_MODES:
        raise ValueError(f"Unsupported mode: {mode}")

    cell_data = (
        extract_sheet_cells_with_links(file_path)
        if include_cell_links
        else extract_sheet_cells(file_path)
    )
    print_area_data: dict[str, list[PrintArea]] = {}
    if include_print_areas:
        print_area_data = _extract_print_areas_openpyxl(file_path)

    def _cells_and_tables_only(reason: str) -> WorkbookData:
        sheets: dict[str, SheetData] = {}
        for sheet_name, rows in cell_data.items():
            try:
                tables = detect_tables_openpyxl(file_path, sheet_name)
            except Exception:
                tables = []
            sheets[sheet_name] = SheetData(
                rows=rows,
                shapes=[],
                charts=[],
                table_candidates=tables,
                print_areas=print_area_data.get(sheet_name, [])
                if include_print_areas
                else [],
            )
        logger.warning(
            "%s Falling back to cells+tables only; shapes and charts will be empty.",
            reason,
        )
        return WorkbookData(book_name=file_path.name, sheets=sheets)

    if mode == "light":
        return _cells_and_tables_only("Light mode selected.")

    if os.getenv("SKIP_COM_TESTS"):
        return _cells_and_tables_only(
            "SKIP_COM_TESTS is set; skipping COM/xlwings access."
        )

    try:
        wb, close_app = _open_workbook(file_path)
    except Exception as e:
        return _cells_and_tables_only(f"xlwings/Excel COM is unavailable. ({e!r})")

    try:
        try:
            shape_data = get_shapes_with_position(wb, mode=mode)
            if include_print_areas and not print_area_data:
                # openpyxl couldn't read (e.g., .xls). Try COM as a fallback.
                try:
                    print_area_data = _extract_print_areas_com(wb)
                except Exception:
                    print_area_data = {}
            merged = integrate_sheet_content(
                cell_data,
                shape_data,
                wb,
                mode=mode,
                print_area_data=print_area_data if include_print_areas else None,
            )
            return WorkbookData(book_name=file_path.name, sheets=merged)
        except Exception as e:
            logger.warning(
                "Shape extraction failed; falling back to cells+tables. (%r)", e
            )
            return _cells_and_tables_only(f"Shape extraction failed ({e!r}).")
    finally:
        # Close only if we created the app to avoid shutting user sessions.
        try:
            if close_app:
                app = wb.app
                wb.close()
                app.quit()
        except Exception:
            pass
