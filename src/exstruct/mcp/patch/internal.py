from __future__ import annotations

from collections.abc import Callable, Iterator
from contextlib import contextmanager
from copy import copy
from pathlib import Path
import re
from typing import Any, Protocol, cast, runtime_checkable
from uuid import uuid4

from pydantic import BaseModel, Field, field_validator, model_validator
import xlwings as xw

from exstruct.cli.availability import get_com_availability as get_com_availability

from ..extract_runner import OnConflictPolicy
from ..io import PathPolicy
from ..shared.a1 import (
    column_index_to_label as _shared_column_index_to_label,
    column_label_to_index as _shared_column_label_to_index,
    range_cell_count as _shared_range_cell_count,
    split_a1 as _shared_split_a1,
)
from ..shared.output_path import (
    apply_conflict_policy as _shared_apply_conflict_policy,
    next_available_path as _shared_next_available_path,
    resolve_output_path as _shared_resolve_output_path,
)
from .chart_types import (
    SUPPORTED_CHART_TYPES_CSV,
    normalize_chart_type,
    resolve_chart_type_id,
)
from .types import (
    FormulaIssueCode,
    FormulaIssueLevel,
    HorizontalAlignType,
    PatchBackend,
    PatchEngine,
    PatchOpType,
    PatchStatus,
    PatchValueKind,
    VerticalAlignType,
)

_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_A1_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")
_A1_RANGE_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*:[A-Za-z]{1,3}[1-9][0-9]*$")
_SHEET_QUALIFIED_A1_RANGE_PATTERN = re.compile(
    r"^(?P<sheet>(?:'(?:(?:[^']|'')+)'|[^!]+)!)?"
    r"(?P<start>[A-Za-z]{1,3}[1-9][0-9]*):(?P<end>[A-Za-z]{1,3}[1-9][0-9]*)$"
)
_HEX_COLOR_PATTERN = re.compile(r"^#?(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$")
_COLUMN_LABEL_PATTERN = re.compile(r"^[A-Za-z]{1,3}$")
_MAX_STYLE_TARGET_CELLS = 10_000
_SOFT_MAX_OPS_WARNING_THRESHOLD = 200

_XLWINGS_HORIZONTAL_ALIGN_MAP: dict[HorizontalAlignType, int] = {
    "general": -4105,
    "left": -4131,
    "center": -4108,
    "right": -4152,
    "fill": 5,
    "justify": -4130,
    "centerContinuous": 7,
    "distributed": -4117,
}
_XLWINGS_VERTICAL_ALIGN_MAP: dict[VerticalAlignType, int] = {
    "top": -4160,
    "center": -4108,
    "bottom": -4107,
    "justify": -4130,
    "distributed": -4117,
}


class BorderSideSnapshot(BaseModel):
    """Serializable border side state for inverse restoration."""

    style: str | None = None
    color: str | None = None


class BorderSnapshot(BaseModel):
    """Serializable border state for one cell."""

    cell: str
    top: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    right: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    bottom: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)
    left: BorderSideSnapshot = Field(default_factory=BorderSideSnapshot)


class FontSnapshot(BaseModel):
    """Serializable font state for one cell."""

    cell: str
    bold: bool | None = None
    size: float | None = None
    color: str | None = None


class FillSnapshot(BaseModel):
    """Serializable fill state for one cell."""

    cell: str
    fill_type: str | None = None
    start_color: str | None = None
    end_color: str | None = None


class AlignmentSnapshot(BaseModel):
    """Serializable alignment state for one cell."""

    cell: str
    horizontal: str | None = None
    vertical: str | None = None
    wrap_text: bool | None = None


class MergeStateSnapshot(BaseModel):
    """Serializable merged-range state for deterministic restoration."""

    scope: str
    ranges: list[str] = Field(default_factory=list)


class RowDimensionSnapshot(BaseModel):
    """Serializable row height state."""

    row: int
    height: float | None = None


class ColumnDimensionSnapshot(BaseModel):
    """Serializable column width state."""

    column: str
    width: float | None = None


class ListObjectAddAttempt(BaseModel):
    """Typed COM ListObjects.Add invocation attempt."""

    args: tuple[object, ...]
    call_kwargs: dict[str, object] = Field(default_factory=dict)
    signature: str


class DesignSnapshot(BaseModel):
    """Serializable style/dimension snapshot for inverse restore."""

    borders: list[BorderSnapshot] = Field(default_factory=list)
    fonts: list[FontSnapshot] = Field(default_factory=list)
    fills: list[FillSnapshot] = Field(default_factory=list)
    alignments: list[AlignmentSnapshot] = Field(default_factory=list)
    merge_state: MergeStateSnapshot | None = None
    row_dimensions: list[RowDimensionSnapshot] = Field(default_factory=list)
    column_dimensions: list[ColumnDimensionSnapshot] = Field(default_factory=list)


@runtime_checkable
class OpenpyxlCellProtocol(Protocol):
    """Protocol for openpyxl cell access used by patch runner."""

    value: str | int | float | None
    data_type: str | None
    font: OpenpyxlFontProtocol
    fill: OpenpyxlFillProtocol
    border: OpenpyxlBorderProtocol
    alignment: OpenpyxlAlignmentProtocol


@runtime_checkable
class OpenpyxlColorProtocol(Protocol):
    """Protocol for openpyxl color access."""

    rgb: object | None


@runtime_checkable
class OpenpyxlSideProtocol(Protocol):
    """Protocol for openpyxl border side access."""

    style: str | None
    color: OpenpyxlColorProtocol | None


@runtime_checkable
class OpenpyxlBorderProtocol(Protocol):
    """Protocol for openpyxl border access."""

    top: OpenpyxlSideProtocol
    right: OpenpyxlSideProtocol
    bottom: OpenpyxlSideProtocol
    left: OpenpyxlSideProtocol


@runtime_checkable
class OpenpyxlFontProtocol(Protocol):
    """Protocol for openpyxl font access."""

    bold: bool | None
    size: float | None
    color: object | None


@runtime_checkable
class OpenpyxlFillProtocol(Protocol):
    """Protocol for openpyxl fill access."""

    fill_type: str | None
    start_color: OpenpyxlColorProtocol | None
    end_color: OpenpyxlColorProtocol | None


@runtime_checkable
class OpenpyxlAlignmentProtocol(Protocol):
    """Protocol for openpyxl alignment access."""

    horizontal: str | None
    vertical: str | None
    wrap_text: bool | None


@runtime_checkable
class OpenpyxlRowDimensionProtocol(Protocol):
    """Protocol for openpyxl row dimension access."""

    height: float | None


@runtime_checkable
class OpenpyxlColumnDimensionProtocol(Protocol):
    """Protocol for openpyxl column dimension access."""

    width: float | None


@runtime_checkable
class OpenpyxlRowDimensionsProtocol(Protocol):
    """Protocol for openpyxl row dimensions collection."""

    def __getitem__(self, key: int) -> OpenpyxlRowDimensionProtocol: ...


@runtime_checkable
class OpenpyxlColumnDimensionsProtocol(Protocol):
    """Protocol for openpyxl column dimensions collection."""

    def __getitem__(self, key: str) -> OpenpyxlColumnDimensionProtocol: ...


@runtime_checkable
class OpenpyxlWorksheetProtocol(Protocol):
    """Protocol for openpyxl worksheet access used by patch runner."""

    row_dimensions: OpenpyxlRowDimensionsProtocol
    column_dimensions: OpenpyxlColumnDimensionsProtocol

    def __getitem__(self, key: str) -> OpenpyxlCellProtocol: ...

    def merge_cells(self, range_string: str) -> None: ...

    def unmerge_cells(self, range_string: str) -> None: ...


@runtime_checkable
class OpenpyxlTablesProtocol(Protocol):
    """Protocol for openpyxl worksheet tables collection."""

    def items(self) -> Iterator[tuple[object, object]]: ...


@runtime_checkable
class OpenpyxlWorkbookProtocol(Protocol):
    """Protocol for openpyxl workbook access used by patch runner."""

    sheetnames: list[str]

    def __getitem__(self, key: str) -> OpenpyxlWorksheetProtocol: ...

    def create_sheet(self, title: str) -> OpenpyxlWorksheetProtocol: ...

    def save(self, filename: str | Path) -> None: ...

    def close(self) -> None: ...


@runtime_checkable
class XlwingsRangeProtocol(Protocol):
    """Protocol for xlwings range access used by patch runner."""

    value: object | None
    formula: str | None
    api: object


@runtime_checkable
class XlwingsSheetProtocol(Protocol):
    """Protocol for xlwings sheet access used by patch runner."""

    name: str
    api: object

    def range(self, cell: str) -> XlwingsRangeProtocol: ...


@runtime_checkable
class XlwingsSheetsProtocol(Protocol):
    """Protocol for xlwings sheets collection."""

    def __iter__(self) -> Iterator[XlwingsSheetProtocol]: ...

    def __len__(self) -> int: ...

    def __getitem__(self, index: int) -> XlwingsSheetProtocol: ...

    def add(
        self, name: str, after: XlwingsSheetProtocol | None = None
    ) -> XlwingsSheetProtocol: ...


@runtime_checkable
class XlwingsWorkbookProtocol(Protocol):
    """Protocol for xlwings workbook access used by patch runner."""

    sheets: XlwingsSheetsProtocol

    def save(self, filename: str) -> None: ...

    def close(self) -> None: ...


@runtime_checkable
class XlwingsAppProtocol(Protocol):
    """Protocol for xlwings app lifecycle used during cleanup."""

    display_alerts: bool
    screen_updating: bool

    def quit(self) -> None: ...  # noqa: N802

    def kill(self) -> None: ...  # noqa: N802


@runtime_checkable
class XlwingsFontApiProtocol(Protocol):
    """Protocol for xlwings COM font API."""

    Bold: bool
    Size: float
    Color: int


@runtime_checkable
class XlwingsInteriorApiProtocol(Protocol):
    """Protocol for xlwings COM interior API."""

    Color: int


@runtime_checkable
class XlwingsBorderApiProtocol(Protocol):
    """Protocol for xlwings COM border API."""

    LineStyle: int
    Color: int


@runtime_checkable
class XlwingsMergeAreaApiProtocol(Protocol):
    """Protocol for xlwings COM merged-area API."""

    def Address(self, row_absolute: bool, column_absolute: bool) -> str: ...  # noqa: N802


@runtime_checkable
class XlwingsRangeApiProtocol(Protocol):
    """Protocol for xlwings COM range API."""

    Font: XlwingsFontApiProtocol
    Interior: XlwingsInteriorApiProtocol
    MergeCells: bool
    MergeArea: XlwingsMergeAreaApiProtocol
    HorizontalAlignment: int
    VerticalAlignment: int
    WrapText: bool
    Left: float
    Top: float

    def Borders(self, edge: int) -> XlwingsBorderApiProtocol: ...  # noqa: N802

    def Merge(self) -> None: ...  # noqa: N802

    def UnMerge(self) -> None: ...  # noqa: N802


@runtime_checkable
class XlwingsRowApiProtocol(Protocol):
    """Protocol for xlwings COM row API."""

    RowHeight: float


@runtime_checkable
class XlwingsColumnApiProtocol(Protocol):
    """Protocol for xlwings COM column API."""

    ColumnWidth: float

    def AutoFit(self) -> None: ...  # noqa: N802


@runtime_checkable
class XlwingsSheetApiProtocol(Protocol):
    """Protocol for xlwings COM sheet API."""

    def Rows(self, index: int) -> XlwingsRowApiProtocol: ...  # noqa: N802

    def Columns(self, key: str) -> XlwingsColumnApiProtocol: ...  # noqa: N802

    def ChartObjects(self) -> XlwingsChartObjectsCollectionProtocol: ...  # noqa: N802

    def ListObjects(self) -> object: ...  # noqa: N802


@runtime_checkable
class XlwingsChartObjectProtocol(Protocol):
    """Protocol for xlwings COM chart object."""

    Name: str
    Chart: object


@runtime_checkable
class XlwingsChartObjectsCollectionProtocol(Protocol):
    """Protocol for xlwings COM ChartObjects collection."""

    Count: int

    def Add(  # noqa: N802
        self, left: float, top: float, width: float, height: float
    ) -> XlwingsChartObjectProtocol: ...

    def __call__(self, index: int) -> XlwingsChartObjectProtocol: ...


@runtime_checkable
class XlwingsChartSeriesProtocol(Protocol):
    """Protocol for xlwings COM chart series object."""

    Name: str
    XValues: object
    Values: object


@runtime_checkable
class XlwingsChartSeriesCollectionProtocol(Protocol):
    """Protocol for xlwings COM chart series collection."""

    Count: int

    def NewSeries(self) -> XlwingsChartSeriesProtocol: ...  # noqa: N802


class PatchOp(BaseModel):
    """Single patch operation for an Excel workbook.

    Operation types and their required fields:

    - ``set_value``: Set a cell value. Requires ``sheet``, ``cell``, ``value``.
    - ``set_formula``: Set a cell formula. Requires ``sheet``, ``cell``, ``formula`` (must start with ``=``).
    - ``add_sheet``: Add a new worksheet. Requires ``sheet`` (new sheet name). No ``cell``/``value``/``formula``.
    - ``set_range_values``: Set values for a rectangular range. Requires ``sheet``, ``range`` (e.g. ``A1:C3``), ``values`` (2D list matching range shape).
    - ``fill_formula``: Fill a formula across a single row or column. Requires ``sheet``, ``range``, ``base_cell``, ``formula``.
    - ``set_value_if``: Conditionally set value. Requires ``sheet``, ``cell``, ``value``. ``expected`` is optional; ``null`` matches an empty cell. Skips if current value != expected.
    - ``set_formula_if``: Conditionally set formula. Requires ``sheet``, ``cell``, ``formula``. ``expected`` is optional; ``null`` matches an empty cell. Skips if current value != expected.
    - ``draw_grid_border``: Draw thin black borders on a target rectangle.
    - ``set_bold``: Set bold style for one cell or one range.
    - ``set_font_size``: Set font size for one cell or one range.
    - ``set_font_color``: Set font color for one cell or one range.
    - ``set_fill_color``: Set solid fill color for one cell or one range.
    - ``set_dimensions``: Set row height and/or column width.
    - ``auto_fit_columns``: Auto-fit column widths with optional bounds.
    - ``merge_cells``: Merge a rectangular range.
    - ``unmerge_cells``: Unmerge all merged ranges intersecting target range.
    - ``set_alignment``: Set horizontal/vertical alignment and/or wrap_text.
    - ``set_style``: Set multiple style attributes in one operation.
    - ``apply_table_style``: Create an Excel table and apply table style.
    - ``create_chart``: Create a new chart from source ranges (COM only).
    - ``restore_design_snapshot``: Restore style/dimension snapshot (internal inverse op).
    """

    op: PatchOpType = Field(
        description=(
            "Operation type: 'set_value', 'set_formula', 'add_sheet', "
            "'set_range_values', 'fill_formula', 'set_value_if', 'set_formula_if', "
            "'draw_grid_border', 'set_bold', 'set_font_size', 'set_font_color', "
            "'set_fill_color', "
            "'set_dimensions', "
            "'auto_fit_columns', "
            "'merge_cells', 'unmerge_cells', 'set_alignment', 'set_style', "
            "'apply_table_style', "
            "'create_chart', "
            "or 'restore_design_snapshot'."
        )
    )
    sheet: str = Field(
        description="Target sheet name. For add_sheet, this is the new sheet name."
    )
    cell: str | None = Field(
        default=None,
        description="Cell reference in A1 notation (e.g. 'B2'). Required for set_value, set_formula, set_value_if, set_formula_if.",
    )
    range: str | None = Field(
        default=None,
        description="Range reference in A1 notation (e.g. 'A1:C3'). Required for set_range_values and fill_formula.",
    )
    base_cell: str | None = Field(
        default=None,
        description="Base cell for formula translation in fill_formula (e.g. 'C2').",
    )
    expected: str | int | float | None = Field(
        default=None,
        description="Expected current value for conditional ops (set_value_if, set_formula_if). Operation is skipped if mismatch.",
    )
    value: str | int | float | None = Field(
        default=None,
        description="Value to set. Use null to clear a cell. For set_value and set_value_if.",
    )
    values: list[list[str | int | float | None]] | None = Field(
        default=None,
        description="2D list of values for set_range_values. Shape must match the range dimensions.",
    )
    formula: str | None = Field(
        default=None,
        description="Formula string starting with '=' (e.g. '=SUM(A1:A10)'). For set_formula, set_formula_if, fill_formula.",
    )
    row_count: int | None = Field(
        default=None,
        description="Row count for draw_grid_border.",
    )
    col_count: int | None = Field(
        default=None,
        description="Column count for draw_grid_border.",
    )
    bold: bool | None = Field(
        default=None,
        description="Bold flag for set_bold. Defaults to true.",
    )
    font_size: float | None = Field(
        default=None,
        description="Font size for set_font_size. Must be > 0.",
    )
    color: str | None = Field(
        default=None,
        description="Font color for set_font_color in RRGGBB/AARRGGBB (with optional '#').",
    )
    fill_color: str | None = Field(
        default=None,
        description="Fill color for set_fill_color in RRGGBB/AARRGGBB (with optional '#').",
    )
    rows: list[int] | None = Field(
        default=None,
        description="Row indexes for set_dimensions.",
    )
    columns: list[str | int] | None = Field(
        default=None,
        description="Column identifiers for set_dimensions. Accepts letters (A/AA) or positive indexes.",
    )
    row_height: float | None = Field(
        default=None,
        description="Target row height for set_dimensions.",
    )
    column_width: float | None = Field(
        default=None,
        description="Target column width for set_dimensions.",
    )
    min_width: float | None = Field(
        default=None,
        description="Optional minimum width bound for auto_fit_columns.",
    )
    max_width: float | None = Field(
        default=None,
        description="Optional maximum width bound for auto_fit_columns.",
    )
    horizontal_align: HorizontalAlignType | None = Field(
        default=None,
        description="Horizontal alignment for set_alignment/set_style.",
    )
    vertical_align: VerticalAlignType | None = Field(
        default=None,
        description="Vertical alignment for set_alignment/set_style.",
    )
    wrap_text: bool | None = Field(
        default=None,
        description="Wrap text flag for set_alignment/set_style.",
    )
    style: str | None = Field(
        default=None,
        description="Table style name for apply_table_style.",
    )
    table_name: str | None = Field(
        default=None,
        description="Optional table name for apply_table_style.",
    )
    design_snapshot: DesignSnapshot | None = Field(
        default=None,
        description="Design snapshot payload for restore_design_snapshot.",
    )
    chart_type: str | None = Field(
        default=None,
        description=(
            "Chart type for create_chart: line, column, bar, area, pie, "
            "doughnut, scatter, radar."
        ),
    )
    data_range: str | list[str] | None = Field(
        default=None,
        description=(
            "Data range in A1 notation for create_chart. "
            "Accepts a single range or a list of ranges."
        ),
    )
    category_range: str | None = Field(
        default=None,
        description="Optional category range in A1 notation for create_chart.",
    )
    anchor_cell: str | None = Field(
        default=None,
        description="Top-left anchor cell in A1 notation for chart placement.",
    )
    chart_name: str | None = Field(
        default=None,
        description="Optional chart object name for create_chart.",
    )
    width: float | None = Field(
        default=None,
        description="Optional chart width (points) for create_chart.",
    )
    height: float | None = Field(
        default=None,
        description="Optional chart height (points) for create_chart.",
    )
    titles_from_data: bool | None = Field(
        default=None,
        description="Whether to infer titles from source data for create_chart.",
    )
    series_from_rows: bool | None = Field(
        default=None,
        description="Whether chart series are oriented by rows for create_chart.",
    )
    chart_title: str | None = Field(
        default=None,
        description="Optional chart title text for create_chart.",
    )
    x_axis_title: str | None = Field(
        default=None,
        description="Optional X-axis title text for create_chart.",
    )
    y_axis_title: str | None = Field(
        default=None,
        description="Optional Y-axis title text for create_chart.",
    )

    @field_validator("sheet")
    @classmethod
    def _validate_sheet(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("sheet must not be empty.")
        return value

    @field_validator("cell")
    @classmethod
    def _validate_cell(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_PATTERN.match(candidate):
            raise ValueError(f"Invalid cell reference: {value}")
        return candidate.upper()

    @field_validator("base_cell")
    @classmethod
    def _validate_base_cell(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_PATTERN.match(candidate):
            raise ValueError(f"Invalid base_cell reference: {value}")
        return candidate.upper()

    @field_validator("range")
    @classmethod
    def _validate_range(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_RANGE_PATTERN.match(candidate):
            raise ValueError(f"Invalid range reference: {value}")
        start, end = candidate.split(":", maxsplit=1)
        return f"{start.upper()}:{end.upper()}"

    @field_validator("data_range")
    @classmethod
    def _validate_data_range(
        cls, value: str | list[str] | None
    ) -> str | list[str] | None:
        if value is None:
            return None
        if isinstance(value, str):
            return _normalize_chart_range_reference(value)
        if not value:
            raise ValueError("data_range list must not be empty.")
        normalized: list[str] = []
        for item in value:
            normalized.append(_normalize_chart_range_reference(item))
        return normalized

    @field_validator("category_range")
    @classmethod
    def _validate_category_range(cls, value: str | None) -> str | None:
        if value is None:
            return None
        return _normalize_chart_range_reference(value)

    @field_validator("anchor_cell")
    @classmethod
    def _validate_anchor_cell(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not _A1_PATTERN.match(candidate):
            raise ValueError(f"Invalid anchor_cell reference: {value}")
        return candidate.upper()

    @field_validator("chart_type")
    @classmethod
    def _validate_chart_type(cls, value: str | None) -> str | None:
        if value is None:
            return None
        normalized = normalize_chart_type(value)
        if normalized is None:
            raise ValueError(f"chart_type must be one of: {SUPPORTED_CHART_TYPES_CSV}.")
        return normalized

    @field_validator("fill_color")
    @classmethod
    def _validate_fill_color(cls, value: str | None) -> str | None:
        if value is None:
            return None
        return _normalize_hex_input(value, field_name="fill_color")

    @field_validator("color")
    @classmethod
    def _validate_color(cls, value: str | None) -> str | None:
        if value is None:
            return None
        return _normalize_hex_input(value, field_name="color")

    @field_validator("rows")
    @classmethod
    def _validate_rows(cls, value: list[int] | None) -> list[int] | None:
        if value is None:
            return None
        if not value:
            raise ValueError("rows must not be empty.")
        normalized: list[int] = []
        for row in value:
            if row < 1:
                raise ValueError("rows must contain positive integers.")
            normalized.append(row)
        return normalized

    @field_validator("columns")
    @classmethod
    def _validate_columns(cls, value: list[str | int] | None) -> list[str | int] | None:
        if value is None:
            return None
        if not value:
            raise ValueError("columns must not be empty.")
        normalized: list[str | int] = []
        for column in value:
            normalized.append(_normalize_column_identifier(column))
        return normalized

    @field_validator(
        "style",
        "table_name",
        "chart_name",
        "chart_title",
        "x_axis_title",
        "y_axis_title",
    )
    @classmethod
    def _validate_non_empty_optional_text(cls, value: str | None) -> str | None:
        if value is None:
            return None
        candidate = value.strip()
        if not candidate:
            raise ValueError(
                "style/table_name/chart_name/chart_title/x_axis_title/y_axis_title "
                "must not be empty when provided."
            )
        return candidate

    @field_validator("min_width", "max_width", "width", "height")
    @classmethod
    def _validate_optional_positive_width(cls, value: float | None) -> float | None:
        if value is None:
            return None
        if value <= 0:
            raise ValueError("min_width/max_width/width/height must be > 0.")
        return value

    @model_validator(mode="after")
    def _validate_op(self) -> PatchOp:
        validator = _validator_for_op(self.op)
        if validator is None:
            return self
        if self.op in _CELL_REQUIRED_OPS:
            _validate_cell_required(self)
        validator(self)
        return self


_CELL_REQUIRED_OPS: set[PatchOpType] = {
    "set_value",
    "set_formula",
    "set_value_if",
    "set_formula_if",
}


def _validator_for_op(op_type: PatchOpType) -> Callable[[PatchOp], None] | None:
    """Return per-op validator function."""
    validators: dict[PatchOpType, Callable[[PatchOp], None]] = {
        "add_sheet": _validate_add_sheet,
        "set_value": _validate_set_value,
        "set_formula": _validate_set_formula,
        "set_range_values": _validate_set_range_values,
        "fill_formula": _validate_fill_formula,
        "set_value_if": _validate_set_value_if,
        "set_formula_if": _validate_set_formula_if,
        "draw_grid_border": _validate_draw_grid_border,
        "set_bold": _validate_set_bold,
        "set_font_size": _validate_set_font_size,
        "set_font_color": _validate_set_font_color,
        "set_fill_color": _validate_set_fill_color,
        "set_dimensions": _validate_set_dimensions,
        "auto_fit_columns": _validate_auto_fit_columns,
        "merge_cells": _validate_merge_cells,
        "unmerge_cells": _validate_unmerge_cells,
        "set_alignment": _validate_set_alignment,
        "set_style": _validate_set_style,
        "apply_table_style": _validate_apply_table_style,
        "create_chart": _validate_create_chart,
        "restore_design_snapshot": _validate_restore_design_snapshot,
    }
    return validators.get(op_type)


def _validate_add_sheet(op: PatchOp) -> None:
    """Validate add_sheet operation."""
    _validate_no_design_fields(op, op_name="add_sheet")
    if op.cell is not None:
        raise ValueError("add_sheet does not accept cell.")
    if op.range is not None:
        raise ValueError("add_sheet does not accept range.")
    if op.base_cell is not None:
        raise ValueError("add_sheet does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("add_sheet does not accept expected.")
    if op.value is not None:
        raise ValueError("add_sheet does not accept value.")
    if op.values is not None:
        raise ValueError("add_sheet does not accept values.")
    if op.formula is not None:
        raise ValueError("add_sheet does not accept formula.")


def _validate_cell_required(op: PatchOp) -> None:
    """Validate that the operation has a cell value."""
    if op.cell is None:
        raise ValueError(f"{op.op} requires cell.")


def _validate_set_value(op: PatchOp) -> None:
    """Validate set_value operation."""
    _validate_no_design_fields(op, op_name="set_value")
    if op.range is not None:
        raise ValueError("set_value does not accept range.")
    if op.base_cell is not None:
        raise ValueError("set_value does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_value does not accept expected.")
    if op.values is not None:
        raise ValueError("set_value does not accept values.")
    if op.formula is not None:
        raise ValueError("set_value does not accept formula.")


def _validate_set_formula(op: PatchOp) -> None:
    """Validate set_formula operation."""
    _validate_no_design_fields(op, op_name="set_formula")
    if op.range is not None:
        raise ValueError("set_formula does not accept range.")
    if op.base_cell is not None:
        raise ValueError("set_formula does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_formula does not accept expected.")
    if op.values is not None:
        raise ValueError("set_formula does not accept values.")
    if op.value is not None:
        raise ValueError("set_formula does not accept value.")
    if op.formula is None:
        raise ValueError("set_formula requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("set_formula requires formula starting with '='.")


def _validate_set_range_values(op: PatchOp) -> None:
    """Validate set_range_values operation."""
    _validate_no_design_fields(op, op_name="set_range_values")
    if op.cell is not None:
        raise ValueError("set_range_values does not accept cell.")
    if op.base_cell is not None:
        raise ValueError("set_range_values does not accept base_cell.")
    if op.expected is not None:
        raise ValueError("set_range_values does not accept expected.")
    if op.formula is not None:
        raise ValueError("set_range_values does not accept formula.")
    if op.range is None:
        raise ValueError("set_range_values requires range.")
    if op.values is None:
        raise ValueError("set_range_values requires values.")
    if not op.values:
        raise ValueError("set_range_values requires non-empty values.")
    if not all(op.values):
        raise ValueError("set_range_values values rows must not be empty.")
    expected_width = len(op.values[0])
    if any(len(row) != expected_width for row in op.values):
        raise ValueError("set_range_values requires rectangular values.")


def _validate_fill_formula(op: PatchOp) -> None:
    """Validate fill_formula operation."""
    _validate_no_design_fields(op, op_name="fill_formula")
    if op.cell is not None:
        raise ValueError("fill_formula does not accept cell.")
    if op.expected is not None:
        raise ValueError("fill_formula does not accept expected.")
    if op.value is not None:
        raise ValueError("fill_formula does not accept value.")
    if op.values is not None:
        raise ValueError("fill_formula does not accept values.")
    if op.range is None:
        raise ValueError("fill_formula requires range.")
    if op.base_cell is None:
        raise ValueError("fill_formula requires base_cell.")
    if op.formula is None:
        raise ValueError("fill_formula requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("fill_formula requires formula starting with '='.")


def _validate_set_value_if(op: PatchOp) -> None:
    """Validate set_value_if operation."""
    _validate_no_design_fields(op, op_name="set_value_if")
    if op.formula is not None:
        raise ValueError("set_value_if does not accept formula.")
    if op.range is not None:
        raise ValueError("set_value_if does not accept range.")
    if op.values is not None:
        raise ValueError("set_value_if does not accept values.")
    if op.base_cell is not None:
        raise ValueError("set_value_if does not accept base_cell.")


def _validate_set_formula_if(op: PatchOp) -> None:
    """Validate set_formula_if operation."""
    _validate_no_design_fields(op, op_name="set_formula_if")
    if op.value is not None:
        raise ValueError("set_formula_if does not accept value.")
    if op.range is not None:
        raise ValueError("set_formula_if does not accept range.")
    if op.values is not None:
        raise ValueError("set_formula_if does not accept values.")
    if op.base_cell is not None:
        raise ValueError("set_formula_if does not accept base_cell.")
    if op.formula is None:
        raise ValueError("set_formula_if requires formula.")
    if not op.formula.startswith("="):
        raise ValueError("set_formula_if requires formula starting with '='.")


def _validate_draw_grid_border(op: PatchOp) -> None:
    """Validate draw_grid_border operation."""
    _validate_no_legacy_edit_fields(op, op_name="draw_grid_border")
    if op.cell is not None or op.range is not None:
        raise ValueError("draw_grid_border does not accept cell or range.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("draw_grid_border does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("draw_grid_border does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("draw_grid_border does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("draw_grid_border does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("draw_grid_border does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="draw_grid_border")
    if op.base_cell is None:
        raise ValueError("draw_grid_border requires base_cell.")
    if op.row_count is None or op.col_count is None:
        raise ValueError("draw_grid_border requires row_count and col_count.")
    if op.row_count < 1 or op.col_count < 1:
        raise ValueError("draw_grid_border requires row_count >= 1 and col_count >= 1.")
    if op.row_count * op.col_count > _MAX_STYLE_TARGET_CELLS:
        raise ValueError(
            f"draw_grid_border target exceeds max cells: {_MAX_STYLE_TARGET_CELLS}."
        )


def _validate_set_bold(op: PatchOp) -> None:
    """Validate set_bold operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_bold")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_bold does not accept row_count or col_count.")
    if op.color is not None or op.fill_color is not None:
        raise ValueError("set_bold does not accept color or fill_color.")
    if op.font_size is not None:
        raise ValueError("set_bold does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_bold does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_bold does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_bold does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_bold")
    _validate_exactly_one_cell_or_range(op, op_name="set_bold")
    if op.bold is None:
        op.bold = True
    _validate_style_target_size(op, op_name="set_bold")


def _validate_set_font_size(op: PatchOp) -> None:
    """Validate set_font_size operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_font_size")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_font_size does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("set_font_size does not accept bold, color, or fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_font_size does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_font_size does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_font_size does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_font_size")
    _validate_exactly_one_cell_or_range(op, op_name="set_font_size")
    if op.font_size is None:
        raise ValueError("set_font_size requires font_size.")
    if op.font_size <= 0:
        raise ValueError("set_font_size font_size must be > 0.")
    _validate_style_target_size(op, op_name="set_font_size")


def _validate_set_font_color(op: PatchOp) -> None:
    """Validate set_font_color operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_font_color")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_font_color does not accept row_count or col_count.")
    if op.bold is not None:
        raise ValueError("set_font_color does not accept bold.")
    if op.font_size is not None:
        raise ValueError("set_font_color does not accept font_size.")
    if op.fill_color is not None:
        raise ValueError("set_font_color does not accept fill_color.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_font_color does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_font_color does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_font_color does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_font_color")
    _validate_exactly_one_cell_or_range(op, op_name="set_font_color")
    if op.color is None:
        raise ValueError("set_font_color requires color.")
    _validate_style_target_size(op, op_name="set_font_color")


def _validate_set_fill_color(op: PatchOp) -> None:
    """Validate set_fill_color operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_fill_color")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_fill_color does not accept row_count or col_count.")
    if op.bold is not None:
        raise ValueError("set_fill_color does not accept bold.")
    if op.color is not None:
        raise ValueError("set_fill_color does not accept color.")
    if op.font_size is not None:
        raise ValueError("set_fill_color does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_fill_color does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_fill_color does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_fill_color does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_fill_color")
    _validate_exactly_one_cell_or_range(op, op_name="set_fill_color")
    if op.fill_color is None:
        raise ValueError("set_fill_color requires fill_color.")
    _validate_style_target_size(op, op_name="set_fill_color")


def _validate_set_dimensions(op: PatchOp) -> None:
    """Validate set_dimensions operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_dimensions")
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError("set_dimensions does not accept cell/range/base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_dimensions does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("set_dimensions does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("set_dimensions does not accept font_size.")
    if op.design_snapshot is not None:
        raise ValueError("set_dimensions does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="set_dimensions")
    has_rows = op.rows is not None
    has_columns = op.columns is not None
    if not has_rows and not has_columns:
        raise ValueError("set_dimensions requires rows and/or columns.")
    if has_rows and op.row_height is None:
        raise ValueError("set_dimensions requires row_height when rows is provided.")
    if has_columns and op.column_width is None:
        raise ValueError(
            "set_dimensions requires column_width when columns is provided."
        )
    if op.row_height is not None and op.row_height <= 0:
        raise ValueError("set_dimensions row_height must be > 0.")
    if op.column_width is not None and op.column_width <= 0:
        raise ValueError("set_dimensions column_width must be > 0.")


def _validate_auto_fit_columns(op: PatchOp) -> None:
    """Validate auto_fit_columns operation."""
    _validate_no_legacy_edit_fields(
        op, op_name="auto_fit_columns", allow_auto_fit_fields=True
    )
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError("auto_fit_columns does not accept cell/range/base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("auto_fit_columns does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("auto_fit_columns does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("auto_fit_columns does not accept font_size.")
    if op.rows is not None or op.row_height is not None or op.column_width is not None:
        raise ValueError(
            "auto_fit_columns does not accept rows, row_height, or column_width."
        )
    if op.design_snapshot is not None:
        raise ValueError("auto_fit_columns does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="auto_fit_columns")
    if (
        op.min_width is not None
        and op.max_width is not None
        and op.min_width > op.max_width
    ):
        raise ValueError("auto_fit_columns requires min_width <= max_width.")


def _validate_merge_cells(op: PatchOp) -> None:
    """Validate merge_cells operation."""
    _validate_no_legacy_edit_fields(op, op_name="merge_cells")
    if op.cell is not None or op.base_cell is not None:
        raise ValueError("merge_cells does not accept cell or base_cell.")
    if op.range is None:
        raise ValueError("merge_cells requires range.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("merge_cells does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("merge_cells does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("merge_cells does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("merge_cells does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("merge_cells does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("merge_cells does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="merge_cells")
    if _range_cell_count(op.range) < 2:
        raise ValueError("merge_cells requires a multi-cell range.")


def _validate_unmerge_cells(op: PatchOp) -> None:
    """Validate unmerge_cells operation."""
    _validate_no_legacy_edit_fields(op, op_name="unmerge_cells")
    if op.cell is not None or op.base_cell is not None:
        raise ValueError("unmerge_cells does not accept cell or base_cell.")
    if op.range is None:
        raise ValueError("unmerge_cells requires range.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("unmerge_cells does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("unmerge_cells does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("unmerge_cells does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("unmerge_cells does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("unmerge_cells does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("unmerge_cells does not accept design_snapshot.")
    _validate_no_alignment_fields(op, op_name="unmerge_cells")


def _validate_set_alignment(op: PatchOp) -> None:
    """Validate set_alignment operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_alignment")
    if op.base_cell is not None:
        raise ValueError("set_alignment does not accept base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_alignment does not accept row_count or col_count.")
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError("set_alignment does not accept bold, color, or fill_color.")
    if op.font_size is not None:
        raise ValueError("set_alignment does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_alignment does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_alignment does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_alignment does not accept design_snapshot.")
    _validate_exactly_one_cell_or_range(op, op_name="set_alignment")
    if (
        op.horizontal_align is None
        and op.vertical_align is None
        and op.wrap_text is None
    ):
        raise ValueError(
            "set_alignment requires at least one of horizontal_align, vertical_align, or wrap_text."
        )
    _validate_style_target_size(op, op_name="set_alignment")


def _validate_set_style(op: PatchOp) -> None:
    """Validate set_style operation."""
    _validate_no_legacy_edit_fields(op, op_name="set_style")
    if op.base_cell is not None:
        raise ValueError("set_style does not accept base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("set_style does not accept row_count or col_count.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("set_style does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("set_style does not accept row_height or column_width.")
    if op.design_snapshot is not None:
        raise ValueError("set_style does not accept design_snapshot.")
    _validate_exactly_one_cell_or_range(op, op_name="set_style")
    if (
        op.bold is None
        and op.font_size is None
        and op.color is None
        and op.fill_color is None
        and op.horizontal_align is None
        and op.vertical_align is None
        and op.wrap_text is None
    ):
        raise ValueError(
            "set_style requires at least one style field from: "
            "bold, font_size, color, fill_color, horizontal_align, vertical_align, wrap_text."
        )
    if op.font_size is not None and op.font_size <= 0:
        raise ValueError("set_style font_size must be > 0.")
    _validate_style_target_size(op, op_name="set_style")


def _validate_apply_table_style(op: PatchOp) -> None:
    """Validate apply_table_style operation."""
    _validate_no_legacy_edit_fields(
        op, op_name="apply_table_style", allow_table_fields=True
    )
    if op.cell is not None or op.base_cell is not None:
        raise ValueError("apply_table_style does not accept cell or base_cell.")
    if op.range is None:
        raise ValueError("apply_table_style requires range.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("apply_table_style does not accept row_count or col_count.")
    if (
        op.bold is not None
        or op.color is not None
        or op.fill_color is not None
        or op.font_size is not None
    ):
        raise ValueError(
            "apply_table_style does not accept bold, color, fill_color, or font_size."
        )
    if op.rows is not None or op.columns is not None:
        raise ValueError("apply_table_style does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError(
            "apply_table_style does not accept row_height or column_width."
        )
    _validate_no_alignment_fields(op, op_name="apply_table_style")
    if op.design_snapshot is not None:
        raise ValueError("apply_table_style does not accept design_snapshot.")
    if op.style is None:
        raise ValueError("apply_table_style requires style.")


def _validate_restore_design_snapshot(op: PatchOp) -> None:
    """Validate restore_design_snapshot operation."""
    _validate_no_legacy_edit_fields(op, op_name="restore_design_snapshot")
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError(
            "restore_design_snapshot does not accept cell/range/base_cell."
        )
    if op.row_count is not None or op.col_count is not None:
        raise ValueError(
            "restore_design_snapshot does not accept row_count or col_count."
        )
    if op.bold is not None or op.color is not None or op.fill_color is not None:
        raise ValueError(
            "restore_design_snapshot does not accept bold, color, or fill_color."
        )
    if op.font_size is not None:
        raise ValueError("restore_design_snapshot does not accept font_size.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("restore_design_snapshot does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError(
            "restore_design_snapshot does not accept row_height or column_width."
        )
    _validate_no_alignment_fields(op, op_name="restore_design_snapshot")
    if op.design_snapshot is None:
        raise ValueError("restore_design_snapshot requires design_snapshot.")


def _validate_create_chart(op: PatchOp) -> None:
    """Validate create_chart operation."""
    _validate_no_legacy_edit_fields(op, op_name="create_chart", allow_chart_fields=True)
    if op.cell is not None or op.range is not None or op.base_cell is not None:
        raise ValueError("create_chart does not accept cell/range/base_cell.")
    if op.row_count is not None or op.col_count is not None:
        raise ValueError("create_chart does not accept row_count or col_count.")
    if (
        op.bold is not None
        or op.color is not None
        or op.fill_color is not None
        or op.font_size is not None
    ):
        raise ValueError("create_chart does not accept style fields.")
    if op.rows is not None or op.columns is not None:
        raise ValueError("create_chart does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError("create_chart does not accept row_height or column_width.")
    _validate_no_alignment_fields(op, op_name="create_chart")
    if op.design_snapshot is not None:
        raise ValueError("create_chart does not accept design_snapshot.")
    if op.chart_type is None:
        raise ValueError("create_chart requires chart_type.")
    if op.data_range is None:
        raise ValueError("create_chart requires data_range.")
    if op.anchor_cell is None:
        raise ValueError("create_chart requires anchor_cell.")
    if op.titles_from_data is None:
        op.titles_from_data = True
    if op.series_from_rows is None:
        op.series_from_rows = False


def _validate_no_legacy_edit_fields(
    op: PatchOp,
    *,
    op_name: str,
    allow_table_fields: bool = False,
    allow_auto_fit_fields: bool = False,
    allow_chart_fields: bool = False,
) -> None:
    """Reject fields that are unrelated to design operations."""
    if op.expected is not None:
        raise ValueError(f"{op_name} does not accept expected.")
    if op.value is not None:
        raise ValueError(f"{op_name} does not accept value.")
    if op.values is not None:
        raise ValueError(f"{op_name} does not accept values.")
    if op.formula is not None:
        raise ValueError(f"{op_name} does not accept formula.")
    if not allow_table_fields:
        if op.style is not None:
            raise ValueError(f"{op_name} does not accept style.")
        if op.table_name is not None:
            raise ValueError(f"{op_name} does not accept table_name.")
    if not allow_auto_fit_fields:
        if op.min_width is not None:
            raise ValueError(f"{op_name} does not accept min_width.")
        if op.max_width is not None:
            raise ValueError(f"{op_name} does not accept max_width.")
    if not allow_chart_fields:
        _reject_optional_field(op_name, "chart_type", op.chart_type)
        _reject_optional_field(op_name, "data_range", op.data_range)
        _reject_optional_field(op_name, "category_range", op.category_range)
        _reject_optional_field(op_name, "anchor_cell", op.anchor_cell)
        _reject_optional_field(op_name, "chart_name", op.chart_name)
        _reject_optional_field(op_name, "width", op.width)
        _reject_optional_field(op_name, "height", op.height)
        _reject_optional_field(op_name, "titles_from_data", op.titles_from_data)
        _reject_optional_field(op_name, "series_from_rows", op.series_from_rows)
        _reject_optional_field(op_name, "chart_title", op.chart_title)
        _reject_optional_field(op_name, "x_axis_title", op.x_axis_title)
        _reject_optional_field(op_name, "y_axis_title", op.y_axis_title)


def _validate_no_design_fields(op: PatchOp, *, op_name: str) -> None:
    """Reject design-only fields for legacy value edit operations."""
    if op.row_count is not None or op.col_count is not None:
        raise ValueError(f"{op_name} does not accept row_count or col_count.")
    if op.rows is not None or op.columns is not None:
        raise ValueError(f"{op_name} does not accept rows or columns.")
    if op.row_height is not None or op.column_width is not None:
        raise ValueError(f"{op_name} does not accept row_height or column_width.")
    _reject_optional_field(op_name, "bold", op.bold)
    _reject_optional_field(op_name, "color", op.color)
    _reject_optional_field(op_name, "font_size", op.font_size)
    _reject_optional_field(op_name, "fill_color", op.fill_color)
    _reject_optional_field(op_name, "style", op.style)
    _reject_optional_field(op_name, "table_name", op.table_name)
    _validate_no_alignment_fields(op, op_name=op_name)
    _reject_optional_field(op_name, "design_snapshot", op.design_snapshot)
    _reject_optional_field(op_name, "min_width", op.min_width)
    _reject_optional_field(op_name, "max_width", op.max_width)
    _reject_optional_field(op_name, "chart_type", op.chart_type)
    _reject_optional_field(op_name, "data_range", op.data_range)
    _reject_optional_field(op_name, "category_range", op.category_range)
    _reject_optional_field(op_name, "anchor_cell", op.anchor_cell)
    _reject_optional_field(op_name, "chart_name", op.chart_name)
    _reject_optional_field(op_name, "width", op.width)
    _reject_optional_field(op_name, "height", op.height)
    _reject_optional_field(op_name, "titles_from_data", op.titles_from_data)
    _reject_optional_field(op_name, "series_from_rows", op.series_from_rows)
    _reject_optional_field(op_name, "chart_title", op.chart_title)
    _reject_optional_field(op_name, "x_axis_title", op.x_axis_title)
    _reject_optional_field(op_name, "y_axis_title", op.y_axis_title)


def _reject_optional_field(op_name: str, field_name: str, value: object) -> None:
    """Raise when an optional field is provided for an unsupported op."""
    if value is not None:
        raise ValueError(f"{op_name} does not accept {field_name}.")


def _validate_no_alignment_fields(op: PatchOp, *, op_name: str) -> None:
    """Reject alignment-only fields for unrelated operations."""
    if op.horizontal_align is not None:
        raise ValueError(f"{op_name} does not accept horizontal_align.")
    if op.vertical_align is not None:
        raise ValueError(f"{op_name} does not accept vertical_align.")
    if op.wrap_text is not None:
        raise ValueError(f"{op_name} does not accept wrap_text.")


def _validate_exactly_one_cell_or_range(op: PatchOp, *, op_name: str) -> None:
    """Ensure exactly one of cell/range is provided."""
    if op.base_cell is not None:
        raise ValueError(f"{op_name} does not accept base_cell.")
    has_cell = op.cell is not None
    has_range = op.range is not None
    if has_cell == has_range:
        raise ValueError(f"{op_name} requires exactly one of cell or range.")


def _validate_style_target_size(op: PatchOp, *, op_name: str) -> None:
    """Guard style edits against accidental huge targets."""
    target_count = 1 if op.cell is not None else _range_cell_count(op.range)
    if target_count > _MAX_STYLE_TARGET_CELLS:
        raise ValueError(
            f"{op_name} target exceeds max cells: {_MAX_STYLE_TARGET_CELLS}."
        )


def _range_cell_count(range_ref: str | None) -> int:
    """Return the number of cells represented by an A1 range."""
    if range_ref is None:
        raise ValueError("range is required.")
    return _shared_range_cell_count(range_ref)


def _split_a1(value: str) -> tuple[str, int]:
    """Split A1 notation into normalized (column_label, row_index)."""
    return _shared_split_a1(value)


def _normalize_column_identifier(value: str | int) -> str | int:
    """Normalize a column identifier preserving letter/index semantics."""
    if isinstance(value, int):
        if value < 1:
            raise ValueError("columns numeric values must be positive.")
        return value
    label = value.strip().upper()
    if not _COLUMN_LABEL_PATTERN.match(label):
        raise ValueError(f"Invalid column identifier: {value}")
    return label


def _column_label_to_index(label: str) -> int:
    """Convert Excel-style column label (A/AA) to 1-based index."""
    return _shared_column_label_to_index(label)


def _column_index_to_label(index: int) -> str:
    """Convert 1-based column index to Excel-style column label."""
    return _shared_column_index_to_label(index)


class PatchValue(BaseModel):
    """Normalized before/after value in patch diff."""

    kind: PatchValueKind
    value: str | int | float | None


class PatchDiffItem(BaseModel):
    """Applied change record for patch operations."""

    op_index: int
    op: PatchOpType
    sheet: str
    cell: str | None = None
    before: PatchValue | None = None
    after: PatchValue | None = None
    status: PatchStatus = "applied"


class PatchErrorDetail(BaseModel):
    """Structured error details for patch failures."""

    op_index: int
    op: PatchOpType
    sheet: str
    cell: str | None
    message: str
    hint: str | None = None
    expected_fields: list[str] = Field(default_factory=list)
    example_op: str | None = None
    error_code: str | None = None
    failed_field: str | None = None
    raw_com_message: str | None = None


class FormulaIssue(BaseModel):
    """Formula health-check finding."""

    sheet: str
    cell: str
    level: FormulaIssueLevel
    code: FormulaIssueCode
    message: str


def _validate_backend_feature_constraints(
    *,
    backend: PatchBackend,
    ops: list[PatchOp],
    dry_run: bool,
    return_inverse_ops: bool,
    preflight_formula_check: bool,
) -> None:
    """Validate backend-specific feature constraints for patch/make requests."""
    has_create_chart = any(op.op == "create_chart" for op in ops)
    if has_create_chart and backend == "openpyxl":
        raise ValueError(
            "create_chart is supported only on COM backend; backend='openpyxl' is not allowed."
        )
    if backend == "com":
        if dry_run or return_inverse_ops or preflight_formula_check:
            raise ValueError(
                "backend='com' does not support dry_run, return_inverse_ops, "
                "or preflight_formula_check."
            )
        if any(op.op == "restore_design_snapshot" for op in ops):
            raise ValueError(
                "backend='com' does not support restore_design_snapshot operation."
            )
    if has_create_chart and (dry_run or return_inverse_ops or preflight_formula_check):
        raise ValueError(
            "create_chart does not support dry_run, return_inverse_ops, or preflight_formula_check."
        )


class PatchRequest(BaseModel):
    """Input model for ExStruct MCP patch."""

    xlsx_path: Path
    ops: list[PatchOp]
    sheet: str | None = None
    out_dir: Path | None = None
    out_name: str | None = None
    on_conflict: OnConflictPolicy = "overwrite"
    auto_formula: bool = False
    dry_run: bool = False
    return_inverse_ops: bool = False
    preflight_formula_check: bool = False
    backend: PatchBackend = "auto"

    @model_validator(mode="after")
    def _validate_backend_constraints(self) -> PatchRequest:
        _validate_backend_feature_constraints(
            backend=self.backend,
            ops=self.ops,
            dry_run=self.dry_run,
            return_inverse_ops=self.return_inverse_ops,
            preflight_formula_check=self.preflight_formula_check,
        )
        return self


class MakeRequest(BaseModel):
    """Input model for ExStruct MCP workbook creation."""

    out_path: Path
    ops: list[PatchOp] = Field(default_factory=list)
    sheet: str | None = None
    on_conflict: OnConflictPolicy = "overwrite"
    auto_formula: bool = False
    dry_run: bool = False
    return_inverse_ops: bool = False
    preflight_formula_check: bool = False
    backend: PatchBackend = "auto"

    @model_validator(mode="after")
    def _validate_backend_constraints(self) -> MakeRequest:
        _validate_backend_feature_constraints(
            backend=self.backend,
            ops=self.ops,
            dry_run=self.dry_run,
            return_inverse_ops=self.return_inverse_ops,
            preflight_formula_check=self.preflight_formula_check,
        )
        return self


class PatchResult(BaseModel):
    """Output model for ExStruct MCP patch."""

    out_path: str
    patch_diff: list[PatchDiffItem] = Field(default_factory=list)
    inverse_ops: list[PatchOp] = Field(default_factory=list)
    formula_issues: list[FormulaIssue] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    error: PatchErrorDetail | None = None
    engine: PatchEngine


def run_make(request: MakeRequest, *, policy: PathPolicy | None = None) -> PatchResult:
    """Create a new workbook and apply patch operations in one call.

    Args:
        request: Workbook creation request payload.
        policy: Optional path policy for access control.

    Returns:
        Patch-compatible result with output path and diff.

    Raises:
        ValueError: If request validation fails.
        RuntimeError: If backend operations fail.
    """
    from .service import run_make as _service_run_make

    return cast(PatchResult, _service_run_make(cast(Any, request), policy=policy))


def run_patch(
    request: PatchRequest, *, policy: PathPolicy | None = None
) -> PatchResult:
    """Run a patch operation and write the updated workbook.

    Args:
        request: Patch request payload.
        policy: Optional path policy for access control.

    Returns:
        Patch result with output path and diff.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If validation fails or the path violates policy.
        RuntimeError: If a backend operation fails.
    """
    from .service import run_patch as _service_run_patch

    return cast(PatchResult, _service_run_patch(cast(Any, request), policy=policy))


def _apply_with_openpyxl(
    request: PatchRequest,
    input_path: Path,
    output_path: Path,
    warnings: list[str],
) -> PatchResult:
    """Apply patch operations using openpyxl."""
    try:
        diff, inverse_ops, formula_issues, op_warnings = _apply_ops_openpyxl(
            request,
            input_path,
            output_path,
        )
    except PatchOpError as exc:
        return PatchResult(
            out_path=str(output_path),
            patch_diff=[],
            inverse_ops=[],
            formula_issues=[],
            warnings=warnings,
            error=exc.detail,
            engine="openpyxl",
        )
    except ValueError:
        raise
    except FileNotFoundError:
        raise
    except OSError:
        raise
    except Exception as exc:
        raise RuntimeError(f"openpyxl patch failed: {exc}") from exc

    warnings.extend(op_warnings)
    if not request.dry_run:
        warnings.append(
            "openpyxl editing may drop shapes/charts or unsupported elements."
        )
    _append_skip_warnings(warnings, diff)
    if (
        not request.dry_run
        and request.preflight_formula_check
        and any(issue.level == "error" for issue in formula_issues)
    ):
        issue = formula_issues[0]
        op_index, op_name = _find_preflight_issue_origin(issue, request.ops)
        error = PatchErrorDetail(
            op_index=op_index,
            op=op_name,
            sheet=issue.sheet,
            cell=issue.cell,
            message=f"Formula health check failed: {issue.message}",
            hint=None,
            expected_fields=[],
            example_op=None,
        )
        return PatchResult(
            out_path=str(output_path),
            patch_diff=[],
            inverse_ops=[],
            formula_issues=formula_issues,
            warnings=warnings,
            error=error,
            engine="openpyxl",
        )
    return PatchResult(
        out_path=str(output_path),
        patch_diff=diff,
        inverse_ops=inverse_ops,
        formula_issues=formula_issues,
        warnings=warnings,
        engine="openpyxl",
    )


def _append_skip_warnings(warnings: list[str], diff: list[PatchDiffItem]) -> None:
    """Append warning messages for skipped conditional operations."""
    for item in diff:
        if item.status != "skipped":
            continue
        warnings.append(
            f"Skipped op[{item.op_index}] {item.op} at {item.sheet}!{item.cell} due to condition mismatch."
        )


def _find_preflight_issue_origin(
    issue: FormulaIssue, ops: list[PatchOp]
) -> tuple[int, PatchOpType]:
    """Find the most likely op index/op name for a preflight formula issue."""
    for index, op in enumerate(ops):
        if _op_targets_issue_cell(op, issue.sheet, issue.cell):
            return index, op.op
    return -1, "set_value"


def _op_targets_issue_cell(op: PatchOp, sheet: str, cell: str) -> bool:
    """Return True when an op can affect the specified sheet/cell."""
    if op.sheet != sheet:
        return False
    if op.cell is not None:
        return op.cell == cell
    if op.range is None:
        return False
    for row in _expand_range_coordinates(op.range):
        if cell in row:
            return True
    return False


def _allow_auto_openpyxl_fallback(request: PatchRequest, input_path: Path) -> bool:
    """Return True when COM failure can fallback to openpyxl."""
    if request.backend != "auto":
        return False
    if _contains_create_chart_op(request.ops):
        return False
    return input_path.suffix.lower() in {".xlsx", ".xlsm"}


def _requires_openpyxl_backend(request: PatchRequest) -> bool:
    """Return True if request requires openpyxl backend for extended features."""
    if request.dry_run or request.return_inverse_ops or request.preflight_formula_check:
        return True
    return any(op.op == "restore_design_snapshot" for op in request.ops)


def _raise_create_chart_com_unavailable_error(
    *,
    has_apply_table_style: bool,
) -> None:
    """Raise a COM availability error for create_chart requests."""
    if has_apply_table_style:
        raise ValueError(
            "create_chart + apply_table_style requests require Windows Excel COM availability in this environment."
        )
    raise ValueError(
        "create_chart requires Windows Excel COM availability in this environment."
    )


def _select_patch_engine(
    *, request: PatchRequest, input_path: Path, com_available: bool
) -> PatchEngine:
    """Select concrete patch engine based on request and environment."""
    extension = input_path.suffix.lower()
    has_create_chart = _contains_create_chart_op(request.ops)
    has_apply_table_style = _contains_apply_table_style_op(request.ops)
    if request.backend == "openpyxl":
        if has_create_chart:
            raise ValueError("create_chart is supported only on COM backend.")
        if extension == ".xls":
            raise ValueError("backend='openpyxl' cannot edit .xls files.")
        return "openpyxl"
    if request.backend == "com":
        if not com_available:
            raise ValueError("backend='com' requires Windows Excel COM availability.")
        return "com"
    if extension == ".xls":
        if not com_available:
            raise ValueError(
                ".xls editing requires Windows Excel COM (xlwings) in this environment."
            )
        return "com"
    if _requires_openpyxl_backend(request):
        if has_create_chart:
            raise ValueError(
                "create_chart does not support dry_run, return_inverse_ops, or preflight_formula_check."
            )
        return "openpyxl"
    if com_available:
        return "com"
    if has_create_chart:
        _raise_create_chart_com_unavailable_error(
            has_apply_table_style=has_apply_table_style
        )
    return "openpyxl"


def _contains_design_ops(ops: list[PatchOp]) -> bool:
    """Return True when any style/dimension design operation is present."""
    design_ops = {
        "draw_grid_border",
        "set_bold",
        "set_font_size",
        "set_font_color",
        "set_fill_color",
        "set_dimensions",
        "auto_fit_columns",
        "merge_cells",
        "unmerge_cells",
        "set_alignment",
        "set_style",
        "apply_table_style",
        "restore_design_snapshot",
    }
    return any(op.op in design_ops for op in ops)


def _contains_apply_table_style_op(ops: list[PatchOp]) -> bool:
    """Return True when apply_table_style is present."""
    return any(op.op == "apply_table_style" for op in ops)


def _contains_create_chart_op(ops: list[PatchOp]) -> bool:
    """Return True when create_chart is present."""
    return any(op.op == "create_chart" for op in ops)


def _append_large_ops_warning(warnings: list[str], ops: list[PatchOp]) -> None:
    """Append warning when operation count exceeds the soft threshold."""
    if len(ops) <= _SOFT_MAX_OPS_WARNING_THRESHOLD:
        return
    warnings.append(
        "Large patch request: "
        f"{len(ops)} ops. Recommended maximum is "
        f"{_SOFT_MAX_OPS_WARNING_THRESHOLD}; consider splitting into batches."
    )


def _resolve_input_path(path: Path, *, policy: PathPolicy | None) -> Path:
    """Resolve and validate the input path."""
    resolved = policy.ensure_allowed(path) if policy else path.resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Input file not found: {resolved}")
    if not resolved.is_file():
        raise ValueError(f"Input path is not a file: {resolved}")
    return resolved


def _ensure_supported_extension(path: Path) -> None:
    """Validate that the input file extension is supported."""
    if path.suffix.lower() not in _ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension: {path.suffix}")


def _resolve_output_path(
    input_path: Path,
    *,
    out_dir: Path | None,
    out_name: str | None,
    policy: PathPolicy | None,
) -> Path:
    """Build and validate the output path."""
    return _shared_resolve_output_path(
        input_path,
        out_dir=out_dir,
        out_name=out_name,
        policy=policy,
        default_suffix=input_path.suffix,
        default_name_builder="patched",
    )


def _resolve_make_output_path(path: Path, *, policy: PathPolicy | None) -> Path:
    """Resolve and validate output path for workbook creation."""
    resolved = policy.ensure_allowed(path) if policy else path.resolve()
    if resolved.exists() and resolved.is_dir():
        raise ValueError(f"Output path is a directory: {resolved}")
    return resolved


def _validate_make_request_constraints(request: MakeRequest, output_path: Path) -> None:
    """Validate make-specific constraints by output extension."""
    if output_path.suffix.lower() != ".xls":
        return
    if request.backend == "openpyxl":
        raise ValueError("backend='openpyxl' cannot edit .xls files.")
    if request.dry_run or request.return_inverse_ops or request.preflight_formula_check:
        raise ValueError(
            ".xls creation does not support dry_run, return_inverse_ops, "
            "or preflight_formula_check."
        )
    com = get_com_availability()
    if not com.available:
        raise ValueError(
            ".xls editing requires Windows Excel COM (xlwings) in this environment."
        )


def _build_make_seed_path(output_path: Path) -> Path:
    """Return a temporary seed path in the target output directory."""
    seed_name = f".exstruct_make_seed_{uuid4().hex}{output_path.suffix.lower()}"
    return output_path.parent / seed_name


def _resolve_make_initial_sheet_name(request: MakeRequest) -> str:
    """Resolve initial sheet name for `exstruct_make` seed workbook."""
    if request.sheet is None:
        return "Sheet1"
    requested_sheet = request.sheet.strip()
    if not requested_sheet:
        return "Sheet1"
    normalized_requested_sheet = _normalize_sheet_name_for_make_conflict(
        requested_sheet
    )
    has_conflicting_add_sheet = any(
        op.op == "add_sheet"
        and _normalize_sheet_name_for_make_conflict(op.sheet)
        == normalized_requested_sheet
        for op in request.ops
    )
    if has_conflicting_add_sheet:
        return "Sheet1"
    return requested_sheet


def _normalize_sheet_name_for_make_conflict(sheet_name: str) -> str:
    """Normalize sheet name text for make-time conflict detection."""
    return sheet_name.strip().casefold()


def _create_seed_workbook(
    seed_path: Path, extension: str, *, initial_sheet_name: str
) -> None:
    """Create an empty workbook seed with the resolved initial sheet name."""
    _ensure_output_dir(seed_path)
    if extension == ".xls":
        _create_xls_seed_with_com(seed_path, initial_sheet_name=initial_sheet_name)
        return
    _create_openpyxl_seed(seed_path, initial_sheet_name=initial_sheet_name)


def _create_openpyxl_seed(seed_path: Path, *, initial_sheet_name: str) -> None:
    """Create an empty workbook via openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    workbook = Workbook()
    try:
        active_sheet = workbook.active
        if active_sheet is None:
            raise RuntimeError("Failed to create default worksheet.")
        active_sheet.title = initial_sheet_name
        workbook.save(seed_path)
    finally:
        workbook.close()


def _create_xls_seed_with_com(seed_path: Path, *, initial_sheet_name: str) -> None:
    """Create an empty .xls workbook via Excel COM."""
    com = get_com_availability()
    if not com.available:
        raise ValueError(
            ".xls editing requires Windows Excel COM (xlwings) in this environment."
        )
    app = xw.App(add_book=False, visible=False)
    app.display_alerts = False
    app.screen_updating = False
    workbook = app.books.add()
    try:
        workbook.sheets[0].name = initial_sheet_name
        workbook.save(str(seed_path))
    except Exception as exc:
        raise RuntimeError(f"COM workbook creation failed: {exc}") from exc
    finally:
        _close_workbook_safely(workbook)
        _quit_app_safely(app)


def _normalize_output_name(input_path: Path, out_name: str | None) -> str:
    """Normalize output filename with a safe suffix."""
    if out_name:
        candidate = Path(out_name)
        return (
            candidate.name
            if candidate.suffix
            else f"{candidate.name}{input_path.suffix}"
        )
    return f"{input_path.stem}_patched{input_path.suffix}"


def _ensure_output_dir(path: Path) -> None:
    """Ensure the output directory exists before writing."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _apply_conflict_policy(
    output_path: Path, on_conflict: OnConflictPolicy
) -> tuple[Path, str | None, bool]:
    """Apply output conflict policy to a resolved output path."""
    return _shared_apply_conflict_policy(output_path, on_conflict)


def _next_available_path(path: Path) -> Path:
    """Return the next available path by appending a numeric suffix."""
    return _shared_next_available_path(path)


def _apply_ops_openpyxl(
    request: PatchRequest,
    input_path: Path,
    output_path: Path,
) -> tuple[list[PatchDiffItem], list[PatchOp], list[FormulaIssue], list[str]]:
    """Apply operations using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    if input_path.suffix.lower() == ".xls":
        raise ValueError("openpyxl cannot edit .xls files.")

    if input_path.suffix.lower() == ".xlsm":
        workbook = load_workbook(input_path, keep_vba=True)
    else:
        workbook = load_workbook(input_path)
    try:
        diff, inverse_ops, op_warnings = _apply_ops_to_openpyxl_workbook(
            workbook,
            request.ops,
            request.auto_formula,
            return_inverse_ops=request.return_inverse_ops,
        )
        formula_issues = (
            _collect_formula_issues_openpyxl(workbook)
            if request.preflight_formula_check
            else []
        )
        if not request.dry_run and not (
            request.preflight_formula_check
            and any(issue.level == "error" for issue in formula_issues)
        ):
            workbook.save(output_path)
    finally:
        workbook.close()
    return diff, inverse_ops, formula_issues, op_warnings


def _apply_ops_to_openpyxl_workbook(
    workbook: OpenpyxlWorkbookProtocol,
    ops: list[PatchOp],
    auto_formula: bool,
    *,
    return_inverse_ops: bool,
) -> tuple[list[PatchDiffItem], list[PatchOp], list[str]]:
    """Apply ops to an openpyxl workbook instance."""
    sheets = _openpyxl_sheet_map(workbook)
    diff: list[PatchDiffItem] = []
    inverse_ops: list[PatchOp] = []
    op_warnings: list[str] = []
    for index, op in enumerate(ops):
        try:
            item, inverse = _apply_openpyxl_op(
                workbook, sheets, op, index, auto_formula, op_warnings
            )
            diff.append(item)
            if return_inverse_ops and item.status == "applied" and inverse is not None:
                inverse_ops.append(inverse)
        except ValueError as exc:
            raise PatchOpError.from_op(index, op, exc) from exc
    if return_inverse_ops:
        inverse_ops.reverse()
    return diff, inverse_ops, op_warnings


def _openpyxl_sheet_map(
    workbook: OpenpyxlWorkbookProtocol,
) -> dict[str, OpenpyxlWorksheetProtocol]:
    """Build a sheet map for openpyxl workbooks."""
    sheet_names = getattr(workbook, "sheetnames", None)
    if not isinstance(sheet_names, list):
        raise ValueError("Invalid workbook: sheetnames missing.")
    return {name: workbook[name] for name in sheet_names}


def _apply_openpyxl_op(
    workbook: OpenpyxlWorkbookProtocol,
    sheets: dict[str, OpenpyxlWorksheetProtocol],
    op: PatchOp,
    index: int,
    auto_formula: bool,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply a single op to openpyxl workbook."""
    if op.op == "add_sheet":
        return _apply_openpyxl_add_sheet(workbook, sheets, op, index)

    existing_sheet = sheets.get(op.sheet)
    if existing_sheet is None:
        raise ValueError(f"Sheet not found: {op.sheet}")
    return _apply_openpyxl_sheet_op(
        existing_sheet,
        op,
        index,
        auto_formula=auto_formula,
        warnings=warnings,
    )


def _apply_openpyxl_sheet_op(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    *,
    auto_formula: bool,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply openpyxl operation that targets an existing sheet."""
    if op.op in {"set_value", "set_formula", "set_value_if", "set_formula_if"}:
        return _apply_openpyxl_cell_op(sheet, op, index, auto_formula)
    handlers: dict[PatchOpType, Callable[[], tuple[PatchDiffItem, PatchOp | None]]] = {
        "set_range_values": lambda: _apply_openpyxl_set_range_values(sheet, op, index),
        "fill_formula": lambda: _apply_openpyxl_fill_formula(sheet, op, index),
        "draw_grid_border": lambda: _apply_openpyxl_draw_grid_border(sheet, op, index),
        "set_bold": lambda: _apply_openpyxl_set_bold(sheet, op, index),
        "set_font_size": lambda: _apply_openpyxl_set_font_size(sheet, op, index),
        "set_font_color": lambda: _apply_openpyxl_set_font_color(sheet, op, index),
        "set_fill_color": lambda: _apply_openpyxl_set_fill_color(sheet, op, index),
        "set_dimensions": lambda: _apply_openpyxl_set_dimensions(sheet, op, index),
        "auto_fit_columns": lambda: _apply_openpyxl_auto_fit_columns(sheet, op, index),
        "merge_cells": lambda: _apply_openpyxl_merge_cells(sheet, op, index, warnings),
        "unmerge_cells": lambda: _apply_openpyxl_unmerge_cells(sheet, op, index),
        "set_alignment": lambda: _apply_openpyxl_set_alignment(sheet, op, index),
        "set_style": lambda: _apply_openpyxl_set_style(sheet, op, index),
        "apply_table_style": lambda: _apply_openpyxl_apply_table_style(
            sheet, op, index
        ),
        "create_chart": lambda: _apply_openpyxl_create_chart(op),
        "restore_design_snapshot": lambda: _apply_openpyxl_restore_design_snapshot(
            sheet, op, index
        ),
    }
    handler = handlers.get(op.op)
    if handler is None:
        raise ValueError(f"Unsupported op: {op.op}")
    return handler()


def _apply_openpyxl_add_sheet(
    workbook: OpenpyxlWorkbookProtocol,
    sheets: dict[str, OpenpyxlWorksheetProtocol],
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply add_sheet op."""
    if op.sheet in sheets:
        raise ValueError(f"Sheet already exists: {op.sheet}")
    sheet = workbook.create_sheet(title=op.sheet)
    sheets[op.sheet] = sheet
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="sheet", value=op.sheet),
        ),
        None,
    )


def _apply_openpyxl_set_range_values(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_range_values op."""
    if op.range is None or op.values is None:
        raise ValueError("set_range_values requires range and values.")
    coordinates = _expand_range_coordinates(op.range)
    rows, cols = _shape_of_coordinates(coordinates)
    if len(op.values) != rows:
        raise ValueError("set_range_values values height does not match range.")
    if any(len(row) != cols for row in op.values):
        raise ValueError("set_range_values values width does not match range.")
    for r_idx, row in enumerate(coordinates):
        for c_idx, coord in enumerate(row):
            sheet[coord].value = op.values[r_idx][c_idx]
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="value", value=f"{rows}x{cols}"),
        ),
        None,
    )


def _apply_openpyxl_fill_formula(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply fill_formula op."""
    if op.range is None or op.formula is None or op.base_cell is None:
        raise ValueError("fill_formula requires range, base_cell and formula.")
    coordinates = _expand_range_coordinates(op.range)
    rows, cols = _shape_of_coordinates(coordinates)
    if rows != 1 and cols != 1:
        raise ValueError("fill_formula range must be a single row or a single column.")
    for row in coordinates:
        for coord in row:
            sheet[coord].value = _translate_formula(op.formula, op.base_cell, coord)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="formula", value=op.formula),
        ),
        None,
    )


def _apply_openpyxl_draw_grid_border(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply draw_grid_border op with thin black border."""
    if op.base_cell is None or op.row_count is None or op.col_count is None:
        raise ValueError(
            "draw_grid_border requires base_cell, row_count and col_count."
        )
    coordinates = _expand_rect_coordinates(op.base_cell, op.row_count, op.col_count)
    snapshot = DesignSnapshot(
        borders=[_snapshot_border(sheet[coord], coord) for coord in coordinates]
    )
    for coord in coordinates:
        _set_grid_border(sheet[coord])
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=f"{op.base_cell}:{coordinates[-1]}",
            before=None,
            after=PatchValue(kind="style", value="grid_border(thin,black)"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_bold(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_bold op."""
    targets = _resolve_style_targets(op)
    target_bold = True if op.bold is None else op.bold
    snapshot = DesignSnapshot(
        fonts=[_snapshot_font(sheet[coord], coord) for coord in targets]
    )
    for coord in targets:
        cell = sheet[coord]
        font = copy(cell.font)
        font.bold = target_bold
        cell.font = font
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"bold={target_bold}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_font_size(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_font_size op."""
    if op.font_size is None:
        raise ValueError("set_font_size requires font_size.")
    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        fonts=[_snapshot_font(sheet[coord], coord) for coord in targets]
    )
    for coord in targets:
        cell = sheet[coord]
        font = copy(cell.font)
        font.size = op.font_size
        cell.font = font
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"font_size={op.font_size}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_font_color(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_font_color op."""
    if op.color is None:
        raise ValueError("set_font_color requires color.")
    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        fonts=[_snapshot_font(sheet[coord], coord) for coord in targets]
    )
    normalized = _normalize_hex_color(op.color)
    for coord in targets:
        cell = sheet[coord]
        font = copy(cell.font)
        font.color = normalized
        cell.font = font
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"font_color={op.color}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_fill_color(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_fill_color op."""
    if op.fill_color is None:
        raise ValueError("set_fill_color requires fill_color.")
    try:
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        fills=[_snapshot_fill(sheet[coord], coord) for coord in targets]
    )
    normalized = _normalize_hex_color(op.fill_color)
    for coord in targets:
        sheet[coord].fill = PatternFill(
            fill_type="solid",
            start_color=normalized,
            end_color=normalized,
        )
    location = op.cell if op.cell is not None else op.range
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=f"fill={op.fill_color}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_dimensions(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_dimensions op."""
    snapshot = DesignSnapshot()
    parts: list[str] = []
    if op.rows is not None and op.row_height is not None:
        for row in op.rows:
            row_dimension = sheet.row_dimensions[row]
            snapshot.row_dimensions.append(
                RowDimensionSnapshot(
                    row=row,
                    height=getattr(row_dimension, "height", None),
                )
            )
            row_dimension.height = op.row_height
        parts.append(f"rows={_summarize_int_targets(op.rows)}")
    if op.columns is not None and op.column_width is not None:
        normalized_columns = _normalize_columns_for_dimensions(op.columns)
        for column in normalized_columns:
            column_dimension = sheet.column_dimensions[column]
            snapshot.column_dimensions.append(
                ColumnDimensionSnapshot(
                    column=column,
                    width=getattr(column_dimension, "width", None),
                )
            )
            column_dimension.width = op.column_width
        parts.append(f"columns={_summarize_column_targets(normalized_columns)}")
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="dimension", value=", ".join(parts)),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_auto_fit_columns(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply auto_fit_columns op using openpyxl text-length estimation."""
    target_columns = _resolve_auto_fit_columns_openpyxl(sheet, op.columns)
    if not target_columns:
        raise ValueError("auto_fit_columns could not resolve target columns.")
    target_column_indexes = {
        _column_label_to_index(column) for column in target_columns
    }
    max_lengths = _collect_openpyxl_target_column_max_lengths(
        sheet, target_column_indexes
    )
    snapshot = DesignSnapshot()
    for column in target_columns:
        column_dimension = sheet.column_dimensions[column]
        snapshot.column_dimensions.append(
            ColumnDimensionSnapshot(
                column=column,
                width=getattr(column_dimension, "width", None),
            )
        )
        max_len = max_lengths.get(_column_label_to_index(column), 0)
        estimated_width = _resolve_openpyxl_estimated_width(column_dimension, max_len)
        column_dimension.width = _clamp_column_width(
            estimated_width, min_width=op.min_width, max_width=op.max_width
        )
    parts = [f"columns={_summarize_column_targets(target_columns)}"]
    if op.min_width is not None:
        parts.append(f"min_width={op.min_width}")
    if op.max_width is not None:
        parts.append(f"max_width={op.max_width}")
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="dimension", value=", ".join(parts)),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_merge_cells(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    warnings: list[str],
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply merge_cells op."""
    if op.range is None:
        raise ValueError("merge_cells requires range.")
    overlapped = _intersecting_merged_ranges(sheet, op.range)
    if overlapped:
        raise ValueError(
            "merge_cells range overlaps existing merged ranges: "
            + ", ".join(overlapped)
            + "."
        )
    merge_warning = _build_merge_value_loss_warning(sheet, op.sheet, op.range)
    if merge_warning is not None:
        warnings.append(merge_warning)
    snapshot = DesignSnapshot(
        merge_state=MergeStateSnapshot(scope=op.range, ranges=[]),
    )
    sheet.merge_cells(op.range)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="style", value=f"merged={op.range}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_unmerge_cells(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply unmerge_cells op."""
    if op.range is None:
        raise ValueError("unmerge_cells requires range.")
    target_ranges = _intersecting_merged_ranges(sheet, op.range)
    snapshot = DesignSnapshot(
        merge_state=MergeStateSnapshot(scope=op.range, ranges=target_ranges),
    )
    for range_ref in target_ranges:
        sheet.unmerge_cells(range_ref)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(kind="style", value=f"unmerged={len(target_ranges)}"),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_alignment(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_alignment op."""
    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        alignments=[_snapshot_alignment(sheet[coord], coord) for coord in targets]
    )
    for coord in targets:
        cell = sheet[coord]
        alignment = copy(cell.alignment)
        if op.horizontal_align is not None:
            alignment.horizontal = op.horizontal_align
        if op.vertical_align is not None:
            alignment.vertical = op.vertical_align
        if op.wrap_text is not None:
            alignment.wrap_text = op.wrap_text
        cell.alignment = alignment
    location = op.cell if op.cell is not None else op.range
    summary = (
        f"horizontal={op.horizontal_align},"
        f"vertical={op.vertical_align},"
        f"wrap_text={op.wrap_text}"
    )
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=summary),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_set_style(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply set_style op."""
    targets = _resolve_style_targets(op)
    snapshot = DesignSnapshot(
        fonts=[_snapshot_font(sheet[coord], coord) for coord in targets],
        fills=[_snapshot_fill(sheet[coord], coord) for coord in targets],
        alignments=[_snapshot_alignment(sheet[coord], coord) for coord in targets],
    )
    font_color = _normalize_hex_color(op.color) if op.color is not None else None
    fill_color = (
        _normalize_hex_color(op.fill_color) if op.fill_color is not None else None
    )
    pattern_fill_factory: Callable[..., OpenpyxlFillProtocol] | None = None
    if fill_color is not None:
        try:
            from openpyxl.styles import PatternFill
        except ImportError as exc:
            raise RuntimeError(f"openpyxl is not available: {exc}") from exc
        pattern_fill_factory = PatternFill
    for coord in targets:
        cell = sheet[coord]
        font = copy(cell.font)
        if op.bold is not None:
            font.bold = op.bold
        if op.font_size is not None:
            font.size = op.font_size
        if font_color is not None:
            font.color = font_color
        cell.font = font
        if fill_color is not None and pattern_fill_factory is not None:
            cell.fill = pattern_fill_factory(
                fill_type="solid",
                start_color=fill_color,
                end_color=fill_color,
            )
        if (
            op.horizontal_align is not None
            or op.vertical_align is not None
            or op.wrap_text is not None
        ):
            alignment = copy(cell.alignment)
            if op.horizontal_align is not None:
                alignment.horizontal = op.horizontal_align
            if op.vertical_align is not None:
                alignment.vertical = op.vertical_align
            if op.wrap_text is not None:
                alignment.wrap_text = op.wrap_text
            cell.alignment = alignment
    location = op.cell if op.cell is not None else op.range
    parts = _build_set_style_summary_parts(op)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=location,
            before=None,
            after=PatchValue(kind="style", value=";".join(parts)),
        ),
        _build_restore_snapshot_op(op.sheet, snapshot),
    )


def _apply_openpyxl_apply_table_style(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply apply_table_style op."""
    if op.range is None or op.style is None:
        raise ValueError("apply_table_style requires range and style.")
    try:
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    _ensure_range_not_intersects_existing_tables(sheet, op.range)
    table_name = op.table_name or _next_openpyxl_table_name(sheet)
    _ensure_table_name_available(sheet, table_name)
    table = Table(displayName=table_name, ref=op.range)
    table.tableStyleInfo = TableStyleInfo(
        name=op.style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    add_table = getattr(sheet, "add_table", None)
    if not callable(add_table):
        raise ValueError("apply_table_style requires worksheet.add_table support.")
    add_table(table)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.range,
            before=None,
            after=PatchValue(
                kind="style",
                value=f"table={table_name};table_style={op.style}",
            ),
        ),
        None,
    )


def _apply_openpyxl_restore_design_snapshot(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply restore_design_snapshot op."""
    if op.design_snapshot is None:
        raise ValueError("restore_design_snapshot requires design_snapshot.")
    _restore_design_snapshot(sheet, op.design_snapshot)
    return (
        PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="style", value="design_snapshot_restored"),
        ),
        None,
    )


def _apply_openpyxl_create_chart(op: PatchOp) -> tuple[PatchDiffItem, PatchOp | None]:
    """Reject create_chart on openpyxl backend."""
    raise ValueError(
        f"create_chart is supported only on COM backend (sheet={op.sheet})."
    )


def _apply_openpyxl_cell_op(
    sheet: OpenpyxlWorksheetProtocol,
    op: PatchOp,
    index: int,
    auto_formula: bool,
) -> tuple[PatchDiffItem, PatchOp | None]:
    """Apply single-cell operations."""
    cell_ref = op.cell
    if cell_ref is None:
        raise ValueError(f"{op.op} requires cell.")
    cell = sheet[cell_ref]
    before = _openpyxl_cell_value(cell)

    if op.op == "set_value":
        after = _set_cell_value(cell, op.value, auto_formula, op_name="set_value")
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    if op.op == "set_formula":
        formula = _require_formula(op.formula, "set_formula")
        cell.value = formula
        after = PatchValue(kind="formula", value=formula)
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    if op.op == "set_value_if":
        if not _values_equal_for_condition(
            _patch_value_to_primitive(before), op.expected
        ):
            return _build_skipped_result(op, index, cell_ref, before), None
        after = _set_cell_value(cell, op.value, auto_formula, op_name="set_value_if")
        return _build_cell_result(
            op, index, cell_ref, before, after
        ), _build_inverse_cell_op(op, cell_ref, before)
    formula_if = _require_formula(op.formula, "set_formula_if")
    if not _values_equal_for_condition(_patch_value_to_primitive(before), op.expected):
        return _build_skipped_result(op, index, cell_ref, before), None
    cell.value = formula_if
    after = PatchValue(kind="formula", value=formula_if)
    return _build_cell_result(
        op, index, cell_ref, before, after
    ), _build_inverse_cell_op(op, cell_ref, before)


def _set_cell_value(
    cell: OpenpyxlCellProtocol,
    value: str | int | float | None,
    auto_formula: bool,
    *,
    op_name: str,
) -> PatchValue:
    """Set cell value with auto_formula handling."""
    if isinstance(value, str) and value.startswith("="):
        if not auto_formula:
            raise ValueError(f"{op_name} rejects values starting with '='.")
        cell.value = value
        return PatchValue(kind="formula", value=value)
    cell.value = value
    return PatchValue(kind="value", value=value)


def _build_cell_result(
    op: PatchOp,
    index: int,
    cell_ref: str,
    before: PatchValue | None,
    after: PatchValue | None,
) -> PatchDiffItem:
    """Build applied diff item for single-cell op."""
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=cell_ref,
        before=before,
        after=after,
    )


def _build_skipped_result(
    op: PatchOp,
    index: int,
    cell_ref: str,
    before: PatchValue | None,
) -> PatchDiffItem:
    """Build skipped diff item."""
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=cell_ref,
        before=before,
        after=before,
        status="skipped",
    )


def _build_set_style_summary_parts(op: PatchOp) -> list[str]:
    """Build summary parts for set_style diff output."""
    parts: list[str] = []
    if op.bold is not None:
        parts.append(f"bold={op.bold}")
    if op.font_size is not None:
        parts.append(f"font_size={op.font_size}")
    if op.color is not None:
        parts.append(f"color={_normalize_hex_input(op.color, field_name='color')}")
    if op.fill_color is not None:
        parts.append(
            f"fill_color={_normalize_hex_input(op.fill_color, field_name='fill_color')}"
        )
    if op.horizontal_align is not None:
        parts.append(f"horizontal_align={op.horizontal_align}")
    if op.vertical_align is not None:
        parts.append(f"vertical_align={op.vertical_align}")
    if op.wrap_text is not None:
        parts.append(f"wrap_text={op.wrap_text}")
    return parts


def _ensure_range_not_intersects_existing_tables(
    sheet: OpenpyxlWorksheetProtocol, range_ref: str
) -> None:
    """Raise ValueError if range intersects with existing table ranges."""
    for table_name, existing_ref in _collect_openpyxl_table_ranges(sheet):
        if _ranges_overlap(range_ref, existing_ref):
            raise ValueError(
                "apply_table_style range intersects existing table "
                f"'{table_name}' ({existing_ref})."
            )


def _ensure_table_name_available(
    sheet: OpenpyxlWorksheetProtocol, table_name: str
) -> None:
    """Raise ValueError when table name already exists in sheet."""
    existing_names = {name for name, _ in _collect_openpyxl_table_ranges(sheet)}
    if table_name in existing_names:
        raise ValueError(f"Table name already exists: {table_name}")


def _next_openpyxl_table_name(sheet: OpenpyxlWorksheetProtocol) -> str:
    """Generate next available table name like Table1, Table2, ..."""
    existing_names = {name for name, _ in _collect_openpyxl_table_ranges(sheet)}
    for index in range(1, 10_000):
        candidate = f"Table{index}"
        if candidate not in existing_names:
            return candidate
    raise RuntimeError("Failed to generate unique table name.")


def _collect_openpyxl_table_ranges(
    sheet: OpenpyxlWorksheetProtocol,
) -> list[tuple[str, str]]:
    """Collect (table_name, range_ref) pairs from worksheet tables."""
    tables = getattr(sheet, "tables", None)
    if tables is None or not isinstance(tables, OpenpyxlTablesProtocol):
        return []
    pairs: list[tuple[str, str]] = []
    for key, value in tables.items():
        table_name = str(getattr(value, "displayName", key))
        ref_raw = getattr(value, "ref", None)
        if isinstance(ref_raw, str):
            pairs.append((table_name, ref_raw))
            continue
        if isinstance(value, str):
            pairs.append((str(key), value))
    return pairs


def _require_formula(formula: str | None, op_name: str) -> str:
    """Require a non-null formula string."""
    if formula is None:
        raise ValueError(f"{op_name} requires formula.")
    return formula


def _openpyxl_cell_value(cell: OpenpyxlCellProtocol) -> PatchValue | None:
    """Normalize an openpyxl cell value into PatchValue."""
    value = getattr(cell, "value", None)
    if value is None:
        return None
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        text = _normalize_formula(value)
        return PatchValue(kind="formula", value=text)
    return PatchValue(kind="value", value=value)


def _normalize_formula(value: object) -> str:
    """Ensure formula string starts with '='."""
    text = str(value)
    return text if text.startswith("=") else f"={text}"


def _expand_range_coordinates(range_ref: str) -> list[list[str]]:
    """Expand A1 range string into a 2D list of coordinates."""
    try:
        from openpyxl.utils.cell import get_column_letter, range_boundaries
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"Invalid range reference: {range_ref}")
    rows: list[list[str]] = []
    for row_idx in range(min_row, max_row + 1):
        row: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            row.append(f"{get_column_letter(col_idx)}{row_idx}")
        rows.append(row)
    return rows


def _shape_of_coordinates(coordinates: list[list[str]]) -> tuple[int, int]:
    """Return rows/cols for expanded coordinates."""
    if not coordinates or not coordinates[0]:
        raise ValueError("Range expansion resulted in an empty coordinate set.")
    return len(coordinates), len(coordinates[0])


def _expand_rect_coordinates(base_cell: str, rows: int, cols: int) -> list[str]:
    """Expand base cell + size into a flat coordinate list."""
    base_column, base_row = _split_a1(base_cell)
    start_col = _column_label_to_index(base_column)
    coordinates: list[str] = []
    for row_offset in range(rows):
        for col_offset in range(cols):
            column = _column_index_to_label(start_col + col_offset)
            coordinates.append(f"{column}{base_row + row_offset}")
    return coordinates


def _resolve_style_targets(op: PatchOp) -> list[str]:
    """Resolve style operation target coordinates."""
    if op.cell is not None:
        return [op.cell]
    if op.range is None:
        raise ValueError(f"{op.op} requires cell or range.")
    coordinates = _expand_range_coordinates(op.range)
    targets: list[str] = []
    for row in coordinates:
        targets.extend(row)
    return targets


def _merged_range_strings(sheet: OpenpyxlWorksheetProtocol) -> list[str]:
    """Return normalized merged range strings from worksheet."""
    merged_cells = getattr(sheet, "merged_cells", None)
    ranges = getattr(merged_cells, "ranges", None)
    if ranges is None:
        return []
    return [str(item) for item in ranges]


def _intersecting_merged_ranges(
    sheet: OpenpyxlWorksheetProtocol, scope_range: str
) -> list[str]:
    """Return merged ranges that intersect the scope."""
    intersections: list[str] = []
    for merged_range in _merged_range_strings(sheet):
        if _ranges_overlap(scope_range, merged_range):
            intersections.append(merged_range)
    return intersections


def _ranges_overlap(left: str, right: str) -> bool:
    """Return True if two A1 ranges overlap."""
    left_min_col, left_min_row, left_max_col, left_max_row = _range_bounds(left)
    right_min_col, right_min_row, right_max_col, right_max_row = _range_bounds(right)
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _range_bounds(range_ref: str) -> tuple[int, int, int, int]:
    """Return range boundaries in (min_col, min_row, max_col, max_row)."""
    try:
        from openpyxl.utils.cell import range_boundaries
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    return cast(tuple[int, int, int, int], range_boundaries(range_ref))


def _build_merge_value_loss_warning(
    sheet: OpenpyxlWorksheetProtocol,
    sheet_name: str,
    range_ref: str,
) -> str | None:
    """Build warning when merge can clear non-top-left cell values."""
    coordinates = _expand_range_coordinates(range_ref)
    top_left = coordinates[0][0]
    risky_cells: list[str] = []
    for row in coordinates:
        for coord in row:
            if coord == top_left:
                continue
            value = sheet[coord].value
            if _has_non_empty_cell_value(value):
                risky_cells.append(coord)
    if not risky_cells:
        return None
    joined = ", ".join(risky_cells)
    return (
        f"merge_cells may clear non-top-left values at {sheet_name}!{range_ref}: "
        f"{joined}"
    )


def _has_non_empty_cell_value(value: str | int | float | None) -> bool:
    """Return True when cell has a non-empty value."""
    if value is None:
        return False
    if isinstance(value, str):
        return value != ""
    return True


def _normalize_hex_input(value: str, *, field_name: str) -> str:
    """Normalize HEX input into #RRGGBB or #AARRGGBB form.

    Args:
        value: Raw user input value.
        field_name: Field name used in validation messages.

    Returns:
        Normalized uppercase HEX string with '#'.

    Raises:
        ValueError: If the value is not valid HEX color text.
    """
    text = value.strip().upper()
    if not _HEX_COLOR_PATTERN.match(text):
        raise ValueError(
            f"Invalid {field_name} format. Use 'RRGGBB', 'AARRGGBB', "
            "'#RRGGBB', or '#AARRGGBB'."
        )
    return text if text.startswith("#") else f"#{text}"


def _normalize_chart_range_reference(value: str) -> str:
    """Normalize chart range reference with optional sheet qualifier."""
    candidate = value.strip()
    match = _SHEET_QUALIFIED_A1_RANGE_PATTERN.match(candidate)
    if match is None:
        raise ValueError(f"Invalid chart range reference: {value}")
    sheet_prefix = match.group("sheet") or ""
    start = match.group("start").upper()
    end = match.group("end").upper()
    return f"{sheet_prefix}{start}:{end}"


def _normalize_hex_color(value: str) -> str:
    """Normalize HEX input into AARRGGBB form for workbook internals."""
    normalized = _normalize_hex_input(value, field_name="color/fill_color")
    raw = normalized[1:]
    return raw if len(raw) == 8 else f"FF{raw}"


def _normalize_columns_for_dimensions(columns: list[str | int]) -> list[str]:
    """Normalize columns list to unique Excel-style labels."""
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in columns:
        label = (
            _column_index_to_label(raw) if isinstance(raw, int) else raw.strip().upper()
        )
        if label in seen:
            continue
        seen.add(label)
        normalized.append(label)
    return normalized


def _summarize_column_targets(columns: list[str], *, preview_limit: int = 5) -> str:
    """Return a concise summary for column target labels."""
    return _summarize_targets(columns, preview_limit=preview_limit)


def _summarize_int_targets(values: list[int], *, preview_limit: int = 5) -> str:
    """Return a concise summary for numeric target lists."""
    text_values = [str(value) for value in values]
    return _summarize_targets(text_values, preview_limit=preview_limit)


def _summarize_targets(values: list[str], *, preview_limit: int = 5) -> str:
    """Return preview text with total count for diff logs."""
    if not values:
        return "(0)"
    preview = ", ".join(values[:preview_limit])
    if len(values) > preview_limit:
        preview = f"{preview}, ..."
    return f"{preview} ({len(values)})"


def _clamp_column_width(
    width: float, *, min_width: float | None, max_width: float | None
) -> float:
    """Clamp a column width by optional lower/upper bounds."""
    clamped = width
    if min_width is not None and clamped < min_width:
        clamped = min_width
    if max_width is not None and clamped > max_width:
        clamped = max_width
    return float(clamped)


def _resolve_auto_fit_columns_openpyxl(
    sheet: OpenpyxlWorksheetProtocol,
    columns: list[str | int] | None,
) -> list[str]:
    """Resolve auto-fit target columns for openpyxl backend."""
    if columns is not None:
        return _normalize_columns_for_dimensions(columns)
    used_columns = _detect_openpyxl_used_column_indexes(sheet)
    if not used_columns:
        return ["A"]
    return [_column_index_to_label(index) for index in used_columns]


def _detect_openpyxl_used_column_indexes(
    sheet: OpenpyxlWorksheetProtocol,
) -> list[int]:
    """Detect used column indexes from non-empty openpyxl cells."""
    iter_rows = getattr(sheet, "iter_rows", None)
    if iter_rows is None:
        return [1]
    used_indexes: set[int] = set()
    for row in iter_rows():
        for cell in row:
            if _is_blank_cell_value(getattr(cell, "value", None)):
                continue
            used_index = _extract_openpyxl_cell_column_index(cell)
            if used_index is not None:
                used_indexes.add(used_index)
    if used_indexes:
        return sorted(used_indexes)
    max_column = getattr(sheet, "max_column", None)
    if isinstance(max_column, int) and max_column > 0:
        return list(range(1, max_column + 1))
    return [1]


def _collect_openpyxl_target_column_max_lengths(
    sheet: OpenpyxlWorksheetProtocol, target_indexes: set[int]
) -> dict[int, int]:
    """Collect max display lengths for target columns in a single sheet pass."""
    iter_rows = getattr(sheet, "iter_rows", None)
    if iter_rows is None:
        return {}
    max_lengths: dict[int, int] = {}
    for row in iter_rows():
        for cell in row:
            column_index = _extract_openpyxl_cell_column_index(cell)
            if column_index is None or column_index not in target_indexes:
                continue
            cell_value = getattr(cell, "value", None)
            if _is_blank_cell_value(cell_value):
                continue
            text_len = _text_display_length(cell_value)
            prev = max_lengths.get(column_index, 0)
            if text_len > prev:
                max_lengths[column_index] = text_len
    return max_lengths


def _resolve_openpyxl_estimated_width(
    column_dimension: OpenpyxlColumnDimensionProtocol, max_len: int
) -> float:
    """Resolve estimated width from max text length or current default width."""
    if max_len <= 0:
        default_width = getattr(column_dimension, "width", None)
        if isinstance(default_width, int | float) and default_width > 0:
            return float(default_width)
        return 8.43
    return float(max_len + 2)


def _extract_openpyxl_cell_column_index(cell: object) -> int | None:
    """Extract 1-based column index from an openpyxl cell-like object."""
    raw_column = getattr(cell, "column", None)
    if isinstance(raw_column, int):
        return raw_column if raw_column > 0 else None
    if isinstance(raw_column, str):
        normalized = raw_column.strip().upper()
        if not normalized:
            return None
        return _column_label_to_index(normalized)
    coordinate = str(getattr(cell, "coordinate", "")).strip()
    if not coordinate:
        return None
    if not _A1_PATTERN.match(coordinate):
        return None
    column_label, _ = _split_a1(coordinate)
    return _column_label_to_index(column_label)


def _is_blank_cell_value(value: object) -> bool:
    """Return True when the value is considered blank for width detection."""
    if value is None:
        return True
    return isinstance(value, str) and value == ""


def _text_display_length(value: object) -> int:
    """Estimate visible text length for one cell value."""
    text = str(value)
    lines = text.splitlines() or [text]
    return max(len(line) for line in lines)


def _set_grid_border(cell: OpenpyxlCellProtocol) -> None:
    """Set thin black border on all sides."""
    try:
        from openpyxl.styles import Side
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    side = Side(style="thin", color="FF000000")
    border = copy(cell.border)
    border.top = side
    border.right = side
    border.bottom = side
    border.left = side
    cell.border = border


def _snapshot_border(cell: OpenpyxlCellProtocol, coordinate: str) -> BorderSnapshot:
    """Capture border snapshot for one cell."""
    border = cell.border
    return BorderSnapshot(
        cell=coordinate,
        top=_snapshot_border_side(border.top),
        right=_snapshot_border_side(border.right),
        bottom=_snapshot_border_side(border.bottom),
        left=_snapshot_border_side(border.left),
    )


def _snapshot_border_side(side: object) -> BorderSideSnapshot:
    """Capture one border side state."""
    style = getattr(side, "style", None)
    color = _extract_openpyxl_color(getattr(side, "color", None))
    return BorderSideSnapshot(style=style, color=color)


def _snapshot_font(cell: OpenpyxlCellProtocol, coordinate: str) -> FontSnapshot:
    """Capture font snapshot for one cell."""
    font = cell.font
    return FontSnapshot(
        cell=coordinate,
        bold=getattr(font, "bold", None),
        size=getattr(font, "size", None),
        color=_extract_openpyxl_color(getattr(font, "color", None)),
    )


def _snapshot_fill(cell: OpenpyxlCellProtocol, coordinate: str) -> FillSnapshot:
    """Capture fill snapshot for one cell."""
    fill = cell.fill
    return FillSnapshot(
        cell=coordinate,
        fill_type=getattr(fill, "fill_type", None),
        start_color=_extract_openpyxl_color(getattr(fill, "start_color", None)),
        end_color=_extract_openpyxl_color(getattr(fill, "end_color", None)),
    )


def _snapshot_alignment(
    cell: OpenpyxlCellProtocol, coordinate: str
) -> AlignmentSnapshot:
    """Capture alignment snapshot for one cell."""
    alignment = cell.alignment
    return AlignmentSnapshot(
        cell=coordinate,
        horizontal=getattr(alignment, "horizontal", None),
        vertical=getattr(alignment, "vertical", None),
        wrap_text=getattr(alignment, "wrap_text", None),
    )


def _extract_openpyxl_color(color: object) -> str | None:
    """Extract RGB-like color text from openpyxl color object."""
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return None
    text = str(rgb).upper()
    return text if len(text) == 8 else None


def _build_restore_snapshot_op(sheet: str, snapshot: DesignSnapshot) -> PatchOp | None:
    """Build a restore op when snapshot contains data."""
    if (
        not snapshot.borders
        and not snapshot.fonts
        and not snapshot.fills
        and not snapshot.alignments
        and snapshot.merge_state is None
        and not snapshot.row_dimensions
        and not snapshot.column_dimensions
    ):
        return None
    return PatchOp(op="restore_design_snapshot", sheet=sheet, design_snapshot=snapshot)


def _restore_design_snapshot(
    sheet: OpenpyxlWorksheetProtocol,
    snapshot: DesignSnapshot,
) -> None:
    """Restore cell style and dimension snapshot."""
    if snapshot.merge_state is not None:
        _restore_merge_state(sheet, snapshot.merge_state)
    for border_snapshot in snapshot.borders:
        _restore_border(sheet[border_snapshot.cell], border_snapshot)
    for font_snapshot in snapshot.fonts:
        cell = sheet[font_snapshot.cell]
        font = copy(cell.font)
        font.bold = font_snapshot.bold
        font.size = font_snapshot.size
        font.color = font_snapshot.color
        cell.font = font
    for fill_snapshot in snapshot.fills:
        _restore_fill(sheet[fill_snapshot.cell], fill_snapshot)
    for alignment_snapshot in snapshot.alignments:
        _restore_alignment(sheet[alignment_snapshot.cell], alignment_snapshot)
    for row_snapshot in snapshot.row_dimensions:
        sheet.row_dimensions[row_snapshot.row].height = row_snapshot.height
    for column_snapshot in snapshot.column_dimensions:
        sheet.column_dimensions[column_snapshot.column].width = column_snapshot.width


def _restore_merge_state(
    sheet: OpenpyxlWorksheetProtocol,
    snapshot: MergeStateSnapshot,
) -> None:
    """Restore merged ranges for a scope deterministically."""
    for range_ref in _intersecting_merged_ranges(sheet, snapshot.scope):
        sheet.unmerge_cells(range_ref)
    for range_ref in snapshot.ranges:
        sheet.merge_cells(range_ref)


def _restore_border(cell: OpenpyxlCellProtocol, snapshot: BorderSnapshot) -> None:
    """Restore border from snapshot."""
    border = copy(cell.border)
    border.top = _build_side_from_snapshot(snapshot.top)
    border.right = _build_side_from_snapshot(snapshot.right)
    border.bottom = _build_side_from_snapshot(snapshot.bottom)
    border.left = _build_side_from_snapshot(snapshot.left)
    cell.border = border


def _build_side_from_snapshot(snapshot: BorderSideSnapshot) -> OpenpyxlSideProtocol:
    """Build openpyxl Side object from serializable snapshot."""
    try:
        from openpyxl.styles import Side
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    kwargs: dict[str, str] = {}
    if snapshot.style is not None:
        kwargs["style"] = snapshot.style
    if snapshot.color is not None:
        kwargs["color"] = snapshot.color
    return cast(OpenpyxlSideProtocol, Side(**kwargs))


def _restore_fill(cell: OpenpyxlCellProtocol, snapshot: FillSnapshot) -> None:
    """Restore fill from snapshot."""
    try:
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    cell.fill = PatternFill(
        fill_type=snapshot.fill_type,
        start_color=snapshot.start_color,
        end_color=snapshot.end_color,
    )


def _restore_alignment(cell: OpenpyxlCellProtocol, snapshot: AlignmentSnapshot) -> None:
    """Restore alignment from snapshot."""
    alignment = copy(cell.alignment)
    alignment.horizontal = snapshot.horizontal
    alignment.vertical = snapshot.vertical
    alignment.wrap_text = snapshot.wrap_text
    cell.alignment = alignment


def _translate_formula(formula: str, origin: str, target: str) -> str:
    """Translate formula with relative references from origin to target."""
    try:
        from openpyxl.formula.translate import Translator
    except ImportError as exc:
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc
    translated = Translator(formula, origin=origin).translate_formula(target)
    return str(translated)


def _patch_value_to_primitive(value: PatchValue | None) -> str | int | float | None:
    """Convert PatchValue into primitive value for condition checks."""
    if value is None:
        return None
    return value.value


def _values_equal_for_condition(
    current: str | int | float | None,
    expected: str | int | float | None,
) -> bool:
    """Compare values for conditional update checks."""
    return current == expected


def _build_inverse_cell_op(
    op: PatchOp,
    cell_ref: str,
    before: PatchValue | None,
) -> PatchOp | None:
    """Build inverse operation for single-cell updates."""
    if op.op not in {"set_value", "set_formula", "set_value_if", "set_formula_if"}:
        return None
    if before is None:
        return PatchOp(op="set_value", sheet=op.sheet, cell=cell_ref, value=None)
    if before.kind == "formula":
        return PatchOp(
            op="set_formula",
            sheet=op.sheet,
            cell=cell_ref,
            formula=str(before.value),
        )
    return PatchOp(op="set_value", sheet=op.sheet, cell=cell_ref, value=before.value)


def _collect_formula_issues_openpyxl(
    workbook: OpenpyxlWorkbookProtocol,
) -> list[FormulaIssue]:
    """Collect simple formula issues by scanning formula text."""
    token_map: dict[str, tuple[FormulaIssueCode, FormulaIssueLevel]] = {
        "#REF!": ("ref_error", "error"),
        "#NAME?": ("name_error", "error"),
        "#DIV/0!": ("div0_error", "error"),
        "#VALUE!": ("value_error", "error"),
        "#N/A": ("na_error", "warning"),
    }
    issues: list[FormulaIssue] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        iter_rows = getattr(sheet, "iter_rows", None)
        if iter_rows is None:
            continue
        for row in iter_rows():
            for cell in row:
                raw = getattr(cell, "value", None)
                if not isinstance(raw, str) or not raw.startswith("="):
                    continue
                normalized = raw.upper()
                if "==" in normalized:
                    issues.append(
                        FormulaIssue(
                            sheet=sheet_name,
                            cell=str(getattr(cell, "coordinate", "")),
                            level="warning",
                            code="invalid_token",
                            message="Formula contains duplicated '=' token.",
                        )
                    )
                for token, (code, level) in token_map.items():
                    if token in normalized:
                        issues.append(
                            FormulaIssue(
                                sheet=sheet_name,
                                cell=str(getattr(cell, "coordinate", "")),
                                level=level,
                                code=code,
                                message=f"Formula contains error token {token}.",
                            )
                        )
    return issues


def _apply_ops_xlwings(
    input_path: Path,
    output_path: Path,
    ops: list[PatchOp],
    auto_formula: bool,
) -> list[PatchDiffItem]:
    """Apply operations using Excel COM via xlwings."""
    diff: list[PatchDiffItem] = []
    try:
        with _xlwings_workbook(input_path) as workbook:
            sheets = {sheet.name: sheet for sheet in workbook.sheets}
            for index, op in enumerate(ops):
                try:
                    diff.append(
                        _apply_xlwings_op(workbook, sheets, op, index, auto_formula)
                    )
                except Exception as exc:
                    raise PatchOpError.from_op(index, op, exc) from exc
            workbook.save(str(output_path))
    except PatchOpError:
        raise
    except ValueError:
        raise
    except Exception as exc:
        raise RuntimeError(f"COM patch failed: {exc}") from exc
    return diff


def _apply_xlwings_op(
    workbook: XlwingsWorkbookProtocol,
    sheets: dict[str, XlwingsSheetProtocol],
    op: PatchOp,
    index: int,
    auto_formula: bool,
) -> PatchDiffItem:
    """Apply a single op to an xlwings workbook."""
    if op.op == "add_sheet":
        if op.sheet in sheets:
            raise ValueError(f"Sheet already exists: {op.sheet}")
        last = workbook.sheets[-1] if workbook.sheets else None
        sheet = workbook.sheets.add(name=op.sheet, after=last)
        sheets[op.sheet] = sheet
        return PatchDiffItem(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=None,
            before=None,
            after=PatchValue(kind="sheet", value=op.sheet),
        )

    existing_sheet = sheets.get(op.sheet)
    if existing_sheet is None:
        raise ValueError(f"Sheet not found: {op.sheet}")
    if op.op in {"set_value", "set_formula", "set_value_if", "set_formula_if"}:
        return _apply_xlwings_cell_op(existing_sheet, op, index, auto_formula)
    return _apply_xlwings_extended_op(existing_sheet, op, index)


def _apply_xlwings_extended_op(
    sheet: XlwingsSheetProtocol,
    op: PatchOp,
    index: int,
) -> PatchDiffItem:
    """Apply non-cell operations on xlwings sheets."""
    handlers: dict[PatchOpType, Callable[[], PatchDiffItem]] = {
        "set_range_values": lambda: _apply_xlwings_set_range_values(sheet, op, index),
        "fill_formula": lambda: _apply_xlwings_fill_formula(sheet, op, index),
        "draw_grid_border": lambda: _apply_xlwings_draw_grid_border(sheet, op, index),
        "set_bold": lambda: _apply_xlwings_set_bold(sheet, op, index),
        "set_font_size": lambda: _apply_xlwings_set_font_size(sheet, op, index),
        "set_font_color": lambda: _apply_xlwings_set_font_color(sheet, op, index),
        "set_fill_color": lambda: _apply_xlwings_set_fill_color(sheet, op, index),
        "set_dimensions": lambda: _apply_xlwings_set_dimensions(sheet, op, index),
        "auto_fit_columns": lambda: _apply_xlwings_auto_fit_columns(sheet, op, index),
        "merge_cells": lambda: _apply_xlwings_merge_cells(sheet, op, index),
        "unmerge_cells": lambda: _apply_xlwings_unmerge_cells(sheet, op, index),
        "set_alignment": lambda: _apply_xlwings_set_alignment(sheet, op, index),
        "set_style": lambda: _apply_xlwings_set_style(sheet, op, index),
        "apply_table_style": lambda: _apply_xlwings_apply_table_style(sheet, op, index),
        "create_chart": lambda: _apply_xlwings_create_chart(sheet, op, index),
        "restore_design_snapshot": lambda: _apply_xlwings_restore_design_snapshot(op),
    }
    handler = handlers.get(op.op)
    if handler is None:
        raise ValueError(f"Unsupported op: {op.op}")
    return handler()


def _apply_xlwings_set_range_values(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_range_values with xlwings."""
    if op.range is None or op.values is None:
        raise ValueError("set_range_values requires range and values.")
    coordinates_2d = _expand_range_coordinates(op.range)
    row_count, col_count = _shape_of_coordinates(coordinates_2d)
    if len(op.values) != row_count:
        raise ValueError("set_range_values values height does not match range.")
    if any(len(value_row) != col_count for value_row in op.values):
        raise ValueError("set_range_values values width does not match range.")
    sheet.range(op.range).value = op.values
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.range,
        before=None,
        after=PatchValue(kind="value", value=f"{row_count}x{col_count}"),
    )


def _apply_xlwings_fill_formula(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply fill_formula with xlwings."""
    if op.range is None or op.formula is None or op.base_cell is None:
        raise ValueError("fill_formula requires range, base_cell and formula.")
    coordinates_2d = _expand_range_coordinates(op.range)
    row_count, col_count = _shape_of_coordinates(coordinates_2d)
    if row_count != 1 and col_count != 1:
        raise ValueError("fill_formula range must be a single row or a single column.")
    for coord_row in coordinates_2d:
        for coord in coord_row:
            translated = _translate_formula(op.formula, op.base_cell, coord)
            sheet.range(coord).formula = translated
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.range,
        before=None,
        after=PatchValue(kind="formula", value=op.formula),
    )


def _apply_xlwings_draw_grid_border(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply draw_grid_border with xlwings."""
    if op.base_cell is None or op.row_count is None or op.col_count is None:
        raise ValueError(
            "draw_grid_border requires base_cell, row_count and col_count."
        )
    coordinates = _expand_rect_coordinates(op.base_cell, op.row_count, op.col_count)
    for coord in coordinates:
        _set_xlwings_grid_border(sheet.range(coord))
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=f"{op.base_cell}:{coordinates[-1]}",
        before=None,
        after=PatchValue(kind="style", value="grid_border(thin,black)"),
    )


def _apply_xlwings_set_bold(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_bold with xlwings."""
    target_range_ref = _xlwings_target_range_ref(op)
    target_bold = True if op.bold is None else op.bold
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    target_api.Font.Bold = target_bold
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(kind="style", value=f"bold={target_bold}"),
    )


def _apply_xlwings_set_font_size(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_font_size with xlwings."""
    if op.font_size is None:
        raise ValueError("set_font_size requires font_size.")
    target_range_ref = _xlwings_target_range_ref(op)
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    target_api.Font.Size = op.font_size
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(kind="style", value=f"font_size={op.font_size}"),
    )


def _apply_xlwings_set_font_color(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_font_color with xlwings."""
    if op.color is None:
        raise ValueError("set_font_color requires color.")
    target_range_ref = _xlwings_target_range_ref(op)
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    normalized = _normalize_hex_input(op.color, field_name="color")
    target_api.Font.Color = _hex_color_to_excel_rgb(op.color)
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(kind="style", value=f"font_color={normalized}"),
    )


def _apply_xlwings_set_fill_color(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_fill_color with xlwings."""
    if op.fill_color is None:
        raise ValueError("set_fill_color requires fill_color.")
    target_range_ref = _xlwings_target_range_ref(op)
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    target_api.Interior.Color = _hex_color_to_excel_rgb(op.fill_color)
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(
            kind="style",
            value=f"fill={_normalize_hex_input(op.fill_color, field_name='fill_color')}",
        ),
    )


def _apply_xlwings_set_dimensions(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_dimensions with xlwings."""
    parts: list[str] = []
    sheet_api = _xlwings_sheet_api(sheet)
    if op.rows is not None and op.row_height is not None:
        for row_index in op.rows:
            sheet_api.Rows(row_index).RowHeight = op.row_height
        parts.append(f"rows={_summarize_int_targets(op.rows)}")
    if op.columns is not None and op.column_width is not None:
        normalized_columns = _normalize_columns_for_dimensions(op.columns)
        for column in normalized_columns:
            sheet_api.Columns(column).ColumnWidth = op.column_width
        parts.append(f"columns={_summarize_column_targets(normalized_columns)}")
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=None,
        before=None,
        after=PatchValue(kind="dimension", value=", ".join(parts)),
    )


def _apply_xlwings_auto_fit_columns(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply auto_fit_columns with xlwings COM AutoFit."""
    sheet_api = _xlwings_sheet_api(sheet)
    target_columns = _resolve_auto_fit_columns_xlwings(sheet, op.columns)
    if not target_columns:
        raise ValueError("auto_fit_columns could not resolve target columns.")
    for column in target_columns:
        column_api = sheet_api.Columns(column)
        auto_fit = getattr(column_api, "AutoFit", None)
        if callable(auto_fit):
            auto_fit()
        current_width = getattr(column_api, "ColumnWidth", None)
        if isinstance(current_width, int | float):
            width_value = float(current_width)
        else:
            width_value = 8.43
        column_api.ColumnWidth = _clamp_column_width(
            width_value, min_width=op.min_width, max_width=op.max_width
        )
    parts = [f"columns={_summarize_column_targets(target_columns)}"]
    if op.min_width is not None:
        parts.append(f"min_width={op.min_width}")
    if op.max_width is not None:
        parts.append(f"max_width={op.max_width}")
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=None,
        before=None,
        after=PatchValue(kind="dimension", value=", ".join(parts)),
    )


def _apply_xlwings_merge_cells(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply merge_cells with xlwings."""
    if op.range is None:
        raise ValueError("merge_cells requires range.")
    _xlwings_range_api(sheet.range(op.range)).Merge()
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.range,
        before=None,
        after=PatchValue(kind="style", value=f"merged={op.range}"),
    )


def _apply_xlwings_unmerge_cells(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply unmerge_cells with xlwings."""
    if op.range is None:
        raise ValueError("unmerge_cells requires range.")
    merged_areas = _collect_xlwings_merged_areas(sheet, op.range)
    for area in merged_areas:
        _xlwings_range_api(sheet.range(area)).UnMerge()
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.range,
        before=None,
        after=PatchValue(kind="style", value=f"unmerged={len(merged_areas)}"),
    )


def _apply_xlwings_set_alignment(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_alignment with xlwings."""
    target_range_ref = _xlwings_target_range_ref(op)
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    if op.horizontal_align is not None:
        target_api.HorizontalAlignment = _XLWINGS_HORIZONTAL_ALIGN_MAP[
            op.horizontal_align
        ]
    if op.vertical_align is not None:
        target_api.VerticalAlignment = _XLWINGS_VERTICAL_ALIGN_MAP[op.vertical_align]
    if op.wrap_text is not None:
        target_api.WrapText = op.wrap_text
    summary = (
        f"horizontal={op.horizontal_align},"
        f"vertical={op.vertical_align},"
        f"wrap_text={op.wrap_text}"
    )
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(kind="style", value=summary),
    )


def _apply_xlwings_set_style(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply set_style with xlwings."""
    target_range_ref = _xlwings_target_range_ref(op)
    target_api = _xlwings_range_api(sheet.range(target_range_ref))
    if op.bold is not None:
        target_api.Font.Bold = op.bold
    if op.font_size is not None:
        target_api.Font.Size = op.font_size
    if op.color is not None:
        target_api.Font.Color = _hex_color_to_excel_rgb(op.color)
    if op.fill_color is not None:
        target_api.Interior.Color = _hex_color_to_excel_rgb(op.fill_color)
    if op.horizontal_align is not None:
        target_api.HorizontalAlignment = _XLWINGS_HORIZONTAL_ALIGN_MAP[
            op.horizontal_align
        ]
    if op.vertical_align is not None:
        target_api.VerticalAlignment = _XLWINGS_VERTICAL_ALIGN_MAP[op.vertical_align]
    if op.wrap_text is not None:
        target_api.WrapText = op.wrap_text
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=target_range_ref,
        before=None,
        after=PatchValue(
            kind="style", value=";".join(_build_set_style_summary_parts(op))
        ),
    )


def _apply_xlwings_apply_table_style(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply apply_table_style with xlwings COM API."""
    if op.range is None or op.style is None:
        raise ValueError("apply_table_style requires range and style.")
    sheet_api = _xlwings_sheet_api(sheet)
    list_objects = _resolve_xlwings_list_objects(sheet_api)
    _ensure_xlwings_table_range_not_intersects_existing_tables(list_objects, op.range)
    table_name = op.table_name or _next_xlwings_table_name(list_objects)
    _ensure_xlwings_table_name_available(list_objects, table_name)
    source_range = _resolve_chart_range_api(sheet, op.range)
    table = _xlwings_add_list_object(list_objects, source_range)
    table_any = cast(Any, table)
    table_any.Name = table_name
    _apply_xlwings_table_style(table_any, op.style)
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.range,
        before=None,
        after=PatchValue(
            kind="style", value=f"table={table_name};table_style={op.style}"
        ),
    )


def _apply_xlwings_create_chart(
    sheet: XlwingsSheetProtocol, op: PatchOp, index: int
) -> PatchDiffItem:
    """Apply create_chart with xlwings COM API."""
    if op.chart_type is None or op.data_range is None or op.anchor_cell is None:
        raise ValueError(
            "create_chart requires chart_type, data_range, and anchor_cell."
        )

    chart_type_id = _resolve_chart_type_id(op.chart_type)
    if chart_type_id is None:
        raise ValueError(
            f"create_chart chart_type must be one of: {SUPPORTED_CHART_TYPES_CSV}."
        )

    sheet_api = _xlwings_sheet_api(sheet)
    anchor_left, anchor_top = _resolve_chart_anchor(sheet, op.anchor_cell)
    chart_width = float(op.width if op.width is not None else 360.0)
    chart_height = float(op.height if op.height is not None else 220.0)
    chart_objects = _resolve_chart_objects(sheet_api)
    _validate_chart_name_uniqueness(chart_objects, op.chart_name)

    chart_object = chart_objects().Add(
        anchor_left, anchor_top, chart_width, chart_height
    )
    chart = getattr(chart_object, "Chart", None)
    if chart is None:
        raise ValueError("create_chart failed to acquire chart COM object.")

    chart.ChartType = chart_type_id
    normalized_data_ranges = _normalize_chart_data_ranges(op.data_range)
    category_range = op.category_range
    if len(normalized_data_ranges) == 1:
        chart.SetSourceData(_resolve_chart_range_api(sheet, normalized_data_ranges[0]))
    else:
        if category_range is None:
            if len(normalized_data_ranges) < 2:
                raise ValueError(
                    "create_chart data_range list requires at least two ranges when "
                    "category_range is omitted."
                )
            category_range = normalized_data_ranges[0]
            value_ranges = normalized_data_ranges[1:]
        else:
            value_ranges = normalized_data_ranges
        first_series_range = value_ranges[0]
        chart.SetSourceData(_resolve_chart_range_api(sheet, first_series_range))
        series_collection = _resolve_chart_series_collection(chart)
        first_series = _get_com_collection_item(series_collection, 1)
        cast(
            XlwingsChartSeriesProtocol, first_series
        ).Values = _resolve_chart_range_api(sheet, first_series_range)
        for series_range in value_ranges[1:]:
            series = series_collection.NewSeries()
            series.Values = _resolve_chart_range_api(sheet, series_range)
    _apply_chart_category_range(sheet, chart, category_range)
    if op.series_from_rows is not None:
        plot_by = 1 if op.series_from_rows else 2
        chart.PlotBy = plot_by
    _apply_titles_from_data_flag(chart, op.titles_from_data)
    _apply_chart_text_overrides(chart, op)
    if op.chart_name is not None:
        chart_object.Name = op.chart_name

    chart_label = op.chart_name or str(getattr(chart_object, "Name", "Chart"))
    if isinstance(op.data_range, list):
        data_summary = ",".join(op.data_range)
    else:
        data_summary = op.data_range
    chart_summary = f"type={op.chart_type};data={data_summary};anchor={op.anchor_cell};name={chart_label}"
    return PatchDiffItem(
        op_index=index,
        op=op.op,
        sheet=op.sheet,
        cell=op.anchor_cell,
        before=None,
        after=PatchValue(kind="chart", value=chart_summary),
    )


def _resolve_xlwings_list_objects(sheet_api: XlwingsSheetApiProtocol) -> object:
    """Resolve ListObjects COM collection for both property and callable forms."""
    accessor = getattr(sheet_api, "ListObjects", None)
    if accessor is None:
        raise ValueError("apply_table_style requires sheet ListObjects COM API.")
    if _looks_like_xlwings_list_objects_collection(accessor):
        return cast(object, accessor)
    if callable(accessor):
        callable_accessor = cast(Callable[..., object], accessor)
        try:
            resolved = callable_accessor()
        except Exception as exc:
            raise ValueError(
                "apply_table_style failed to access sheet ListObjects COM collection."
            ) from exc
        if _looks_like_xlwings_list_objects_collection(resolved):
            return resolved
        raise ValueError("apply_table_style requires sheet ListObjects COM API.")
    raise ValueError("apply_table_style requires sheet ListObjects COM API.")


def _looks_like_xlwings_list_objects_collection(candidate: object) -> bool:
    """Return True when object looks like Excel ListObjects collection."""
    for attr_name in ("Add", "Count"):
        try:
            getattr(candidate, attr_name)
        except Exception:
            return False
    return True


def _resolve_chart_type_id(chart_type: str) -> int | None:
    """Map chart type name to Excel COM chart type id."""
    return resolve_chart_type_id(chart_type)


def _normalize_chart_data_ranges(data_range: str | list[str]) -> list[str]:
    """Normalize create_chart data_range into a non-empty list."""
    if isinstance(data_range, str):
        return [_normalize_chart_range_reference(data_range)]
    if not data_range:
        raise ValueError("create_chart data_range list must not be empty.")
    return [_normalize_chart_range_reference(item) for item in data_range]


def _resolve_chart_anchor(
    sheet: XlwingsSheetProtocol, anchor_cell: str
) -> tuple[float, float]:
    """Return chart anchor coordinates from an A1 anchor cell."""
    anchor_api = _xlwings_range_api(sheet.range(anchor_cell))
    return float(anchor_api.Left), float(anchor_api.Top)


def _resolve_chart_objects(
    sheet_api: XlwingsSheetApiProtocol,
) -> Callable[[], XlwingsChartObjectsCollectionProtocol]:
    """Return callable ChartObjects COM accessor."""
    chart_objects = getattr(sheet_api, "ChartObjects", None)
    if not callable(chart_objects):
        raise ValueError("create_chart requires sheet ChartObjects COM API.")
    return cast(Callable[[], XlwingsChartObjectsCollectionProtocol], chart_objects)


def _resolve_chart_series_collection(
    chart: object,
) -> XlwingsChartSeriesCollectionProtocol:
    """Return series collection for a chart COM object."""
    series_collection = getattr(chart, "SeriesCollection", None)
    if not callable(series_collection):
        raise ValueError("create_chart requires chart SeriesCollection COM API.")
    return cast(XlwingsChartSeriesCollectionProtocol, series_collection())


def _existing_chart_names(
    chart_objects: Callable[[], XlwingsChartObjectsCollectionProtocol],
) -> set[str]:
    """Collect chart object names from a worksheet."""
    chart_collection = chart_objects()
    existing_count = int(getattr(chart_collection, "Count", 0))
    names: set[str] = set()
    for chart_index in range(1, existing_count + 1):
        item = _get_com_collection_item(chart_collection, chart_index)
        name_value = getattr(item, "Name", None)
        if isinstance(name_value, str):
            names.add(name_value)
    return names


def _validate_chart_name_uniqueness(
    chart_objects: Callable[[], XlwingsChartObjectsCollectionProtocol],
    chart_name: str | None,
) -> None:
    """Validate chart_name uniqueness against existing chart objects."""
    if chart_name is None:
        return
    if chart_name in _existing_chart_names(chart_objects):
        raise ValueError(f"create_chart chart_name already exists: {chart_name}")


def _apply_chart_category_range(
    sheet: XlwingsSheetProtocol, chart: object, category_range: str | None
) -> None:
    """Apply category range to all chart series when provided."""
    if category_range is None:
        return
    series_accessor = _resolve_chart_series_collection(chart)
    series_count = int(getattr(series_accessor, "Count", 0))
    category_range_api = _resolve_chart_range_api(sheet, category_range)
    for series_idx in range(1, series_count + 1):
        series_item = cast(
            XlwingsChartSeriesProtocol,
            _get_com_collection_item(series_accessor, series_idx),
        )
        series_item.XValues = category_range_api


def _apply_titles_from_data_flag(chart: object, titles_from_data: bool | None) -> None:
    """Apply titles_from_data behavior for COM chart series names."""
    if titles_from_data is not False:
        return
    series_collection = getattr(chart, "SeriesCollection", None)
    if not callable(series_collection):
        return
    series_accessor = series_collection()
    series_count = int(getattr(series_accessor, "Count", 0))
    for series_idx in range(1, series_count + 1):
        series_item = cast(
            XlwingsChartSeriesProtocol,
            _get_com_collection_item(series_accessor, series_idx),
        )
        series_item.Name = f"Series {series_idx}"


def _apply_chart_text_overrides(chart: object, op: PatchOp) -> None:
    """Apply explicit chart and axis title overrides.

    Args:
        chart: Target chart COM object.
        op: Patch operation that may include explicit title fields.

    Returns:
        None.
    """
    _set_chart_title(chart, op.chart_title)
    _set_chart_axis_title(chart, axis_type=1, text=op.x_axis_title)
    _set_chart_axis_title(chart, axis_type=2, text=op.y_axis_title)


def _set_chart_title(chart: object, title: str | None) -> None:
    """Set chart title text when provided.

    Args:
        chart: Target chart COM object.
        title: Title text. No-op when ``None``.

    Returns:
        None.
    """
    if title is None:
        return
    chart_any = cast(Any, chart)
    chart_any.HasTitle = True
    chart_title = getattr(chart_any, "ChartTitle", None)
    if chart_title is None:
        return
    cast(Any, chart_title).Text = title


def _set_chart_axis_title(chart: object, *, axis_type: int, text: str | None) -> None:
    """Set chart axis title text when provided.

    Args:
        chart: Target chart COM object.
        axis_type: Excel axis type ID (for example ``1`` for X, ``2`` for Y).
        text: Axis title text. No-op when ``None``.

    Returns:
        None.
    """
    if text is None:
        return
    axes_accessor = getattr(chart, "Axes", None)
    if not callable(axes_accessor):
        return
    try:
        axis = axes_accessor(axis_type)
    except Exception:
        return
    axis_any = cast(Any, axis)
    axis_any.HasTitle = True
    axis_title = getattr(axis, "AxisTitle", None)
    if axis_title is None:
        return
    cast(Any, axis_title).Text = text


def _resolve_chart_range_api(sheet: XlwingsSheetProtocol, range_ref: str) -> object:
    """Resolve chart source/category range API with optional sheet qualifier."""
    target_sheet_name, target_range = _split_chart_range_reference(range_ref)
    target_sheet = (
        _resolve_sheet_by_name_for_chart_range(sheet, target_sheet_name)
        if target_sheet_name is not None
        else sheet
    )
    return target_sheet.range(target_range).api


def _split_chart_range_reference(range_ref: str) -> tuple[str | None, str]:
    """Split chart range into optional sheet name and local range."""
    normalized = _normalize_chart_range_reference(range_ref)
    match = _SHEET_QUALIFIED_A1_RANGE_PATTERN.match(normalized)
    if match is None:
        raise ValueError(f"Invalid chart range reference: {range_ref}")
    sheet_prefix = match.group("sheet")
    start = match.group("start").upper()
    end = match.group("end").upper()
    local_range = f"{start}:{end}"
    if sheet_prefix is None:
        return None, local_range
    sheet_token = sheet_prefix[:-1]
    if sheet_token.startswith("'") and sheet_token.endswith("'"):
        return sheet_token[1:-1].replace("''", "'"), local_range
    return sheet_token, local_range


def _resolve_sheet_by_name_for_chart_range(
    current_sheet: XlwingsSheetProtocol, sheet_name: str
) -> XlwingsSheetProtocol:
    """Resolve target sheet by name for sheet-qualified chart ranges."""
    workbook = getattr(current_sheet, "book", None)
    if workbook is None:
        raise ValueError("create_chart requires sheet.book for sheet-qualified ranges.")
    sheets = getattr(workbook, "sheets", None)
    if sheets is None:
        raise ValueError(
            "create_chart requires workbook.sheets for sheet-qualified ranges."
        )
    try:
        return cast(XlwingsSheetProtocol, sheets[sheet_name])
    except Exception:
        try:
            for candidate in cast(list[XlwingsSheetProtocol], list(sheets)):
                if candidate.name == sheet_name:
                    return candidate
        except Exception:
            pass
    raise ValueError(f"create_chart sheet not found for range reference: {sheet_name}")


def _existing_xlwings_table_ranges(list_objects: object) -> list[tuple[str, str]]:
    """Collect existing COM table names and ranges."""
    table_count = int(getattr(list_objects, "Count", 0))
    pairs: list[tuple[str, str]] = []
    for table_index in range(1, table_count + 1):
        table = _get_com_collection_item(list_objects, table_index)
        table_name = str(getattr(table, "Name", f"Table{table_index}"))
        table_range = getattr(table, "Range", None)
        raw_address = _resolve_com_range_address(table_range)
        normalized = _normalize_table_range_address(raw_address)
        pairs.append((table_name, normalized))
    return pairs


def _resolve_com_range_address(range_api: object | None) -> str:
    """Resolve COM range address with fallback signatures."""
    if range_api is None:
        return ""
    address_method = getattr(range_api, "Address", None)
    if callable(address_method):
        address_callable = cast(Callable[..., object], address_method)
        for args in ((False, False, 1, False), (False, False), ()):
            try:
                resolved = str(address_callable(*args))
            except Exception:
                continue
            if resolved:
                return resolved
    address_value = getattr(range_api, "Address", "")
    if callable(address_value):
        try:
            return str(cast(Callable[[], object], address_value)())
        except Exception:
            return ""
    return str(address_value)


def _normalize_table_range_address(raw_address: str) -> str:
    """Normalize COM table range address for overlap checks."""
    normalized = raw_address.strip()
    if normalized.startswith("="):
        normalized = normalized[1:]
    normalized = normalized.replace("$", "")
    if "!" in normalized:
        normalized = normalized.rsplit("!", maxsplit=1)[1]
    normalized = normalized.strip().strip("'")
    range_match = _SHEET_QUALIFIED_A1_RANGE_PATTERN.match(normalized)
    if range_match is not None:
        start = range_match.group("start").upper()
        end = range_match.group("end").upper()
        return f"{start}:{end}"
    single_ref = normalized.upper()
    if _A1_PATTERN.match(single_ref):
        return single_ref
    return normalized


def _ensure_xlwings_table_range_not_intersects_existing_tables(
    list_objects: object, target_range: str
) -> None:
    """Raise when target range intersects with an existing COM table range."""
    for table_name, existing_range in _existing_xlwings_table_ranges(list_objects):
        if not existing_range:
            continue
        if _ranges_overlap(target_range, existing_range):
            raise ValueError(
                "apply_table_style range intersects existing table "
                f"'{table_name}' ({existing_range})."
            )


def _ensure_xlwings_table_name_available(list_objects: object, table_name: str) -> None:
    """Raise when table name already exists in COM tables."""
    existing_names = {name for name, _ in _existing_xlwings_table_ranges(list_objects)}
    if table_name in existing_names:
        raise ValueError(f"Table name already exists: {table_name}")


def _next_xlwings_table_name(list_objects: object) -> str:
    """Generate next available table name for COM tables."""
    existing_names = {name for name, _ in _existing_xlwings_table_ranges(list_objects)}
    for index in range(1, 10_000):
        candidate = f"Table{index}"
        if candidate not in existing_names:
            return candidate
    raise RuntimeError("Failed to generate unique table name.")


def _xlwings_add_list_object(list_objects: object, source_range_api: object) -> object:
    """Create COM ListObject with a robust Add-call fallback sequence."""
    add_method = getattr(list_objects, "Add", None)
    if not callable(add_method):
        raise ValueError("apply_table_style requires ListObjects.Add COM API.")
    add_callable = cast(Callable[..., object], add_method)
    errors: list[str] = []
    for source in _xlwings_list_object_add_sources(source_range_api):
        for attempt in _xlwings_list_object_add_attempts(source):
            try:
                return add_callable(*attempt.args, **attempt.call_kwargs)
            except Exception as exc:
                source_label = _describe_list_object_source(source)
                errors.append(f"{attempt.signature} [{source_label}] -> {exc!r}")
    tail = " | ".join(errors[-4:])
    raise ValueError(
        "apply_table_style failed to add table after COM Add signature retries. "
        f"{tail}"
    )


def _xlwings_list_object_add_sources(source_range_api: object) -> list[object]:
    """Build ListObjects.Add source variants for COM compatibility."""
    sources: list[object] = [source_range_api]
    address = _normalize_table_range_address(
        _resolve_com_range_address(source_range_api)
    )
    if address and all(
        not isinstance(item, str) or item != address for item in sources
    ):
        sources.append(address)
    return sources


def _xlwings_list_object_add_attempts(
    source: object,
) -> tuple[ListObjectAddAttempt, ...]:
    """Return Add call signatures tried for a given COM source."""
    return (
        ListObjectAddAttempt(args=(1, source), signature="Add(1, Source)"),
        ListObjectAddAttempt(
            args=(1, source, None, 1),
            signature="Add(1, Source, None, 1)",
        ),
        ListObjectAddAttempt(
            args=(1, source, None, 1, None),
            signature="Add(1, Source, None, 1, None)",
        ),
        ListObjectAddAttempt(
            args=(1, source, None, 1, None, None),
            signature="Add(1, Source, None, 1, None, None)",
        ),
        ListObjectAddAttempt(
            args=(),
            call_kwargs={"SourceType": 1, "Source": source},
            signature="Add(SourceType=1, Source=...)",
        ),
        ListObjectAddAttempt(
            args=(),
            call_kwargs={
                "SourceType": 1,
                "Source": source,
                "XlListObjectHasHeaders": 1,
            },
            signature="Add(SourceType=1, Source=..., XlListObjectHasHeaders=1)",
        ),
    )


def _describe_list_object_source(source: object) -> str:
    """Return short source label for ListObjects.Add diagnostics."""
    if isinstance(source, str):
        return f"address:{source}"
    return "range_api"


def _apply_xlwings_table_style(table: object, style_name: str) -> None:
    """Apply table style using compatible COM attributes."""
    table_any = cast(Any, table)
    style_errors: list[Exception] = []
    for attr_name in ("TableStyle", "TableStyle2"):
        if not hasattr(table_any, attr_name):
            continue
        try:
            setattr(table_any, attr_name, style_name)
            return
        except Exception as exc:
            style_errors.append(exc)
    if style_errors:
        raise ValueError(
            f"apply_table_style invalid table style: {style_name!r}. "
            f"({style_errors[-1]!r})"
        )
    raise ValueError("apply_table_style requires ListObject table style COM API.")


def _get_com_collection_item(collection: object, index: int) -> object:
    """Return indexed COM collection item with call/Item fallback."""
    last_error: Exception | None = None
    collection_call: Callable[[int], object] | None = None
    if callable(collection):
        collection_call = cast(Callable[[int], object], collection)
    try:
        if collection_call is not None:
            return collection_call(index)
    except Exception as exc:
        last_error = exc
    item_method = getattr(collection, "Item", None)
    if callable(item_method):
        item_callable = cast(Callable[[int], object], item_method)
        try:
            return item_callable(index)
        except Exception as exc:
            last_error = exc
    raise ValueError(
        f"COM collection item access failed at index {index}: {last_error!r}"
    )


def _apply_xlwings_restore_design_snapshot(op: PatchOp) -> PatchDiffItem:
    """Reject restore_design_snapshot on COM backend."""
    raise ValueError("restore_design_snapshot is supported only on openpyxl backend.")


def _apply_xlwings_cell_op(
    sheet: XlwingsSheetProtocol,
    op: PatchOp,
    index: int,
    auto_formula: bool,
) -> PatchDiffItem:
    """Apply single-cell operations on xlwings sheets."""
    cell_ref = op.cell
    if cell_ref is None:
        raise ValueError(f"{op.op} requires cell.")
    rng = sheet.range(cell_ref)
    before = _xlwings_cell_value(rng)
    if op.op == "set_value":
        after = _set_xlwings_cell_value(
            rng, op.value, auto_formula, op_name="set_value"
        )
        return _build_cell_result(op, index, cell_ref, before, after)
    if op.op == "set_formula":
        formula = _require_formula(op.formula, "set_formula")
        rng.formula = formula
        return _build_cell_result(
            op,
            index,
            cell_ref,
            before,
            PatchValue(kind="formula", value=formula),
        )
    if op.op == "set_value_if":
        if not _values_equal_for_condition(
            _patch_value_to_primitive(before), op.expected
        ):
            return _build_skipped_result(op, index, cell_ref, before)
        after = _set_xlwings_cell_value(
            rng,
            op.value,
            auto_formula,
            op_name="set_value_if",
        )
        return _build_cell_result(op, index, cell_ref, before, after)
    formula_if = _require_formula(op.formula, "set_formula_if")
    if not _values_equal_for_condition(_patch_value_to_primitive(before), op.expected):
        return _build_skipped_result(op, index, cell_ref, before)
    rng.formula = formula_if
    return _build_cell_result(
        op,
        index,
        cell_ref,
        before,
        PatchValue(kind="formula", value=formula_if),
    )


def _set_xlwings_cell_value(
    cell: XlwingsRangeProtocol,
    value: str | int | float | None,
    auto_formula: bool,
    *,
    op_name: str,
) -> PatchValue:
    """Set xlwings cell value with auto_formula handling."""
    if isinstance(value, str) and value.startswith("="):
        if not auto_formula:
            raise ValueError(f"{op_name} rejects values starting with '='.")
        cell.formula = value
        return PatchValue(kind="formula", value=value)
    cell.value = value
    return PatchValue(kind="value", value=value)


def _resolve_auto_fit_columns_xlwings(
    sheet: XlwingsSheetProtocol, columns: list[str | int] | None
) -> list[str]:
    """Resolve auto-fit target columns for xlwings backend."""
    if columns is not None:
        return _normalize_columns_for_dimensions(columns)
    used_range = getattr(sheet, "used_range", None)
    if used_range is None:
        return ["A"]
    last_cell = getattr(used_range, "last_cell", None)
    last_column = getattr(last_cell, "column", None)
    if isinstance(last_column, int) and last_column > 0:
        return [_column_index_to_label(index) for index in range(1, last_column + 1)]
    return ["A"]


def _xlwings_range_api(target: XlwingsRangeProtocol) -> XlwingsRangeApiProtocol:
    """Return COM range API object from xlwings wrapper."""
    return cast(XlwingsRangeApiProtocol, target.api)


def _xlwings_sheet_api(target: XlwingsSheetProtocol) -> XlwingsSheetApiProtocol:
    """Return COM sheet API object from xlwings wrapper."""
    return cast(XlwingsSheetApiProtocol, target.api)


def _xlwings_target_range_ref(op: PatchOp) -> str:
    """Return target range reference from a style operation payload."""
    if op.cell is not None:
        return op.cell
    if op.range is not None:
        return op.range
    raise ValueError(f"{op.op} requires cell or range.")


def _set_xlwings_grid_border(cell: XlwingsRangeProtocol) -> None:
    """Set thin black border on all four sides via Excel COM."""
    cell_api = _xlwings_range_api(cell)
    for edge in (7, 8, 9, 10):
        border = cell_api.Borders(edge)
        border.LineStyle = 1
        border.Color = 0


def _hex_color_to_excel_rgb(fill_color: str) -> int:
    """Convert hex color to Excel COM RGB integer."""
    argb = _normalize_hex_color(fill_color)
    rgb = argb[2:]
    red = int(rgb[0:2], 16)
    green = int(rgb[2:4], 16)
    blue = int(rgb[4:6], 16)
    return red + green * 256 + blue * 65_536


def _collect_xlwings_merged_areas(
    sheet: XlwingsSheetProtocol,
    target_range: str,
) -> list[str]:
    """Collect unique merged range addresses intersecting target range."""
    merged_areas: set[str] = set()
    for coord_row in _expand_range_coordinates(target_range):
        for coord in coord_row:
            cell_api = _xlwings_range_api(sheet.range(coord))
            if not bool(cell_api.MergeCells):
                continue
            merge_area = cell_api.MergeArea
            raw_address = str(merge_area.Address(False, False))
            merged_areas.add(raw_address.replace("$", ""))
    return sorted(merged_areas)


def _xlwings_cell_value(cell: XlwingsRangeProtocol) -> PatchValue | None:
    """Normalize an xlwings cell value into PatchValue."""
    formula = getattr(cell, "formula", None)
    if isinstance(formula, str) and formula.startswith("="):
        return PatchValue(kind="formula", value=formula)
    value = getattr(cell, "value", None)
    if value is None:
        return None
    return PatchValue(kind="value", value=value)


def _close_workbook_safely(workbook: XlwingsWorkbookProtocol) -> None:
    """Close workbook and ignore cleanup failures."""
    try:
        workbook.close()
    except Exception:
        return


def _quit_app_safely(app: XlwingsAppProtocol) -> None:
    """Quit xlwings app and fallback to force-kill on failure."""
    try:
        app.quit()
    except Exception:
        try:
            app.kill()
        except Exception:
            return


@contextmanager
def _xlwings_workbook(file_path: Path) -> Iterator[XlwingsWorkbookProtocol]:
    """Open an Excel workbook with a dedicated COM app."""
    app = xw.App(add_book=False, visible=False)
    app.display_alerts = False
    app.screen_updating = False
    workbook = app.books.open(str(file_path))
    try:
        yield workbook
    finally:
        _close_workbook_safely(workbook)
        _quit_app_safely(app)


class PatchOpError(ValueError):
    """Patch operation error with structured detail."""

    def __init__(self, detail: PatchErrorDetail) -> None:
        super().__init__(detail.message)
        self.detail = detail

    @classmethod
    def from_op(cls, index: int, op: PatchOp, exc: Exception) -> PatchOpError:
        """Build a PatchOpError from an op and exception."""
        message = str(exc)
        hint, expected_fields, example_op = _build_patch_error_guidance(op, message)
        error_code, failed_field, raw_com_message = _classify_patch_error(
            op, message, exc
        )
        detail = PatchErrorDetail(
            op_index=index,
            op=op.op,
            sheet=op.sheet,
            cell=op.cell,
            message=message,
            hint=hint,
            expected_fields=expected_fields,
            example_op=example_op,
            error_code=error_code,
            failed_field=failed_field,
            raw_com_message=raw_com_message,
        )
        return cls(detail)


def _build_patch_error_guidance(
    op: PatchOp, message: str
) -> tuple[str | None, list[str], str | None]:
    """Build structured guidance for common operation mistakes."""
    lowered_message = message.lower()
    if op.op == "set_fill_color" and (
        "does not accept color" in message or "requires fill_color" in message
    ):
        return (
            "set_fill_color では 'color' ではなく 'fill_color' を指定してください。",
            ["op", "sheet", "cell or range", "fill_color"],
            (
                '{"op":"set_fill_color","sheet":"Sheet1",'
                '"cell":"A1","fill_color":"#FFD966"}'
            ),
        )
    if op.op == "set_alignment" and "requires at least one of" in message:
        return (
            "set_alignment は horizontal_align / vertical_align / wrap_text の"
            " いずれかが必須です。alias の 'horizontal' / 'vertical' も利用できます。",
            [
                "op",
                "sheet",
                "cell or range",
                "horizontal_align/vertical_align/wrap_text",
            ],
            (
                '{"op":"set_alignment","sheet":"Sheet1","range":"A1:B1",'
                '"horizontal_align":"center"}'
            ),
        )
    if op.op == "set_style" and "requires at least one style field" in message:
        return (
            "set_style では style 属性を少なくとも1つ指定してください。",
            [
                "op",
                "sheet",
                "cell or range",
                "bold/font_size/color/fill_color/horizontal_align/vertical_align/wrap_text",
            ],
            (
                '{"op":"set_style","sheet":"Sheet1","range":"A1:B1",'
                '"bold":true,"fill_color":"#D9E1F2","horizontal_align":"center"}'
            ),
        )
    if op.op == "create_chart" and "Invalid chart range reference" in message:
        return (
            "create_chart の data_range/category_range は A1 範囲または "
            "'Sheet Name'!A1:B10 形式で指定してください。",
            ["op", "sheet", "chart_type", "data_range", "anchor_cell"],
            (
                '{"op":"create_chart","sheet":"Sheet1","chart_type":"line",'
                '"data_range":["Sheet1!B2:B13","Sheet1!C2:C13"],"anchor_cell":"F2"}'
            ),
        )
    if op.op == "create_chart" and "sheet not found" in message.lower():
        return (
            "指定したシート名が存在しません。シート名の大文字小文字・スペース・"
            "引用符（'Sheet Name'）を確認してください。",
            ["data_range/category_range"],
            (
                '{"op":"create_chart","sheet":"Sheet1","chart_type":"line",'
                '"data_range":"\'Sales 2026\'!B2:C13","anchor_cell":"F2"}'
            ),
        )
    if op.op == "apply_table_style" and "invalid table style" in lowered_message:
        return (
            "style には Excel の有効なテーブルスタイル名を指定してください。"
            " 例: TableStyleMedium2 / TableStyleLight9。",
            ["op", "sheet", "range", "style"],
            (
                '{"op":"apply_table_style","sheet":"Sheet1",'
                '"range":"A1:D11","style":"TableStyleMedium2"}'
            ),
        )
    if op.op == "apply_table_style" and "failed to add table" in lowered_message:
        return (
            "range にはヘッダー行を含む連続した A1 範囲を指定してください。"
            " 既存テーブルとの重複や無効な参照があると失敗します。",
            ["op", "sheet", "range", "style"],
            (
                '{"op":"apply_table_style","sheet":"Sheet1",'
                '"range":"A1:D11","style":"TableStyleMedium2"}'
            ),
        )
    return None, [], None


def _classify_patch_error(
    op: PatchOp, message: str, exc: Exception
) -> tuple[str, str | None, str | None]:
    """Classify operation error into a structured code and likely field."""
    lowered = message.lower()
    raw_com_message = _extract_raw_com_message(exc)
    classified = _classify_known_patch_error(lowered)
    if classified is not None:
        error_code, failed_field = classified
        return error_code, failed_field, raw_com_message
    if raw_com_message is not None:
        return "com_runtime_error", None, raw_com_message
    return "operation_failed", None, raw_com_message


def _classify_known_patch_error(
    lowered_message: str,
) -> tuple[str, str | None] | None:
    """Classify non-COM patch errors using deterministic string patterns."""
    if "invalid chart range reference" in lowered_message:
        detected_field = (
            "category_range" if "category" in lowered_message else "data_range"
        )
        return "invalid_range", detected_field
    if "sheet not found" in lowered_message:
        if "category_range" in lowered_message or "category range" in lowered_message:
            return "sheet_not_found", "category_range"
        if "data_range" in lowered_message or "data range" in lowered_message:
            return "sheet_not_found", "data_range"
        return "sheet_not_found", None
    matchers: tuple[tuple[str, str, str | None], ...] = (
        ("chart_type must be one of", "chart_type_invalid", "chart_type"),
        ("chart_name already exists", "chart_name_conflict", "chart_name"),
        ("table name already exists", "table_name_conflict", "table_name"),
        ("intersects existing table", "table_range_intersection", "range"),
        ("invalid table style", "table_style_invalid", "style"),
        ("failed to add table", "list_object_add_failed", "range"),
        ("requires listobjects.add com api", "com_api_missing", "range"),
        (
            "failed to access sheet listobjects com collection",
            "com_api_missing",
            "range",
        ),
        ("requires listobject table style com api", "com_api_missing", "style"),
        ("requires range and style", "invalid_parameter", "range/style"),
        ("requires chart_type", "invalid_parameter", "chart_type"),
        ("requires data_range", "invalid_parameter", "data_range"),
        ("requires anchor_cell", "invalid_parameter", "anchor_cell"),
    )
    for needle, error_code, failed_field in matchers:
        if needle in lowered_message:
            return error_code, failed_field
    return None


def _extract_raw_com_message(exc: Exception) -> str | None:
    """Extract raw COM exception text when applicable."""
    class_name = exc.__class__.__name__.lower()
    message = str(exc)
    if "com_error" in class_name:
        return message
    if "hresult" in message.lower() or "-2147" in message:
        return message
    return None
