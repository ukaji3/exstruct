from __future__ import annotations

import logging
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

from exstruct import ExtractionMode, process_excel

from .io import PathPolicy
from .shared.output_path import (
    apply_conflict_policy as _shared_apply_conflict_policy,
    next_available_path as _shared_next_available_path,
    resolve_output_path as _shared_resolve_output_path,
)

logger = logging.getLogger(__name__)

OnConflictPolicy = Literal["overwrite", "skip", "rename"]


class WorkbookMeta(BaseModel):
    """Lightweight workbook metadata for MCP responses."""

    sheet_names: list[str] = Field(default_factory=list, description="Sheet names.")
    sheet_count: int = Field(default=0, description="Total number of sheets.")


class ExtractOptions(BaseModel):
    """Optional extraction configuration for MCP requests."""

    pretty: bool | None = Field(default=None, description="Pretty-print JSON output.")
    indent: int | None = Field(
        default=None, description="Indent width for JSON output."
    )
    sheets_dir: Path | None = Field(
        default=None, description="Directory for per-sheet outputs."
    )
    print_areas_dir: Path | None = Field(
        default=None, description="Directory for per-print-area outputs."
    )
    auto_page_breaks_dir: Path | None = Field(
        default=None, description="Directory for auto page-break outputs."
    )
    alpha_col: bool = Field(
        default=True,
        description="When true, convert CellRow column keys to Excel-style ABC names (A, B, ..., Z, AA, ...) instead of 0-based indices. MCP default is true.",
    )


class ExtractRequest(BaseModel):
    """Input model for ExStruct MCP extraction."""

    xlsx_path: Path
    mode: ExtractionMode = "standard"
    format: Literal["json", "yaml", "yml", "toon"] = "json"  # noqa: A003
    out_dir: Path | None = None
    out_name: str | None = None
    on_conflict: OnConflictPolicy = "overwrite"
    options: ExtractOptions = Field(default_factory=ExtractOptions)


class ExtractResult(BaseModel):
    """Output model for ExStruct MCP extraction."""

    out_path: str
    workbook_meta: WorkbookMeta | None = None
    warnings: list[str] = Field(default_factory=list)
    engine: Literal["internal_api", "cli_subprocess"] = "internal_api"


def run_extract(
    request: ExtractRequest, *, policy: PathPolicy | None = None
) -> ExtractResult:
    """Run an extraction with file output.

    Args:
        request: Extraction request payload.
        policy: Optional path policy for access control.

    Returns:
        Extraction result with output path and metadata.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If paths violate the policy.
    """
    resolved_input = _resolve_input_path(request.xlsx_path, policy=policy)
    output_path = _resolve_output_path(
        resolved_input,
        request.format,
        out_dir=request.out_dir,
        out_name=request.out_name,
        policy=policy,
    )
    output_path, warning, skipped = _apply_conflict_policy(
        output_path, request.on_conflict
    )
    warnings: list[str] = []
    if warning:
        warnings.append(warning)
    if skipped:
        return ExtractResult(
            out_path=str(output_path),
            workbook_meta=None,
            warnings=warnings,
            engine="internal_api",
        )
    _ensure_output_dir(output_path)

    options = request.options
    sheets_dir = _resolve_optional_dir(options.sheets_dir, policy=policy)
    print_areas_dir = _resolve_optional_dir(options.print_areas_dir, policy=policy)
    auto_page_breaks_dir = _resolve_optional_dir(
        options.auto_page_breaks_dir, policy=policy
    )
    pretty = options.pretty if options.pretty is not None else False

    process_excel(
        file_path=resolved_input,
        output_path=output_path,
        out_fmt=request.format,
        mode=request.mode,
        pretty=pretty,
        indent=options.indent,
        sheets_dir=sheets_dir,
        print_areas_dir=print_areas_dir,
        auto_page_breaks_dir=auto_page_breaks_dir,
        alpha_col=options.alpha_col,
    )
    meta, meta_warnings = _try_read_workbook_meta(resolved_input)
    warnings.extend(meta_warnings)
    return ExtractResult(
        out_path=str(output_path),
        workbook_meta=meta,
        warnings=warnings,
        engine="internal_api",
    )


def _resolve_input_path(path: Path, *, policy: PathPolicy | None) -> Path:
    """Resolve and validate the input path.

    Args:
        path: Candidate input path.
        policy: Optional path policy.

    Returns:
        Resolved input path.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the path violates the policy or is not a file.
    """
    resolved = policy.ensure_allowed(path) if policy else path.resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Input file not found: {resolved}")
    if not resolved.is_file():
        raise ValueError(f"Input path is not a file: {resolved}")
    return resolved


def _resolve_output_path(
    input_path: Path,
    fmt: Literal["json", "yaml", "yml", "toon"],
    *,
    out_dir: Path | None,
    out_name: str | None,
    policy: PathPolicy | None,
) -> Path:
    """Build and validate the output path.

    Args:
        input_path: Resolved input path.
        fmt: Output format.
        out_dir: Optional output directory.
        out_name: Optional output filename.
        policy: Optional path policy.

    Returns:
        Resolved output path.

    Raises:
        ValueError: If the path violates the policy.
    """
    return _shared_resolve_output_path(
        input_path,
        out_dir=out_dir,
        out_name=out_name,
        policy=policy,
        default_suffix=_format_suffix(fmt),
        default_name_builder="same_stem",
    )


def _normalize_output_name(input_path: Path, out_name: str | None, suffix: str) -> str:
    """Normalize output filename with a suffix.

    Args:
        input_path: Input file path.
        out_name: Optional output filename override.
        suffix: Format-specific suffix.

    Returns:
        Output filename with suffix.
    """
    if out_name:
        candidate = Path(out_name)
        return candidate.name if candidate.suffix else f"{candidate.name}{suffix}"
    return f"{input_path.stem}{suffix}"


def _ensure_output_dir(path: Path) -> None:
    """Ensure the output directory exists.

    Args:
        path: Output file path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)


def _resolve_optional_dir(
    path: Path | None, *, policy: PathPolicy | None
) -> Path | None:
    """Resolve an optional output directory with policy enforcement.

    Args:
        path: Optional directory path.
        policy: Optional path policy.

    Returns:
        Resolved path or None.

    Raises:
        ValueError: If the path violates the policy.
    """
    if path is None:
        return None
    return policy.ensure_allowed(path) if policy else path.resolve()


def _format_suffix(fmt: Literal["json", "yaml", "yml", "toon"]) -> str:
    """Return suffix for output format.

    Args:
        fmt: Output format.

    Returns:
        File suffix for the format.
    """
    return ".yml" if fmt == "yml" else f".{fmt}"


def _apply_conflict_policy(
    output_path: Path, on_conflict: OnConflictPolicy
) -> tuple[Path, str | None, bool]:
    """Apply output conflict policy to a resolved output path.

    Args:
        output_path: Target output file path.
        on_conflict: Conflict handling policy.

    Returns:
        Tuple of (resolved output path, warning message or None, skipped flag).
    """
    return _shared_apply_conflict_policy(output_path, on_conflict)


def _next_available_path(path: Path) -> Path:
    """Return the next available path by appending a numeric suffix."""
    return _shared_next_available_path(path)


def _try_read_workbook_meta(path: Path) -> tuple[WorkbookMeta | None, list[str]]:
    """Try reading lightweight workbook metadata.

    Args:
        path: Excel workbook path.

    Returns:
        Tuple of metadata (or None) and warnings.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return None, [f"openpyxl is not available: {exc}"]

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - surface as warning
        return None, [f"Failed to read workbook metadata: {exc}"]

    sheet_names = list(workbook.sheetnames)
    workbook.close()
    return WorkbookMeta(sheet_names=sheet_names, sheet_count=len(sheet_names)), []
