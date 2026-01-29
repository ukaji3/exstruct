from __future__ import annotations

from pathlib import Path

import openpyxl

from .common import write_text


def extract_openpyxl(
    xlsx_path: Path, out_txt: Path, sheet_scope: list[str] | None = None
) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        sheets = sheet_scope or wb.sheetnames

        lines: list[str] = []
        lines.append("[DOC_META]")
        lines.append(f"source={xlsx_path.name}")
        lines.append("method=openpyxl")
        lines.append("")
        lines.append("[CONTENT]")

        for sname in sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            lines.append(f"\n# SHEET: {sname}")
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1

            for r in range(1, max_row + 1):
                row_cells = []
                for c in range(1, max_col + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    txt = str(v).strip()
                    if not txt:
                        continue
                    # 座標付きで記録（後で人間が確認しやすい）
                    row_cells.append(f"R{r}C{c}:{txt}")
                if row_cells:
                    lines.append(" | ".join(row_cells))

        write_text(out_txt, "\n".join(lines).strip() + "\n")
    finally:
        wb.close()
